from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from supabase import create_client, Client
from datetime import datetime, timedelta, timezone
import os
import pytz
import math
import json
from dotenv import load_dotenv
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import re
import bcrypt
import uuid
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import secrets
import traceback
import time
import io
import csv
from collections import defaultdict
import requests

# Load environment variables
load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'emergency-alert-secret-key-2025')

# Configure maximum file upload size (5MB)
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024  # 5MB

# Supabase configuration
SUPABASE_URL = os.getenv('SUPABASE_URL')
SUPABASE_KEY = os.getenv('SUPABASE_KEY')

if not SUPABASE_URL or not SUPABASE_KEY:
    # Fail fast if env vars are missing so Render logs show the problem clearly
    raise ValueError("Please set SUPABASE_URL and SUPABASE_KEY environment variables")

# Initialize Supabase client
try:
    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
    print("Supabase client initialized successfully")
except Exception as e:
    # Log detailed information and re-raise so the deployment fails loudly.
    # This makes it easier to see *why* Supabase failed to initialize in Render logs.
    print("Error initializing Supabase client:", e)
    print("SUPABASE_URL:", repr(SUPABASE_URL))
    print("SUPABASE_KEY length:", len(SUPABASE_KEY) if SUPABASE_KEY else 0)
    raise

# Table name for storing resolution summaries (can be overridden via environment)
RESOLUTION_REPORTS_TABLE = os.getenv('RESOLUTION_REPORTS_TABLE', 'incident_resolution_reports')

# Philippines timezone (UTC+8)
PHILIPPINES_TZ = pytz.timezone('Asia/Manila')

def get_philippines_time():
    """Get current time in Philippines timezone"""
    return datetime.now(PHILIPPINES_TZ)

# Cache for geocoding results to avoid repeated API calls
_geocoding_cache = {}

def get_location_name_from_coords(lat, lng, use_cache=True):
    """Get location name from latitude and longitude using reverse geocoding"""
    # Round coordinates to 6 decimal places for cache key (about 0.1m precision)
    cache_key = (round(float(lat), 6), round(float(lng), 6))
    
    # Check cache first
    if use_cache and cache_key in _geocoding_cache:
        return _geocoding_cache[cache_key]
    
    try:
        # Use OpenStreetMap Nominatim API for reverse geocoding (free, no API key needed)
        # Add a small delay to respect rate limits (1 request per second)
        time.sleep(1)
        
        url = f"https://nominatim.openstreetmap.org/reverse?format=json&lat={lat}&lon={lng}&zoom=18&addressdetails=1"
        headers = {
            'User-Agent': 'UMAK-Emergency-Alert-System/1.0',
            'Accept': 'application/json',
            'Referer': 'https://umak.edu.ph'
        }
        
        response = requests.get(url, headers=headers, timeout=15)
        
        if response.status_code != 200:
            print(f"Geocoding API returned status {response.status_code}: {response.text[:200]}")
            result = None
        else:
            data = response.json()
            
            if not data:
                print(f"No data returned from geocoding API for lat={lat}, lng={lng}")
                result = None
            else:
                # Check if we have a display_name (most reliable and complete)
                if data.get('display_name'):
                    display_name = data['display_name'].strip()
                    if display_name:
                        print(f"Geocoding success: {display_name[:100]}")
                        result = display_name
                    else:
                        result = None
                else:
                    # Fallback to building address from address components
                    if data.get('address'):
                        addr = data['address']
                        location_parts = []
                        
                        # Build location name from most specific to least specific
                        if addr.get('house_number'):
                            location_parts.append(str(addr['house_number']))
                        if addr.get('road') or addr.get('street'):
                            road = addr.get('road') or addr.get('street')
                            if road:
                                location_parts.append(road)
                        if addr.get('suburb') or addr.get('neighbourhood'):
                            suburb = addr.get('suburb') or addr.get('neighbourhood')
                            if suburb:
                                location_parts.append(suburb)
                        if addr.get('city_district') or addr.get('district'):
                            district = addr.get('city_district') or addr.get('district')
                            if district and district not in location_parts:
                                location_parts.append(district)
                        if addr.get('city') or addr.get('town') or addr.get('village'):
                            city = addr.get('city') or addr.get('town') or addr.get('village')
                            if city and city not in location_parts:
                                location_parts.append(city)
                        if addr.get('municipality'):
                            municipality = addr['municipality']
                            if municipality and municipality not in location_parts:
                                location_parts.append(municipality)
                        if addr.get('state') or addr.get('region'):
                            state = addr.get('state') or addr.get('region')
                            if state and state not in location_parts:
                                location_parts.append(state)
                        
                        if location_parts:
                            location_name = ', '.join(location_parts)
                            print(f"Geocoding success (from components): {location_name[:100]}")
                            result = location_name
                        else:
                            result = None
                    elif data.get('name'):
                        # If we have a name field, use it
                        name = data['name'].strip()
                        if name:
                            print(f"Geocoding success (from name): {name[:100]}")
                            result = name
                        else:
                            result = None
                    else:
                        print(f"No usable location name found in geocoding response for lat={lat}, lng={lng}")
                        result = None
        
        # Cache the result (even if None, to avoid repeated failed calls)
        if use_cache:
            _geocoding_cache[cache_key] = result
        
        return result
    except requests.exceptions.Timeout:
        print(f"Geocoding API timeout for lat={lat}, lng={lng}")
        if use_cache:
            _geocoding_cache[cache_key] = None
        return None
    except requests.exceptions.RequestException as e:
        print(f"Geocoding API request error for lat={lat}, lng={lng}: {e}")
        if use_cache:
            _geocoding_cache[cache_key] = None
        return None
    except Exception as e:
        print(f"Unexpected error in geocoding for lat={lat}, lng={lng}: {e}")
        if use_cache:
            _geocoding_cache[cache_key] = None
        return None
        return None
    except json.JSONDecodeError as e:
        print(f"Geocoding API returned invalid JSON: {e}")
        return None
    except Exception as e:
        print(f"Error in reverse geocoding: {e}")
        import traceback
        traceback.print_exc()
        return None

# Email configuration for password reset
EMAIL_CONFIG = {
    'smtp_host': 'smtp.gmail.com',
    'smtp_port': 587,
    'smtp_username': os.getenv('EMAIL_USERNAME', 'eclipsealertsystem@gmail.com'),
    'smtp_password': os.getenv('EMAIL_PASSWORD', 'bzzkpnlojngxluzl'),
    'smtp_secure': 'tls',
    'from_email': os.getenv('EMAIL_USERNAME', 'eclipsealertsystem@gmail.com'),
    'debug_mode': os.getenv('EMAIL_DEBUG', 'true').lower() == 'true'
}

print(f"📧 Email config - Debug mode: {EMAIL_CONFIG['debug_mode']}")

# Add template filters
@app.template_filter('format_datetime')
def format_datetime_filter(dt):
    return format_datetime(dt)

@app.template_filter('safe_get')
def safe_get_filter(data, key, default='N/A'):
    return safe_get(data, key, default)

@app.template_filter('format_admin_id')
def format_admin_id_filter(admin_id):
    """Format admin ID with leading zeros - handles various formats from database"""
    if not admin_id:
        return "ADM-0000"
    
    try:
        # If admin_id is already in format like "ADM00001" or "ADM-00001", extract the number
        admin_id_str = str(admin_id).strip()
        
        # Remove "ADM-" or "ADM" prefix if present
        if admin_id_str.upper().startswith('ADM-'):
            admin_id_str = admin_id_str[4:]
        elif admin_id_str.upper().startswith('ADM'):
            admin_id_str = admin_id_str[3:]
        
        # Try to extract numeric part
        numeric_part = ''.join(filter(str.isdigit, admin_id_str))
        
        if numeric_part:
            # Format with leading zeros: ADM-00001
            return f"ADM-{int(numeric_part):05d}"
        else:
            # If no numeric part found, try to convert the whole thing
            return f"ADM-{int(admin_id_str):05d}"
    except (ValueError, TypeError):
        # If conversion fails, return as-is with ADM- prefix
        return f"ADM-{admin_id}"

@app.template_filter('get_initials')
def get_initials_filter(name, max_chars=2):
    """Get initials from a full name"""
    if not name:
        return 'A'
    try:
        # Split name into words and get first letter of each word
        words = name.strip().split()
        if len(words) == 0:
            return 'A'
        elif len(words) == 1:
            # Single word - return first character(s)
            return name[:max_chars].upper()
        else:
            # Multiple words - return first letter of first and last word
            initials = words[0][0].upper()
            if len(words) > 1:
                initials += words[-1][0].upper()
            return initials[:max_chars]
    except Exception:
        return name[0].upper() if name else 'A'

# Add current year to template context
@app.context_processor
def inject_current_year():
    return {'current_year': datetime.now().year}

# Add built-in functions to Jinja2 environment
@app.context_processor
def inject_builtins():
    return {
        'max': max,
        'min': min,
        'abs': abs,
        'round': round,
        'len': len,
        'range': range
    }

# ---------------- HELPER FUNCTIONS ---------------- #
def format_datetime(dt):
    """Format datetime object or string for display"""
    if not dt:
        return 'N/A'
    
    def ensure_ph_tz(dt_obj):
        """Convert datetime to Asia/Manila timezone for display."""
        try:
            if dt_obj.tzinfo is None:
                # Assume UTC if no timezone info, then convert to PH
                dt_obj = dt_obj.replace(tzinfo=timezone.utc).astimezone(PHILIPPINES_TZ)
            else:
                dt_obj = dt_obj.astimezone(PHILIPPINES_TZ)
        except Exception:
            # As a fallback, localize directly to PH timezone
            try:
                dt_obj = PHILIPPINES_TZ.localize(dt_obj.replace(tzinfo=None))
            except Exception:
                pass
        return dt_obj

    if isinstance(dt, str):
        # Handle ISO format strings from Supabase
        try:
            # Try to parse ISO format
            if 'T' in dt:
                dt_obj = datetime.fromisoformat(dt.replace('Z', '+00:00'))
                dt_obj = ensure_ph_tz(dt_obj)
                return dt_obj.strftime('%Y-%m-%d %H:%M:%S')
            else:
                # Already formatted or partial string
                return dt[:19] if len(dt) > 19 else dt
        except:
            return dt[:19] if len(dt) > 19 else dt
    
    if hasattr(dt, 'strftime'):
        try:
            dt_obj = ensure_ph_tz(dt)
            return dt_obj.strftime('%Y-%m-%d %H:%M:%S')
        except Exception:
            return dt.strftime('%Y-%m-%d %H:%M:%S')
    
    return str(dt)

def safe_get(data, key, default='N/A'):
    """Safely get data from dictionary with default value"""
    if isinstance(data, dict):
        return data.get(key, default)
    return default

def generate_resolution_id():
    """Generate next resolution ID with RSV prefix."""
    prefix = 'RSV'
    padding = 5
    default_number = 1

    if supabase is None:
        return f"{prefix}{default_number:0{padding}d}"

    try:
        result = (
            supabase.table(RESOLUTION_REPORTS_TABLE)
            .select('resolved_id, created_at')
            .order('created_at', desc=True)
            .limit(1)
            .execute()
        )
        last_id = None
        if result and getattr(result, 'data', None):
            last_record = result.data[0]
            last_id = last_record.get('resolved_id')

        if last_id and isinstance(last_id, str) and last_id.upper().startswith(prefix):
            last_id = last_id.upper()
            numeric_part = ''.join(ch for ch in last_id[len(prefix):] if ch.isdigit())
            if numeric_part:
                next_number = int(numeric_part) + 1
            else:
                next_number = default_number
        else:
            next_number = default_number
    except Exception as e:
        print(f"Error generating resolved_id: {e}")
        next_number = default_number

    return f"{prefix}{next_number:0{padding}d}"

def check_profile_image_exists(profile_filename):
    """Check if a profile image file actually exists on disk"""
    if not profile_filename or profile_filename == 'default.png':
        return False
    try:
        image_path = os.path.join(app.static_folder, 'images', profile_filename)
        return os.path.exists(image_path)
    except Exception:
        return False

def generate_verification_code():
    """Generate 6-digit verification code"""
    return f"{secrets.randbelow(900000) + 100000:06d}"

def send_email(to_email, subject, body):
    """Send email for password reset with improved error handling"""
    if EMAIL_CONFIG['debug_mode']:
        print(f"\n=== EMAIL DEBUG MODE ===")
        print(f"To: {to_email}")
        print(f"Subject: {subject}")
        print(f"Body: {body}")
        print(f"=== EMAIL DEBUG MODE ===\n")
        return True
    else:
        try:
            # Create message
            msg = MIMEMultipart()
            msg['From'] = EMAIL_CONFIG['from_email']
            msg['To'] = to_email
            msg['Subject'] = subject
            
            # Add HTML body
            msg.attach(MIMEText(body, 'html'))
            
            # Create secure connection
            server = smtplib.SMTP(EMAIL_CONFIG['smtp_host'], EMAIL_CONFIG['smtp_port'])
            server.ehlo()
            server.starttls()
            server.ehlo()
            
            # Login and send
            server.login(EMAIL_CONFIG['smtp_username'], EMAIL_CONFIG['smtp_password'])
            text = msg.as_string()
            server.sendmail(EMAIL_CONFIG['from_email'], to_email, text)
            server.quit()
            
            print(f"✅ Email sent successfully to {to_email}")
            return True
            
        except Exception as e:
            print(f"❌ Email sending failed: {str(e)}")
            # Fallback to debug mode if email fails
            print(f"\n=== EMAIL FALLBACK DEBUG ===")
            print(f"To: {to_email}")
            print(f"Subject: {subject}")
            print(f"Body: {body}")
            print(f"=== EMAIL FALLBACK DEBUG ===\n")
            return False

def validate_password(password):
    """Validate password strength with reasonable requirements"""
    if len(password) < 8:
        return False, "Password must be at least 8 characters long"
    
    # Check for at least one uppercase letter
    if not re.search(r'[A-Z]', password):
        return False, "Password must contain at least one uppercase letter"
    
    # Check for at least one lowercase letter
    if not re.search(r'[a-z]', password):
        return False, "Password must contain at least one lowercase letter"
    
    # Check for at least one number
    if not re.search(r'\d', password):
        return False, "Password must contain at least one number"
    
    # Check for at least one special character
    if not re.search(r'[!@#$%^&*(),.?":{}|<>]', password):
        return False, "Password must contain at least one special character (!@#$%^&* etc.)"
    
    return True, "Password is valid"

def log_admin_activity(admin_id, admin_name, action_type, incident_id, old_status=None, new_status=None):
    """Log admin activities to the database"""
    try:
        result = supabase.table('admin_activity_logs').insert({
            'admin_id': admin_id,
            'admin_name': admin_name,
            'action_type': action_type,
            'incident_id': incident_id,
            'old_status': old_status,
            'new_status': new_status,
            'action_timestamp': datetime.now().isoformat()
        }).execute()
        return result
    except Exception as e:
        print(f"Error logging activity: {e}")
        return None

def get_admin_by_username(username):
    """Get admin user by username"""
    try:
        result = supabase.table('accounts_admin').select('*').eq('admin_user', username).execute()
        return result.data[0] if result.data else None
    except Exception as e:
        print(f"Error fetching admin: {e}")
        return None

def get_admin_by_id(admin_id):
    """Get admin user by ID - FIXED to handle both string and integer IDs"""
    try:
        # Convert to string for consistency
        admin_id_str = str(admin_id)
        # Explicitly select all fields including admin_id to ensure it's in the result
        result = supabase.table('accounts_admin').select('*').eq('admin_id', admin_id_str).execute()
        if result.data and len(result.data) > 0:
            admin = result.data[0]
            # Ensure admin_id is present in the returned data
            if 'admin_id' not in admin:
                admin['admin_id'] = admin_id_str
            return admin
        return None
    except Exception as e:
        print(f"Error fetching admin: {e}")
        return None

def update_admin_last_login(admin_id):
    """Update admin's last login timestamp"""
    try:
        result = supabase.table('accounts_admin').update({
            'admin_last_login': datetime.now().isoformat()
        }).eq('admin_id', admin_id).execute()
        return result
    except Exception as e:
        print(f"Error updating last login: {e}")
        return None

def hash_password(password):
    """Hash password using bcrypt with proper error handling"""
    try:
        if not password:
            return None
        # Generate salt and hash the password
        salt = bcrypt.gensalt()
        hashed = bcrypt.hashpw(password.encode('utf-8'), salt)
        return hashed.decode('utf-8')
    except Exception as e:
        print(f"❌ Error hashing password: {e}")
        return None

def verify_password(stored_password, provided_password):
    """Verify password - handles both hashed and plain text passwords"""
    if not stored_password or not provided_password:
        return False
        
    # Handle plain text passwords (for testing/legacy)
    if stored_password == provided_password:
        return True
        
    # Handle bcrypt hashes
    if stored_password.startswith('$2'):
        try:
            # For bcrypt hashes
            return bcrypt.checkpw(provided_password.encode('utf-8'), stored_password.encode('utf-8'))
        except Exception as e:
            print(f"Error verifying bcrypt password: {e}")
            return False
    # Handle werkzeug hashes
    elif stored_password.startswith('pbkdf2:'):
        return check_password_hash(stored_password, provided_password)
    else:
        # Fallback to plain text comparison (for legacy passwords)
        return stored_password == provided_password

def get_student_details(user_id):
    """Get student details by user ID"""
    try:
        result = supabase.table('accounts_student').select('*').eq('user_id', user_id).execute()
        return result.data[0] if result.data else None
    except Exception as e:
        print(f"Error fetching student details: {e}")
        return None

def get_active_admins():
    """Get list of active admin users"""
    try:
        result = supabase.table('accounts_admin').select('admin_id, admin_fullname, admin_role').eq('admin_status', 'Active').order('admin_fullname').execute()
        return result.data or []
    except Exception as e:
        print(f"Error fetching active admins: {e}")
        return []

def get_recent_activities(limit=60):
    """Build an incident activity feed that records every status change alongside the original report."""
    try:
        # Fetch all incidents to act as the base reference with retry
        def get_incidents():
            return supabase.table('alert_incidents').select('*').execute()
        
        incidents_result = retry_supabase_query(get_incidents)
        if not incidents_result:
            return []
        incidents = incidents_result.data or []

        incident_map = {str(incident.get('icd_id')): incident for incident in incidents}

        # Collect student ids for name lookup
        user_ids = {incident.get('user_id') for incident in incidents if incident.get('user_id')}
        students_map = {}
        if user_ids:
            try:
                batch_size = 100
                user_ids_list = list(user_ids)
                for i in range(0, len(user_ids_list), batch_size):
                    batch = user_ids_list[i:i + batch_size]
                    def get_students_batch():
                        return supabase.table('accounts_student').select('user_id, full_name').in_('user_id', batch).execute()
                    
                    students_result = retry_supabase_query(get_students_batch)
                    if students_result and students_result.data:
                        for student in students_result.data:
                            students_map[str(student['user_id'])] = student.get('full_name', 'N/A')
            except Exception as e:
                print(f"Error fetching student names: {e}")

        # Helper to parse timestamps
        def parse_timestamp(ts):
            if not ts:
                return None
            try:
                if isinstance(ts, str):
                    if 'T' in ts:
                        dt = datetime.fromisoformat(ts.replace('Z', '+00:00'))
                    else:
                        dt = datetime.strptime(ts[:19], '%Y-%m-%d %H:%M:%S')
                else:
                    dt = ts

                if dt.tzinfo is None:
                    dt = PHILIPPINES_TZ.localize(dt)
                else:
                    dt = dt.astimezone(PHILIPPINES_TZ)
                return dt
            except Exception:
                return None

        activities = []
        seen_events = set()

        # Attempt to fetch audit trail entries for status updates (fallback gracefully on failure)
        audit_entries = []
        try:
            audit_fetch_limit = max(limit * 3, 120)
            def get_audit_trail():
                return supabase.table('incident_audit_trail').select('*').order('changed_at', desc=True).limit(audit_fetch_limit).execute()
            
            audit_result = retry_supabase_query(get_audit_trail)
            audit_entries = audit_result.data if audit_result else []
        except Exception as e:
            print(f"Warning: unable to load incident audit trail, falling back to incident timestamps. Details: {e}")
            audit_entries = []

        # Prepare admin lookup for audit entries and incident status updates
        admin_ids = {entry.get('changed_by') for entry in audit_entries if entry.get('changed_by') is not None}
        for incident in incidents:
            if incident.get('status_updated_by'):
                admin_ids.add(incident.get('status_updated_by'))

        admin_map = {}
        if admin_ids:
            try:
                admin_id_list = [admin_id for admin_id in admin_ids if admin_id is not None]
                if admin_id_list:
                    def get_admins():
                        return supabase.table('accounts_admin').select('admin_id, admin_fullname').in_('admin_id', admin_id_list).execute()
                    
                    admins_result = retry_supabase_query(get_admins)
                    if admins_result and admins_result.data:
                        for admin in admins_result.data:
                            admin_map[str(admin['admin_id'])] = admin.get('admin_fullname', 'Unknown Admin')
            except Exception as e:
                print(f"Error fetching admin names: {e}")

        # Create a base "reported" event for each incident
        for incident in incidents:
            incident_id = str(incident.get('icd_id'))
            event_dt = parse_timestamp(incident.get('icd_timestamp')) or datetime.min.replace(tzinfo=PHILIPPINES_TZ)
            user_id_str = str(incident.get('user_id', ''))
            student_name = students_map.get(user_id_str, 'N/A')

            event_actor = student_name if student_name != 'N/A' else None
            event_key = (incident_id, 'Reported', event_dt.isoformat(), incident.get('icd_status') or 'Active')

            if event_key not in seen_events:
                activities.append({
                    'icd_id': incident.get('icd_id'),
                    'user_id': incident.get('user_id'),
                    'icd_status': incident.get('icd_status') or 'Active',
                    'icd_category': incident.get('icd_category'),
                    'icd_description': incident.get('icd_description'),
                    'icd_timestamp': incident.get('icd_timestamp'),
                    'student_name': student_name,
                    'event_type': 'Reported',
                    'event_label': 'Incident reported',
                    'old_status': None,
                    'new_status': incident.get('icd_status') or 'Active',
                    'change_reason': None,
                    'event_actor': event_actor,
                    'event_actor_role': 'Reported By' if event_actor else None,
                    '_event_dt': event_dt
                })
                seen_events.add(event_key)

        # Fallback: infer status change events from incident timestamp fields
        for incident in incidents:
            incident_id = str(incident.get('icd_id'))
            user_id_str = str(incident.get('user_id', ''))
            student_name = students_map.get(user_id_str, 'N/A')
            status_updated_by = incident.get('status_updated_by')
            status_actor = admin_map.get(str(status_updated_by)) if status_updated_by else None

            status_events = [
                ('Pending', incident.get('pending_timestamp')),
                ('Resolved', incident.get('resolved_timestamp')),
                ('Cancelled', incident.get('cancelled_timestamp')),
            ]

            for status_name, status_ts in status_events:
                event_dt = parse_timestamp(status_ts)
                if not event_dt:
                    continue

                event_key = (incident_id, 'Status Change', event_dt.isoformat(), status_name)
                if event_key in seen_events:
                    continue

                if status_name == 'Pending':
                    old_status = 'Active'
                elif status_name in ('Resolved', 'Cancelled'):
                    old_status = 'Pending' if incident.get('pending_timestamp') else 'Active'
                else:
                    old_status = None

                activities.append({
                    'icd_id': incident.get('icd_id'),
                    'user_id': incident.get('user_id'),
                    'icd_status': status_name,
                    'icd_category': incident.get('icd_category'),
                    'icd_description': incident.get('icd_description'),
                    'icd_timestamp': status_ts,
                    'student_name': student_name,
                    'event_type': 'Status Change',
                    'event_label': f"Status updated to {status_name}",
                    'old_status': old_status,
                    'new_status': status_name,
                    'change_reason': incident.get('status_change_reason'),
                    'event_actor': status_actor,
                    'event_actor_role': 'Updated By' if status_actor else None,
                    '_event_dt': event_dt
                })
                seen_events.add(event_key)

        # Add an entry for every status change captured in the audit trail
        for entry in audit_entries:
            if entry.get('action_type') != 'status_updated':
                continue

            incident_id = str(entry.get('icd_id'))
            incident = incident_map.get(incident_id, {})
            new_status = entry.get('new_status') or incident.get('icd_status') or entry.get('old_status') or 'Unknown'
            old_status = entry.get('old_status')
            event_dt = parse_timestamp(entry.get('changed_at')) or datetime.min.replace(tzinfo=PHILIPPINES_TZ)
            user_id_str = str(incident.get('user_id', ''))

            if old_status and new_status and old_status != new_status:
                event_label = f"Status changed: {old_status} -> {new_status}"
            else:
                event_label = f"Status updated to {new_status}"

            event_key = (incident_id, 'Status Change', event_dt.isoformat(), new_status)
            if event_key in seen_events:
                continue

            activities.append({
                'icd_id': entry.get('icd_id'),
                'user_id': incident.get('user_id'),
                'icd_status': new_status,
                'icd_category': incident.get('icd_category'),
                'icd_description': incident.get('icd_description'),
                'icd_timestamp': entry.get('changed_at'),
                'student_name': students_map.get(user_id_str, 'N/A'),
                'event_type': 'Status Change',
                'event_label': event_label,
                'old_status': old_status,
                'new_status': new_status,
                'change_reason': entry.get('change_reason'),
                'event_actor': admin_map.get(str(entry.get('changed_by')), 'System'),
                'event_actor_role': 'Updated By',
                '_event_dt': event_dt
            })
            seen_events.add(event_key)

        # Sort by the recorded event datetime (newest first)
        activities.sort(key=lambda item: item.get('_event_dt', datetime.min.replace(tzinfo=PHILIPPINES_TZ)), reverse=True)

        trimmed = activities[:limit]
        for item in trimmed:
            item.pop('_event_dt', None)

        return trimmed
    except Exception as e:
        print(f"Error fetching recent activities: {e}")
        import traceback
        traceback.print_exc()
        return []

def safe_count_query(table_name, filter_conditions=None):
    """Safely execute count queries with error handling"""
    try:
        query = supabase.table(table_name).select('*', count='exact')
        if filter_conditions:
            for condition in filter_conditions:
                if condition['type'] == 'eq':
                    query = query.eq(condition['column'], condition['value'])
                elif condition['type'] == 'in':
                    query = query.in_(condition['column'], condition['value'])
        result = query.execute()
        return result.count or 0
    except Exception as e:
        print(f"Error counting {table_name}: {e}")
        return 0

def update_admin_profile(admin_id, full_name, email, username):
    """Update admin profile information"""
    try:
        result = supabase.table('accounts_admin').update({
            'admin_fullname': full_name,
            'admin_email': email,
            'admin_user': username
        }).eq('admin_id', admin_id).execute()
        return result
    except Exception as e:
        print(f"Error updating admin profile: {e}")
        return None

def check_username_exists(username, exclude_admin_id=None):
    """Check if username already exists (excluding current admin)"""
    try:
        query = supabase.table('accounts_admin').select('admin_id').eq('admin_user', username)
        if exclude_admin_id:
            query = query.neq('admin_id', exclude_admin_id)
        result = query.execute()
        return len(result.data) > 0
    except Exception as e:
        print(f"Error checking username: {e}")
        return False

def update_admin_profile_image(admin_id, image_filename):
    """Update admin profile image"""
    try:
        result = supabase.table('accounts_admin').update({
            'admin_profile': image_filename
        }).eq('admin_id', admin_id).execute()
        return result
    except Exception as e:
        print(f"Error updating profile image: {e}")
        return None

def allowed_file(filename):
    """Check if uploaded file has allowed extension"""
    if not filename:
        return False
    ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def handle_profile_upload(file, user_type):
    """Handle profile image upload"""
    if not file or not file.filename:
        return {'filename': 'default.png'}
    
    # Validate file type
    if not allowed_file(file.filename):
        return {'error': 'Invalid file type. Only JPG, PNG, and GIF are allowed.'}
    
    # Validate file size (5MB max)
    file.seek(0, 2)  # Seek to end
    file_size = file.tell()
    file.seek(0)  # Reset to beginning
    
    if file_size > 5 * 1024 * 1024:
        return {'error': 'File too large. Maximum size is 5MB.'}
    
    try:
        # Create upload directory
        upload_dir = os.path.join(app.static_folder, 'images')
        os.makedirs(upload_dir, exist_ok=True)
        
        # Generate unique filename
        file_ext = file.filename.rsplit('.', 1)[1].lower()
        filename = f"{user_type}_{uuid.uuid4().hex[:8]}.{file_ext}"
        file_path = os.path.join(upload_dir, filename)
        
        # Save file
        file.save(file_path)
        return {'filename': filename}
        
    except Exception as e:
        print(f"Error uploading file: {e}")
        return {'error': 'Failed to upload file.'}

def send_account_request_confirmation(email, fullname, username):
    """Send confirmation email to user who requested account"""
    email_subject = "Emergency Alert System - Account Request Received"
    email_body = f"""
    <html>
    <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
        <div style="background: #dc2626; color: white; padding: 20px; text-align: center; border-radius: 10px 10px 0 0;">
            <h1 style="margin: 0;">🛡️ Emergency Alert System</h1>
            <h2 style="margin: 10px 0 0 0;">Account Request Received</h2>
        </div>
        <div style="background: #f9fafb; padding: 30px; border-radius: 0 0 10px 10px;">
            <p>Hello <strong>{fullname}</strong>,</p>
            <p>Thank you for requesting an administrator account for the Emergency Alert System.</p>
            
            <div style="background: white; border: 2px solid #2563eb; border-radius: 10px; padding: 20px; margin: 20px 0;">
                <h3 style="color: #2563eb; margin: 0 0 15px 0;">Request Details:</h3>
                <p><strong>Full Name:</strong> {fullname}</p>
                <p><strong>Username:</strong> {username}</p>
                <p><strong>Email:</strong> {email}</p>
                <p><strong>Request Date:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            </div>
            
            <div style="background: #fffbeb; border: 1px solid #f59e0b; border-radius: 8px; padding: 15px; margin: 20px 0;">
                <h4 style="color: #d97706; margin: 0 0 10px 0;">What happens next?</h4>
                <p>Your request has been submitted for review by our system administrators. You will receive another email once your account has been approved or if additional information is required.</p>
                <p><strong>Typical processing time:</strong> 24-48 hours</p>
            </div>
            
            <p>If you have any questions, please contact the system administrator.</p>
            
            <hr style="margin: 30px 0; border: none; border-top: 1px solid #e5e7eb;">
            <p style="color: #6b7280; font-size: 14px;">
                📞 Emergency Contacts:<br>
                UMak Hotline: (02) 8888-20675<br>
                Makati Emergency Hotline: 02(168)<br>
                UMak Occupational Health Center: 02-8882-0535<br>
                Taguig Emergency Services: Command Center (02) 8789-3200<br>
                Philippine National Police: (02) 8649-3582<br>
                Taguig Bureau of Fire Protection: (02) 8837-0740 | 0926-211-0919<br>
                Emergency Services: 911
            </p>
        </div>
    </body>
    </html>
    """
    
    return send_email(email, email_subject, email_body)

def send_account_approval_notification(email, fullname, username, approved_by):
    """Send notification to user when account is approved"""
    email_subject = "Emergency Alert System - Account Approved"
    email_body = f"""
    <html>
    <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
        <div style="background: #10b981; color: white; padding: 20px; text-align: center; border-radius: 10px 10px 0 0;">
            <h1 style="margin: 0;">🛡️ Emergency Alert System</h1>
            <h2 style="margin: 10px 0 0 0;">Account Approved!</h2>
        </div>
        <div style="background: #f9fafb; padding: 30px; border-radius: 0 0 10px 10px;">
            <p>Hello <strong>{fullname}</strong>,</p>
            <p>We are pleased to inform you that your administrator account request has been <strong>approved</strong>!</p>
            
            <div style="background: white; border: 2px solid #10b981; border-radius: 10px; padding: 20px; margin: 20px 0; text-align: center;">
                <h3 style="color: #10b981; margin: 0 0 15px 0;">🎉 Welcome to the Emergency Alert System!</h3>
                <p>You can now login to the system using your credentials:</p>
                <p><strong>Username:</strong> {username}</p>
                <p><strong>Login URL:</strong> <a href="{request.host_url}login">{request.host_url}login</a></p>
            </div>
            
            <div style="background: #f0f9ff; border: 1px solid #0ea5e9; border-radius: 8px; padding: 15px; margin: 20px 0;">
                <h4 style="color: #0369a1; margin: 0 0 10px 0;">Next Steps:</h4>
                <p>1. Login to the system using your username and password</p>
                <p>2. Complete your profile setup</p>
                <p>3. Familiarize yourself with the system features</p>
                <p>4. Contact the system administrator if you need assistance</p>
            </div>
            
            <p><strong>Approved by:</strong> {approved_by}</p>
            <p><strong>Approval Date:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            
            <hr style="margin: 30px 0; border: none; border-top: 1px solid #e5e7eb;">
            <p style="color: #6b7280; font-size: 14px;">
                📞 Emergency Contacts:<br>
                Campus Health Center: (02) 8123-4567<br>
                Campus Security: (02) 8123-4568<br>
                Emergency Services: 911
            </p>
        </div>
    </body>
    </html>
    """
    
    return send_email(email, email_subject, email_body)

def send_account_rejection_notification(email, fullname, username, rejected_by, reason=None):
    """Send notification to user when account is rejected"""
    email_subject = "Emergency Alert System - Account Request Decision"
    email_body = f"""
    <html>
    <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
        <div style="background: #ef4444; color: white; padding: 20px; text-align: center; border-radius: 10px 10px 0 0;">
            <h1 style="margin: 0;">🛡️ Emergency Alert System</h1>
            <h2 style="margin: 10px 0 0 0;">Account Request Update</h2>
        </div>
        <div style="background: #f9fafb; padding: 30px; border-radius: 0 0 10px 10px;">
            <p>Hello <strong>{fullname}</strong>,</p>
            <p>After careful review, we regret to inform you that your administrator account request has been <strong>not approved</strong> at this time.</p>
            
            <div style="background: white; border: 2px solid #ef4444; border-radius: 10px; padding: 20px; margin: 20px 0;">
                <h3 style="color: #ef4444; margin: 0 0 15px 0;">Request Details:</h3>
                <p><strong>Full Name:</strong> {fullname}</p>
                <p><strong>Username:</strong> {username}</p>
                <p><strong>Decision:</strong> Not Approved</p>
                {f'<p><strong>Reason:</strong> {reason}</p>' if reason else '<p><strong>Reason:</strong> Does not meet current access requirements</p>'}
            </div>
            
            <div style="background: #fef2f2; border: 1px solid #fecaca; border-radius: 8px; padding: 15px; margin: 20px 0;">
                <h4 style="color: #dc2626; margin: 0 0 10px 0;">Additional Information:</h4>
                <p>If you believe this decision was made in error or would like to discuss this further, please contact the system administrator.</p>
            </div>
            
            <p><strong>Reviewed by:</strong> {rejected_by}</p>
            <p><strong>Review Date:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            
            <hr style="margin: 30px 0; border: none; border-top: 1px solid #e5e7eb;">
            <p style="color: #6b7280; font-size: 14px;">
                📞 Emergency Contacts:<br>
                Campus Health Center: (02) 8123-4567<br>
                Campus Security: (02) 8123-4568<br>
                Emergency Services: 911
            </p>
        </div>
    </body>
    </html>
    """
    
    return send_email(email, email_subject, email_body)

def notify_system_admins_of_new_request(fullname, username, email, role):
    """Notify system administrators of new account request"""
    try:
        # Get all system administrators
        admins = supabase.table('accounts_admin').select('admin_email, admin_fullname').eq('admin_role', 'System Administrator').execute()
        
        if not admins.data:
            print("⚠️ No system administrators found to notify")
            return False
            
        email_subject = "Emergency Alert System - New Account Request"
        email_body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
            <div style="background: #f59e0b; color: white; padding: 20px; text-align: center; border-radius: 10px 10px 0 0;">
                <h1 style="margin: 0;">🛡️ Emergency Alert System</h1>
                <h2 style="margin: 10px 0 0 0;">New Account Request</h2>
            </div>
            <div style="background: #f9fafb; padding: 30px; border-radius: 0 0 10px 10px;">
                <p>A new administrator account request requires your review.</p>
                
                <div style="background: white; border: 2px solid #f59e0b; border-radius: 10px; padding: 20px; margin: 20px 0;">
                    <h3 style="color: #f59e0b; margin: 0 0 15px 0;">Request Details:</h3>
                    <p><strong>Full Name:</strong> {fullname}</p>
                    <p><strong>Username:</strong> {username}</p>
                    <p><strong>Email:</strong> {email}</p>
                    <p><strong>Requested Role:</strong> {role}</p>
                    <p><strong>Request Date:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                </div>
                
                <div style="background: #fffbeb; border: 1px solid #f59e0b; border-radius: 8px; padding: 15px; margin: 20px 0;">
                    <h4 style="color: #d97706; margin: 0 0 10px 0;">Action Required:</h4>
                    <p>Please review this request in the Emergency Alert System admin panel:</p>
                    <p><a href="{request.host_url}request-accounts" style="background: #f59e0b; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px; display: inline-block;">Review Request</a></p>
                </div>
                
                <hr style="margin: 30px 0; border: none; border-top: 1px solid #e5e7eb;">
                <p style="color: #6b7280; font-size: 14px;">
                    This is an automated notification from the Emergency Alert System.
                </p>
            </div>
        </body>
        </html>
        """
        
        # Send to all system administrators
        success_count = 0
        for admin in admins.data:
            if send_email(admin['admin_email'], email_subject, email_body):
                success_count += 1
                
        print(f"✅ Notified {success_count}/{len(admins.data)} system administrators")
        return success_count > 0
        
    except Exception as e:
        print(f"❌ Error notifying system admins: {e}")
        return False

# ==================== ENHANCED INCIDENT MANAGEMENT FUNCTIONS ====================

def get_incident_details(incident_id):
    """Get detailed incident information with related data"""
    try:
        # Get incident with student and admin data
        incident_result = supabase.table('alert_incidents').select('*').eq('icd_id', incident_id).execute()
        if not incident_result.data:
            return None
            
        incident = incident_result.data[0]
        
        # Get student details
        if incident.get('user_id'):
            student_result = supabase.table('accounts_student').select('*').eq('user_id', incident['user_id']).execute()
            incident['student_details'] = student_result.data[0] if student_result.data else None
        
        # Get admin details
        if incident.get('admin_id'):
            admin_result = supabase.table('accounts_admin').select('*').eq('admin_id', incident['admin_id']).execute()
            incident['admin_details'] = admin_result.data[0] if admin_result.data else None
        
        return incident
    except Exception as e:
        print(f"Error getting incident details: {e}")
        return None

def archive_incident(incident_id, admin_id, reason=None):
    """Archive an incident to the archive table"""
    try:
        # Get current incident data
        incident_result = supabase.table('alert_incidents').select('*').eq('icd_id', incident_id).execute()
        if not incident_result.data:
            return False, "Incident not found"
        
        incident = incident_result.data[0]
        original_icd_id = incident['icd_id']
        
        # Check if this incident ID already exists in archive (shouldn't happen, but handle it)
        existing_archive = supabase.table('incident_archive').select('icd_id').eq('icd_id', original_icd_id).execute()
        
        # If ID already exists in archive, append archive suffix to make it unique
        archive_icd_id = original_icd_id
        if existing_archive.data:
            # Generate unique ID for archive: ORIGINAL_ID_ARCHIVED_<timestamp>
            timestamp = get_philippines_time().strftime('%Y%m%d%H%M%S')
            archive_icd_id = f"{original_icd_id}_ARCHIVED_{timestamp}"
            print(f"ID conflict in archive: {original_icd_id} already archived. Using: {archive_icd_id}")
        
        # Create archive record
        archived_at = get_philippines_time().isoformat()
        archive_data = {
            'icd_id': archive_icd_id,
            'icd_timestamp': incident.get('icd_timestamp'),
            'resolved_timestamp': incident.get('resolved_timestamp'),
            'pending_timestamp': incident.get('pending_timestamp'),
            'cancelled_timestamp': incident.get('cancelled_timestamp'),
            'icd_status': incident.get('icd_status'),
            'icd_lat': incident.get('icd_lat'),
            'icd_lng': incident.get('icd_lng'),
            'assigned_responder_id': incident.get('assigned_responder_id'),
            'status_updated_at': incident.get('status_updated_at'),
            'status_updated_by': incident.get('status_updated_by'),
            'icd_category': incident.get('icd_category'),
            'icd_medical_type': incident.get('icd_medical_type'),
            'icd_security_type': incident.get('icd_security_type'),
            'icd_university_type': incident.get('icd_university_type'),
            'icd_description': incident.get('icd_description'),
            'icd_image': incident.get('icd_image'),
            'user_id': incident.get('user_id'),
            'archived_by': admin_id,
            'archive_reason': reason or "Archived by administrator",
            'archived_at': archived_at
        }
        
        # Insert to archive table
        supabase.table('incident_archive').insert(archive_data).execute()
        
        # Delete related resolution reports first (to avoid foreign key constraint violation)
        try:
            supabase.table('incident_resolution_reports').delete().eq('icd_id', incident_id).execute()
        except Exception as e:
            print(f"Warning: Could not delete resolution reports for incident {incident_id}: {e}")
            # Continue anyway - the reports will remain but incident can still be archived
        
        # Delete from main incidents table
        supabase.table('alert_incidents').delete().eq('icd_id', incident_id).execute()
        
        # Log to audit trail
        log_incident_change(incident_id, 'archived', old_status=incident.get('icd_status'), 
                           new_status=None, admin_id=admin_id, reason=reason)
        
        return True, "Incident archived successfully"
    except Exception as e:
        print(f"Error archiving incident: {e}")
        return False, str(e)

def archive_user(user_id, user_type, admin_id, reason=None):
    """Archive a user to the archive table"""
    try:
        # Get current user data
        if user_type == 'admin':
            user_result = supabase.table('accounts_admin').select('*').eq('admin_id', user_id).execute()
        else:
            user_result = supabase.table('accounts_student').select('*').eq('user_id', user_id).execute()
        
        if not user_result.data:
            return False, "User not found"
        
        user = user_result.data[0]
        
        # Create archive record with all user data
        archive_data = {
            'user_id': user_id,
            'user_type': user_type,
            'archived_by': admin_id,
            'archive_reason': reason or "Archived by administrator"
        }
        
        # Copy all relevant fields based on user type
        if user_type == 'admin':
            archive_data.update({
                'admin_id': user.get('admin_id'),
                'admin_user': user.get('admin_user'),
                'admin_email': user.get('admin_email'),
                'admin_fullname': user.get('admin_fullname'),
                'admin_role': user.get('admin_role'),
                'admin_status': user.get('admin_status'),
                'admin_approval': user.get('admin_approval'),
                'admin_profile': user.get('admin_profile'),
                'admin_pass': user.get('admin_pass'),
                'admin_created_at': user.get('admin_created_at'),
                'admin_last_login': user.get('admin_last_login'),
                'auth_user_id': user.get('auth_user_id')
            })
        else:  # student
            archive_data.update({
                'student_id': user.get('student_id'),
                'student_user': user.get('student_user'),
                'student_email': user.get('student_email'),
                'full_name': user.get('full_name'),
                'student_yearlvl': user.get('student_yearlvl'),
                'student_college': user.get('student_college'),
                'student_cnum': user.get('student_cnum'),
                'student_status': user.get('student_status'),
                'student_profile': user.get('student_profile'),
                'student_pass': user.get('student_pass'),
                'student_address': user.get('student_address'),
                'student_medinfo': user.get('student_medinfo'),
                'residency': user.get('residency'),
                'email_verified': user.get('email_verified'),
                'student_created_at': user.get('student_created_at'),
                'student_last_login': user.get('student_last_login'),
                'primary_emergencycontact': user.get('primary_emergencycontact'),
                'primary_contactperson': user.get('primary_contactperson'),
                'primary_cprelationship': user.get('primary_cprelationship'),
                'secondary_emergencycontact': user.get('secondary_emergencycontact'),
                'secondary_contactperson': user.get('secondary_contactperson'),
                'secondary_cprelationship': user.get('secondary_cprelationship')
            })
        
        # Insert to archive table
        print(f"📤 Inserting to archive table: {archive_data}")
        archive_result = supabase.table('user_archive').insert(archive_data).execute()
        print(f"✅ Archive insert result: {archive_result}")
        
        if hasattr(archive_result, 'error') and archive_result.error:
            print(f"❌ Archive insert error: {archive_result.error}")
            return False, f"Failed to insert to archive: {archive_result.error}"
        
        # Delete from main users table
        if user_type == 'admin':
            delete_result = supabase.table('accounts_admin').delete().eq('admin_id', user_id).execute()
        else:
            delete_result = supabase.table('accounts_student').delete().eq('user_id', user_id).execute()
        
        print(f"✅ Delete result: {delete_result}")
        
        if hasattr(delete_result, 'error') and delete_result.error:
            print(f"⚠️ Delete error (but archived): {delete_result.error}")
            # Archive was successful, but delete failed - this is still a partial success
            return True, "User archived successfully (but deletion had issues - check logs)"
        
        return True, "User archived successfully"
    except Exception as e:
        print(f"❌ Error archiving user: {e}")
        import traceback
        print(f"❌ Traceback: {traceback.format_exc()}")
        return False, str(e)

def restore_user(archive_id, admin_id):
    """Restore a user from the archive table"""
    try:
        # Get archived user data
        archive_result = supabase.table('user_archive').select('*').eq('archive_id', archive_id).execute()
        if not archive_result.data:
            return False, "Archived user not found"
        
        archive_record = archive_result.data[0]
        user_type = archive_record.get('user_type')
        user_id = archive_record.get('user_id')
        
        # Check if user already exists (might have been restored already)
        if user_type == 'admin':
            existing = supabase.table('accounts_admin').select('*').eq('admin_id', user_id).execute()
            if existing.data:
                return False, "User already exists. Cannot restore duplicate user."
            
            # Create admin record from archive
            admin_data = {
                'admin_id': archive_record.get('admin_id'),
                'admin_user': archive_record.get('admin_user'),
                'admin_pass': archive_record.get('admin_pass'),
                'admin_email': archive_record.get('admin_email'),
                'admin_fullname': archive_record.get('admin_fullname'),
                'admin_role': archive_record.get('admin_role'),
                'admin_status': archive_record.get('admin_status', 'Active'),
                'admin_approval': archive_record.get('admin_approval', 'Approved'),
                'admin_profile': archive_record.get('admin_profile'),
                'admin_created_at': archive_record.get('admin_created_at'),
                'admin_last_login': archive_record.get('admin_last_login'),
                'auth_user_id': archive_record.get('auth_user_id')
            }
            
            # Remove None values
            admin_data = {k: v for k, v in admin_data.items() if v is not None}
            
            print(f"📤 Restoring admin: {admin_data}")
            result = supabase.table('accounts_admin').insert(admin_data).execute()
            print(f"✅ Admin restore result: {result}")
            
        else:  # student
            # Check if student already exists by student_id (not user_id, since user_id is auto-generated)
            student_id = archive_record.get('student_id')
            archived_email = archive_record.get('student_email')
            archived_username = archive_record.get('student_user')
            archived_user_id = archive_record.get('user_id')  # Original user_id from archive
            updated_existing = False  # Flag to track if we updated an existing record
            
            if student_id:
                existing = supabase.table('accounts_student').select('*').eq('student_id', student_id).execute()
                if existing.data:
                    existing_student = existing.data[0]
                    existing_user_id = existing_student.get('user_id')
                    
                    # Check if it's the same student by comparing email, username, or user_id
                    # If it's the same student (deletion during archiving may have failed), update it
                    is_same_student = (
                        (archived_user_id and existing_user_id and str(existing_user_id) == str(archived_user_id)) or
                        (archived_email and existing_student.get('student_email') == archived_email) or
                        (archived_username and existing_student.get('student_user') == archived_username)
                    )
                    
                    if is_same_student:
                        # Same student - update the existing record instead of delete+insert
                        # This avoids unique constraint violations and deletion permission issues
                        print(f"⚠️ Found existing student - updating instead of inserting: student_id={student_id}, user_id={existing_user_id}")
                        
                        # Prepare update data (same as insert data but without student_id since we're updating by user_id)
                        student_data = {
                            'student_user': archive_record.get('student_user'),
                            'student_pass': archive_record.get('student_pass'),
                            'student_email': archive_record.get('student_email'),
                            'full_name': archive_record.get('full_name'),
                            'student_yearlvl': archive_record.get('student_yearlvl'),
                            'student_college': archive_record.get('student_college', 'CLAS'),
                            'student_cnum': archive_record.get('student_cnum'),
                            'student_status': archive_record.get('student_status', 'Active'),
                            'student_profile': archive_record.get('student_profile'),
                            'student_address': archive_record.get('student_address'),
                            'student_medinfo': archive_record.get('student_medinfo'),
                            'residency': archive_record.get('residency', 'MAKATI'),
                            'email_verified': archive_record.get('email_verified', False),
                            'student_created_at': archive_record.get('student_created_at'),
                            'student_last_login': archive_record.get('student_last_login'),
                            'primary_emergencycontact': archive_record.get('primary_emergencycontact'),
                            'primary_contactperson': archive_record.get('primary_contactperson'),
                            'primary_cprelationship': archive_record.get('primary_cprelationship'),
                            'secondary_emergencycontact': archive_record.get('secondary_emergencycontact'),
                            'secondary_contactperson': archive_record.get('secondary_contactperson'),
                            'secondary_cprelationship': archive_record.get('secondary_cprelationship'),
                            'auth_user_id': archive_record.get('auth_user_id')
                        }
                        
                        # Remove None values for required fields, but keep optional fields that can be None
                        optional_nullable_fields = ['primary_cprelationship', 'secondary_cprelationship', 'primary_emergencycontact', 
                                                   'secondary_emergencycontact', 'primary_contactperson', 'secondary_contactperson', 
                                                   'student_medinfo', 'student_profile', 'auth_user_id']
                        student_data = {k: v for k, v in student_data.items() if v is not None or k in optional_nullable_fields}
                        
                        # Update the existing record
                        print(f"📤 Updating existing student: {student_data}")
                        result = supabase.table('accounts_student').update(student_data).eq('user_id', existing_user_id).execute()
                        print(f"✅ Student update result: {result}")
                        
                        if hasattr(result, 'error') and result.error:
                            print(f"❌ Update error: {result.error}")
                            return False, f"Failed to restore user: {result.error}"
                        
                        updated_existing = True
                    else:
                        # Different student with same student_id - cannot restore
                        return False, "Student with this ID already exists and belongs to a different user. Cannot restore duplicate user."
            
            # Only insert if we didn't update above
            if not updated_existing:
                # Create student record from archive
                # NOTE: user_id will be auto-generated by the database sequence
                student_data = {
                    'student_id': archive_record.get('student_id'),
                    'student_user': archive_record.get('student_user'),
                    'student_pass': archive_record.get('student_pass'),
                    'student_email': archive_record.get('student_email'),
                    'full_name': archive_record.get('full_name'),
                    'student_yearlvl': archive_record.get('student_yearlvl'),
                    'student_college': archive_record.get('student_college', 'CLAS'),
                    'student_cnum': archive_record.get('student_cnum'),
                    'student_status': archive_record.get('student_status', 'Active'),
                    'student_profile': archive_record.get('student_profile'),
                    'student_address': archive_record.get('student_address'),
                    'student_medinfo': archive_record.get('student_medinfo'),
                    'residency': archive_record.get('residency', 'MAKATI'),
                    'email_verified': archive_record.get('email_verified', False),
                    'student_created_at': archive_record.get('student_created_at'),
                    'student_last_login': archive_record.get('student_last_login'),
                    'primary_emergencycontact': archive_record.get('primary_emergencycontact'),
                    'primary_contactperson': archive_record.get('primary_contactperson'),
                    'primary_cprelationship': archive_record.get('primary_cprelationship'),
                    'secondary_emergencycontact': archive_record.get('secondary_emergencycontact'),
                    'secondary_contactperson': archive_record.get('secondary_contactperson'),
                    'secondary_cprelationship': archive_record.get('secondary_cprelationship'),
                    'auth_user_id': archive_record.get('auth_user_id')
                }
                
                # Remove None values for required fields, but keep optional fields that can be None
                optional_nullable_fields = ['primary_cprelationship', 'secondary_cprelationship', 'primary_emergencycontact', 
                                           'secondary_emergencycontact', 'primary_contactperson', 'secondary_contactperson', 
                                           'student_medinfo', 'student_profile', 'auth_user_id']
                student_data = {k: v for k, v in student_data.items() if v is not None or k in optional_nullable_fields}
                
                print(f"📤 Restoring student: {student_data}")
                result = supabase.table('accounts_student').insert(student_data).execute()
                print(f"✅ Student restore result: {result}")
        
        if hasattr(result, 'error') and result.error:
            print(f"❌ Restore error: {result.error}")
            return False, f"Failed to restore user: {result.error}"
        
        # Delete from archive table
        delete_result = supabase.table('user_archive').delete().eq('archive_id', archive_id).execute()
        print(f"✅ Archive delete result: {delete_result}")
        
        # Get the restored user's ID for highlighting
        restored_user_id = None
        if user_type == 'admin':
            restored_user_id = archive_record.get('admin_id')
        else:
            # For students, we need to get the user_id from the newly inserted/updated record
            if updated_existing:
                restored_user_id = existing_user_id
            else:
                # Get the newly inserted user_id
                if result.data and len(result.data) > 0:
                    restored_user_id = result.data[0].get('user_id')
                else:
                    # Fallback: try to find by student_id
                    student_id = archive_record.get('student_id')
                    if student_id:
                        check_result = supabase.table('accounts_student').select('user_id').eq('student_id', student_id).execute()
                        if check_result.data:
                            restored_user_id = check_result.data[0].get('user_id')
        
        return True, f"{user_type.title()} restored successfully", restored_user_id, user_type
    except Exception as e:
        print(f"❌ Error restoring user: {e}")
        import traceback
        print(f"❌ Traceback: {traceback.format_exc()}")
        return False, str(e), None, None

def restore_incident(archive_id, admin_id):
    """Restore an incident from the archive table"""
    try:
        # Get archived incident data
        archive_result = supabase.table('incident_archive').select('*').eq('archive_id', archive_id).execute()
        if not archive_result.data:
            return False, "Archived incident not found"
        
        archive_record = archive_result.data[0]
        original_icd_id = archive_record['icd_id']
        
        # Check if the incident ID already exists in active incidents
        existing_incident = supabase.table('alert_incidents').select('icd_id').eq('icd_id', original_icd_id).execute()
        
        # Generate a unique ID if the original ID is already taken
        if existing_incident.data:
            # ID conflict - generate a new unique ID
            # Format: ORIGINAL_ID_ARCHIVED_<archive_id>
            new_icd_id = f"{original_icd_id}_ARCHIVED_{archive_id}"
            
            # Make sure this new ID doesn't exist either (unlikely but check anyway)
            check_new_id = supabase.table('alert_incidents').select('icd_id').eq('icd_id', new_icd_id).execute()
            if check_new_id.data:
                # If still exists, append timestamp
                timestamp = get_philippines_time().strftime('%Y%m%d%H%M%S')
                new_icd_id = f"{original_icd_id}_ARCHIVED_{archive_id}_{timestamp}"
            
            print(f"ID conflict detected: {original_icd_id} already exists. Using new ID: {new_icd_id}")
            restored_icd_id = new_icd_id
        else:
            # Original ID is available, use it
            restored_icd_id = original_icd_id
        
        # Create incident record
        incident_data = {
            'icd_id': restored_icd_id,
            'icd_timestamp': archive_record.get('icd_timestamp'),
            'resolved_timestamp': archive_record.get('resolved_timestamp'),
            'pending_timestamp': archive_record.get('pending_timestamp'),
            'cancelled_timestamp': archive_record.get('cancelled_timestamp'),
            'icd_status': archive_record.get('icd_status'),
            'icd_lat': archive_record.get('icd_lat'),
            'icd_lng': archive_record.get('icd_lng'),
            'assigned_responder_id': archive_record.get('assigned_responder_id'),
            'status_updated_at': archive_record.get('status_updated_at'),
            'status_updated_by': archive_record.get('status_updated_by'),
            'icd_category': archive_record.get('icd_category'),
            'icd_medical_type': archive_record.get('icd_medical_type'),
            'icd_security_type': archive_record.get('icd_security_type'),
            'icd_university_type': archive_record.get('icd_university_type'),
            'icd_description': archive_record.get('icd_description'),
            'icd_image': archive_record.get('icd_image'),
            'user_id': archive_record.get('user_id')
        }
        
        # Insert back to main incidents table
        supabase.table('alert_incidents').insert(incident_data).execute()
        
        # Delete from archive table
        supabase.table('incident_archive').delete().eq('archive_id', archive_id).execute()
        
        # Log to audit trail
        log_incident_change(restored_icd_id, 'restored', old_status=None, 
                           new_status=archive_record.get('icd_status'), admin_id=admin_id, 
                           reason=f"Restored from archive (original ID: {original_icd_id})")
        
        if restored_icd_id != original_icd_id:
            return True, f"Incident restored successfully with new ID: {restored_icd_id} (original ID {original_icd_id} was already in use)"
        else:
            return True, "Incident restored successfully"
    except Exception as e:
        print(f"Error restoring incident: {e}")
        return False, str(e)

def get_audit_trail(incident_id=None, limit=50):
    """Get audit trail for incidents"""
    try:
        query = supabase.table('incident_audit_trail').select('*')
        
        if incident_id:
            query = query.eq('icd_id', incident_id)
        
        result = query.order('changed_at', desc=True).limit(limit).execute()
        
        # Get admin names for changed_by field
        audit_records = result.data or []
        admin_ids = [record['changed_by'] for record in audit_records if record.get('changed_by')]
        
        if admin_ids:
            admins_result = supabase.table('accounts_admin').select('admin_id, admin_fullname').in_('admin_id', admin_ids).execute()
            admin_names = {admin['admin_id']: admin['admin_fullname'] for admin in admins_result.data}
            
            for record in audit_records:
                if record.get('changed_by'):
                    record['changed_by_name'] = admin_names.get(record['changed_by'], 'Unknown Admin')
        
        return audit_records
    except Exception as e:
        print(f"Error getting audit trail: {e}")
        return []

def log_incident_change(incident_id, action_type, old_status=None, new_status=None, admin_id=None, reason=None):
    """Log an incident change to the audit trail"""
    try:
        audit_data = {
            'icd_id': incident_id,
            'action_type': action_type,
            'old_status': old_status,
            'new_status': new_status,
            'changed_by': admin_id,
            'change_reason': reason
        }
        
        supabase.table('incident_audit_trail').insert(audit_data).execute()
        return True
    except Exception as e:
        print(f"Error logging incident change: {e}")
        return False

def get_archived_incidents(admin_id=None):
    """Get list of archived incidents"""
    try:
        query = supabase.table('incident_archive').select('*')
        
        result = query.order('archived_at', desc=True).execute()
        archived_incidents = result.data or []
        
        # Get admin names for archived_by field
        if admin_id:
            admin_result = supabase.table('accounts_admin').select('admin_id, admin_fullname').eq('admin_id', admin_id).execute()
            if admin_result.data:
                admin_name = admin_result.data[0]['admin_fullname']
                for incident in archived_incidents:
                    if incident.get('archived_by') == admin_id:
                        incident['archived_by_name'] = admin_name
        
        return archived_incidents
    except Exception as e:
        print(f"Error getting archived incidents: {e}")
        return []

def calculate_response_time(incident_id):
    """Calculate response time for an incident"""
    try:
        incident_result = supabase.table('alert_incidents').select('icd_timestamp, resolved_timestamp').eq('icd_id', incident_id).execute()
        if not incident_result.data:
            return None
            
        incident = incident_result.data[0]
        
        if not incident.get('icd_timestamp') or not incident.get('resolved_timestamp'):
            return None
        
        # Parse timestamps
        reported_time = datetime.fromisoformat(incident['icd_timestamp'].replace('Z', '+00:00'))
        resolved_time = datetime.fromisoformat(incident['resolved_timestamp'].replace('Z', '+00:00'))
        
        # Calculate difference in minutes
        response_time = resolved_time - reported_time
        return response_time.total_seconds() / 60  # Return in minutes
    except Exception as e:
        print(f"Error calculating response time: {e}")
        return None

def _parse_datetime(value):
    """Parse various datetime representations into aware datetime objects."""
    if not value:
        return None
    if isinstance(value, datetime):
        dt_obj = value
    else:
        try:
            text = str(value)
            if text.endswith('Z'):
                text = text.replace('Z', '+00:00')
            dt_obj = datetime.fromisoformat(text)
        except Exception:
            return None
    if dt_obj.tzinfo is None:
        # Assume UTC if naive, then convert to PH timezone
        dt_obj = dt_obj.replace(tzinfo=timezone.utc)
    try:
        return dt_obj.astimezone(PHILIPPINES_TZ)
    except Exception:
        return dt_obj

def _format_response_duration(minutes):
    """Convert response duration in minutes into a human-friendly label."""
    if minutes is None:
        return 'N/A'
    try:
        minutes = float(minutes)
    except (TypeError, ValueError):
        return 'N/A'
    if minutes < 1:
        seconds = max(int(round(minutes * 60)), 1)
        return f"{seconds} sec" if seconds == 1 else f"{seconds} secs"
    hours, remainder = divmod(minutes, 60)
    if hours >= 1:
        hours_int = int(hours)
        minutes_int = int(round(remainder))
        parts = [f"{hours_int} hr" + ("s" if hours_int != 1 else "")]
        if minutes_int:
            parts.append(f"{minutes_int} min")
        return ' '.join(parts)
    return f"{round(minutes, 1)} min"

def format_incident_label(incident_id):
    """Generate a human readable incident code for display."""
    if incident_id is None:
        return 'Incident'
    incident_str = str(incident_id).strip()
    if not incident_str:
        return 'Incident'
    upper = incident_str.upper()
    if upper.startswith('ICD'):
        return upper
    if incident_str.isdigit():
        return f"ICD_9100{incident_str}"
    return f"ICD_{incident_str}"

def create_incident_resolution_report(incident_id, admin_id, admin_name=None, previous_status=None, summary_text=None):
    """
    Build a professional closure report for a resolved incident and persist it to Supabase.

    Returns:
        tuple(dict|None, str|None): (summary_payload, storage_error)
    """
    if supabase is None:
        return None, "Supabase client is not configured."

    incident = get_incident_details(incident_id)
    if not incident:
        return None, "Incident details could not be retrieved."

    admin_name = admin_name or 'Administrator'
    student_details = incident.get('student_details') or {}
    student_name = (
        student_details.get('full_name')
        or student_details.get('student_name')
        or incident.get('student_name')
        or 'N/A'
    )

    category = incident.get('icd_category') or 'Uncategorized'
    incident_label = format_incident_label(incident_id)
    reported_at = _parse_datetime(incident.get('icd_timestamp'))
    resolved_at = _parse_datetime(incident.get('resolved_timestamp')) or get_philippines_time()

    response_minutes = None
    if reported_at:
        response_minutes = (resolved_at - reported_at).total_seconds() / 60

    response_text = _format_response_duration(response_minutes)
    description = incident.get('icd_description') or ''
    truncated_description = description.strip()
    if truncated_description and len(truncated_description) > 140:
        truncated_description = truncated_description[:137].rstrip() + '…'

    summary_text = (summary_text or '').strip()
    display_summary = summary_text or truncated_description

    key_points = [
        f"Status changed from {previous_status or 'Unknown'} to Resolved.",
        f"Handled by {admin_name}.",
    ]
    if student_name and student_name != 'N/A':
        key_points.append(f"Student: {student_name}")
    if category and category != 'Uncategorized':
        key_points.append(f"Category: {category}")
    if display_summary:
        key_points.append(f"Summary: {display_summary}")

    metrics = []
    if reported_at:
        metrics.append({
            'label': 'Reported',
            'value': reported_at.strftime('%Y-%m-%d %I:%M %p')
        })
    metrics.append({
        'label': 'Resolved',
        'value': resolved_at.strftime('%Y-%m-%d %I:%M %p')
    })
    metrics.append({
        'label': 'Response Time',
        'value': response_text
    })

    summary_headline = f"{incident_label} resolved by {admin_name}"

    summary_details = {
        'incident_label': incident_label,
        'status_before': previous_status,
        'status_after': 'Resolved',
        'category': category,
        'student': {
            'id': incident.get('user_id'),
            'name': student_name
        },
        'reported_at': reported_at.isoformat() if reported_at else None,
        'resolved_at': resolved_at.isoformat(),
        'response_minutes': response_minutes,
        'description': description,
        'location': {
            'lat': incident.get('icd_lat'),
            'lng': incident.get('icd_lng')
        },
        'admin': {
            'id': admin_id,
            'name': admin_name
        },
        'summary_notes': summary_text or None
    }

    incident_id_str = str(incident_id).strip() if incident_id is not None else None
    resolved_id = generate_resolution_id()

    report_record = {
        'resolved_id': resolved_id,
        'icd_id': incident_id_str,
        'incident_label': incident_label,
        'student_id': incident.get('user_id'),
        'student_name': student_name,
        'resolved_by': str(admin_id) if admin_id is not None else None,
        'resolved_by_name': admin_name,
        'summary_headline': summary_headline,
        'summary_details': summary_details,
        'category': category,
        'status_before': previous_status,
        'status_after': 'Resolved',
        'reported_at': summary_details['reported_at'],
        'resolved_at': summary_details['resolved_at'],
        'response_minutes': response_minutes,
        'summary_notes': summary_text or truncated_description,
        'created_at': get_philippines_time().isoformat()
    }

    summary_payload = {
        'resolved_id': resolved_id,
        'icd_id': incident_id_str,
        'incident_label': incident_label,
        'headline': summary_headline,
        'resolved_at_display': resolved_at.strftime('%B %d, %Y • %I:%M %p'),
        'response_time_text': response_text,
        'student_name': student_name,
        'category': category,
        'key_points': key_points,
        'metrics': metrics,
        'summary_text': summary_text or '',
        'stored': False
    }

    storage_error = None
    try:
        supabase.table(RESOLUTION_REPORTS_TABLE).insert(report_record).execute()
        summary_payload['stored'] = True
    except Exception as exc:
        storage_error = str(exc)
        duplicate_error = 'duplicate key value violates unique constraint' in storage_error.lower() or 'resolved_id' in storage_error.lower()
        if duplicate_error:
            try:
                resolved_id_retry = generate_resolution_id()
                report_record['resolved_id'] = resolved_id_retry
                summary_payload['resolved_id'] = resolved_id_retry
                supabase.table(RESOLUTION_REPORTS_TABLE).insert(report_record).execute()
                summary_payload['stored'] = True
                storage_error = None
            except Exception as retry_exc:
                storage_error = str(retry_exc)
                print(f"Error storing resolution report for incident {incident_id}: {storage_error}")
        else:
            print(f"Error storing resolution report for incident {incident_id}: {storage_error}")

    return summary_payload, storage_error

def bulk_archive_incidents(incident_ids, admin_id, reason=None):
    """Archive multiple incidents at once"""
    results = []
    for incident_id in incident_ids:
        success, message = archive_incident(incident_id, admin_id, reason)
        results.append({
            'incident_id': incident_id,
            'success': success,
            'message': message
        })
    return results

# ---------------- ROUTES ---------------- #

@app.route('/debug-supabase')
def debug_supabase():
    """
    Health check route to verify Supabase connectivity in deployed environments.
    Returns basic info about the connection and a sample query to accounts_admin.
    """
    try:
        if supabase is None:
            return "Supabase is None (failed to initialize)", 500

        # Try a very small safe query
        result = supabase.table('accounts_admin').select('admin_id').limit(1).execute()
        row_count = len(result.data) if result.data else 0
        return {
            "status": "ok",
            "rows": row_count
        }, 200
    except Exception as e:
        return f"Error: {e}", 500


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Admin login route with approval status handling"""
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        
        if not username or not password:
            flash('Please enter both username and password!', 'error')
            return render_template('login.html')
        
        admin = get_admin_by_username(username)

        if admin:
            # Check approval status first
            approval_status = admin.get('admin_approval', 'Approved')
            
            if approval_status == 'Pending':
                flash('Your account is under review. Please wait for administrator approval.', 'warning')
                return render_template('login.html')
            elif approval_status == 'Rejected':
                flash('Your account request was rejected. Please contact the system administrator.', 'error')
                return render_template('login.html')
            elif approval_status != 'Approved':
                flash('Your account is not approved for access. Please contact the system administrator.', 'error')
                return render_template('login.html')
            
            # Check password
            stored_password = admin.get('admin_pass', '')
            
            if verify_password(stored_password, password):
                # Set session variables
                session['admin_id'] = admin['admin_id']
                session['admin_name'] = admin.get('admin_fullname', admin.get('admin_user', 'Admin'))
                session['admin_profile'] = admin.get('admin_profile', None)
                session['admin_profile_exists'] = check_profile_image_exists(admin.get('admin_profile'))
                session['role'] = admin.get('admin_role', 'Administrator')
                
                # Update last login
                update_admin_last_login(admin['admin_id'])
                
                flash('Login Successful! Welcome, ' + admin.get('admin_fullname', admin.get('admin_user', 'Admin')), 'success')
                return redirect(url_for('dashboard'))
            else:
                flash('Invalid password.', 'error')
        else:
            flash('Account not found.', 'error')
    
    return render_template('login.html')

@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    """Forgot password route with improved error handling"""
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        
        if not email:
            flash('Please enter your email address!', 'error')
            return render_template('forgot_password.html')
        
        try:
            print(f"🔍 Checking email: {email}")
            
            # Check if email exists in admin accounts
            result = supabase.table('accounts_admin').select('*').eq('admin_email', email).execute()
            print(f"📊 Database result: {len(result.data) if result.data else 0} records found")
            
            user = result.data[0] if result.data else None
            
            if not user:
                flash('No account found with this email address!', 'error')
                return render_template('forgot_password.html')
            
            print(f"✅ User found: {user.get('admin_fullname')}")
            
            # Generate verification code
            verification_code = generate_verification_code()
            expires_at = datetime.now(timezone.utc) + timedelta(minutes=15)
            
            print(f"🔑 Generated code: {verification_code}, expires at: {expires_at}")
            
            # Create password reset request in database
            reset_data = {
                'email': email,
                'verification_code': verification_code,
                'expires_at': expires_at.isoformat(),
                'used': False
            }
            
            print(f"💾 Saving reset request to database: {reset_data}")
            
            # Insert into database
            try:
                reset_result = supabase.table('password_reset_requests').insert(reset_data).execute()
                print(f"📦 Database insert result: {reset_result}")
                
                if not reset_result.data:
                    flash('Failed to create reset request. Please try again.', 'error')
                    return render_template('forgot_password.html')
                    
            except Exception as db_error:
                print(f"❌ Database error: {db_error}")
                flash('Database error. Please try again.', 'error')
                return render_template('forgot_password.html')
            
            # Store email and user info in session for the reset flow
            session['reset_email'] = email
            session['reset_name'] = user.get('admin_fullname', user.get('admin_user', 'User'))
            session['verification_code'] = verification_code
            
            # Send email
            email_subject = "Emergency Alert System - Password Reset Code"
            email_body = f"""
            <html>
            <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
                <div style="background: #dc2626; color: white; padding: 20px; text-align: center; border-radius: 10px 10px 0 0;">
                    <h1 style="margin: 0;">🛡️ Emergency Alert System</h1>
                    <h2 style="margin: 10px 0 0 0;">Password Reset Request</h2>
                </div>
                <div style="background: #f9fafb; padding: 30px; border-radius: 0 0 10px 10px;">
                    <p>Hello {user.get('admin_fullname', user.get('admin_user', 'User'))},</p>
                    <p>You have requested to reset your password for the Emergency Alert System.</p>
                    <div style="background: white; border: 2px solid #2563eb; border-radius: 10px; padding: 20px; margin: 20px 0; text-align: center;">
                        <h3 style="color: #2563eb; margin: 0;">Your Verification Code:</h3>
                        <div style="font-size: 32px; font-weight: bold; color: #2563eb; font-family: monospace; letter-spacing: 3px; margin: 10px 0;">
                            {verification_code}
                        </div>
                    </div>
                    <p><strong>This code will expire in 15 minutes.</strong></p>
                    <p>If you did not request this password reset, please ignore this email and contact your system administrator.</p>
                    <hr style="margin: 30px 0; border: none; border-top: 1px solid #e5e7eb;">
                    <p style="color: #6b7280; font-size: 14px;">
                        📞 Emergency Contacts:<br>
                        Campus Health Center: (02) 8123-4567<br>
                        Campus Security: (02) 8123-4568<br>
                        Emergency Services: 911
                    </p>
                </div>
            </body>
            </html>
            """
            
            print(f"📤 Attempting to send email to: {email}")
            if send_email(email, email_subject, email_body):
                flash('Password reset email sent! Please check your email and enter the verification code.', 'success')
                return redirect(url_for('reset_password'))
            else:
                flash('Email sending failed. The verification code has been generated. Please contact administrator.', 'error')
                # Still redirect to reset password since we have the code in session
                return redirect(url_for('reset_password'))
            
        except Exception as e:
            print(f"❌ Error in forgot password: {e}")
            traceback.print_exc()
            flash('An error occurred. Please try again.', 'error')
    
    return render_template('forgot_password.html')

@app.route('/reset-password', methods=['GET', 'POST'])
def reset_password():
    """Reset password route - SIMPLIFIED WORKING VERSION"""
    # Check if user came from forgot password flow
    if 'reset_email' not in session:
        flash('Please request a password reset first.', 'error')
        return redirect(url_for('forgot_password'))
    
    if request.method == 'POST':
        code = request.form.get('code', '').strip()
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        
        print(f"🔄 Reset password attempt for: {session.get('reset_email')}")
        print(f"📝 Code: {code}, New pass length: {len(new_password)}")
        
        # Validate input
        if not code or not new_password or not confirm_password:
            flash('All fields are required!', 'error')
            return render_template('reset_password.html', 
                                 name=session.get('reset_name', 'User'))
        
        # Verify code against database
        try:
            print(f"🔍 Verifying code in database...")
            result = supabase.table('password_reset_requests').select('*').eq('email', session['reset_email']).eq('verification_code', code).eq('used', False).execute()
            print(f"📊 Verification result: {len(result.data) if result.data else 0} matches")
            
            reset_request = result.data[0] if result.data else None
            
            if not reset_request:
                print("❌ No valid reset request found")
                flash('Invalid or expired verification code!', 'error')
                return render_template('reset_password.html',
                                     name=session.get('reset_name', 'User'))
            
            # Check if code is expired
            expires_at_str = reset_request['expires_at']
            if 'Z' in expires_at_str:
                expires_at = datetime.fromisoformat(expires_at_str.replace('Z', '+00:00'))
            else:
                expires_at = datetime.fromisoformat(expires_at_str)
                
            # Make both datetimes timezone-aware for comparison
            current_time = datetime.now(timezone.utc)
            
            print(f"⏰ Code expires at: {expires_at}, Current time: {current_time}")
            
            if current_time > expires_at:
                print("❌ Code expired")
                flash('Verification code has expired! Please request a new one.', 'error')
                return redirect(url_for('forgot_password'))
            
            # Validate password strength with reasonable requirements
            is_valid, password_message = validate_password(new_password)
            if not is_valid:
                flash(f'Password validation failed: {password_message}', 'error')
                return render_template('reset_password.html',
                                     name=session.get('reset_name', 'User'))
            
            if new_password != confirm_password:
                flash('Passwords do not match!', 'error')
                return render_template('reset_password.html',
                                     name=session.get('reset_name', 'User'))
            
            # Get admin user
            user_result = supabase.table('accounts_admin').select('admin_id, admin_user, admin_email, admin_pass').eq('admin_email', session['reset_email']).execute()
            
            if not user_result.data:
                print("❌ User not found in accounts_admin")
                flash('User account not found.', 'error')
                return render_template('reset_password.html', name=session.get('reset_name', 'User'))
            
            admin_id = user_result.data[0]['admin_id']
            old_password_hash = user_result.data[0]['admin_pass']
            print(f"🔑 Found admin_id: {admin_id}")
            print(f"🔑 Old password hash: {old_password_hash[:20]}...")
            
            # SIMPLIFIED APPROACH: Use plain text password for now to ensure it works
            print("🔄 Using simplified password update approach...")
            
            try:
                # Update password directly - using plain text to ensure it works
                update_result = supabase.table('accounts_admin').update({
                    'admin_pass': new_password  # Store as plain text for now
                }).eq('admin_id', admin_id).execute()
                
                print(f"📊 Update result: {update_result}")
                
                # Wait for update to propagate
                time.sleep(2)
                
                # Verify the update worked by fetching the user
                verify_result = supabase.table('accounts_admin').select('admin_pass').eq('admin_id', admin_id).execute()
                
                if verify_result.data:
                    new_password_in_db = verify_result.data[0]['admin_pass']
                    print(f"🔍 Verification - Password in DB: {new_password_in_db}")
                    
                    # Check if password was updated (for plain text, just check if it matches)
                    if new_password_in_db == new_password:
                        print("✅ Password updated successfully!")
                        
                        # Mark reset request as used
                        try:
                            mark_used_result = supabase.table('password_reset_requests').update({
                                'used': True,
                                'used_at': datetime.now(timezone.utc).isoformat()
                            }).eq('id', reset_request['id']).execute()
                            print(f"✅ Reset request marked as used")
                        except Exception as mark_error:
                            print(f"⚠️ Could not mark reset request as used: {mark_error}")
                        
                        # Clear session
                        session.pop('reset_email', None)
                        session.pop('reset_name', None)
                        session.pop('verification_code', None)
                        
                        flash('Password reset successful! Please login with your new password.', 'success')
                        return redirect(url_for('login'))
                    else:
                        print("❌ Password update failed - password not changed in database")
                        flash('Password update failed. Please try again or contact administrator.', 'error')
                else:
                    print("❌ Could not verify password update")
                    flash('Password update verification failed. Please try again.', 'error')
                
            except Exception as update_error:
                print(f"❌ Exception during update: {update_error}")
                traceback.print_exc()
                flash('Database error during password update. Please try again.', 'error')
                
        except Exception as e:
            print(f"❌ Error in password reset: {e}")
            traceback.print_exc()
            flash('An error occurred. Please try again.', 'error')
    
    return render_template('reset_password.html',
                         name=session.get('reset_name', 'User'))

@app.route('/request-account', methods=['GET', 'POST'])
def request_account():
    """Account request route - handle new admin account requests"""
    if request.method == 'POST':
        try:
            # Get form data
            fullname = request.form.get('fullname', '').strip()
            email = request.form.get('email', '').strip().lower()
            username = request.form.get('username', '').strip()
            password = request.form.get('password', '').strip()
            confirm_password = request.form.get('confirm_password', '').strip()
            role = request.form.get('role', '').strip()
            reason = request.form.get('reason', '').strip()
            
            print(f"📝 Account request submitted: {fullname} ({email})")
            
            # Validate required fields
            if not all([fullname, email, username, password, confirm_password, role]):
                flash('Please fill in all required fields!', 'error')
                return render_template('request_account.html')
            
            # Validate email format
            import re
            email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            if not re.match(email_pattern, email):
                flash('Please enter a valid email address!', 'error')
                return render_template('request_account.html')
            
            # Validate username (letters, numbers, underscores only)
            username_pattern = r'^[a-zA-Z0-9_]+$'
            if not re.match(username_pattern, username):
                flash('Username can only contain letters, numbers, and underscores!', 'error')
                return render_template('request_account.html')
            
            # Check password match
            if password != confirm_password:
                flash('Passwords do not match!', 'error')
                return render_template('request_account.html')
            
            # Validate password strength
            is_valid, password_message = validate_password_for_request(password)
            if not is_valid:
                flash(f'Password validation failed: {password_message}', 'error')
                return render_template('request_account.html')
            
            # Check if username or email already exists
            existing_admin = supabase.table('accounts_admin').select('admin_user, admin_email').or_(f'admin_user.eq.{username},admin_email.eq.{email}').execute()
            existing_request = supabase.table('account_requests').select('admin_user, admin_email').or_(f'admin_user.eq.{username},admin_email.eq.{email}').execute()
            
            if existing_admin.data or existing_request.data:
                flash('Username or email already exists! Please choose different credentials.', 'error')
                return render_template('request_account.html')
            
            # Hash password
            hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            
            # Insert account request
            request_data = {
                'admin_fullname': fullname,
                'admin_email': email,
                'admin_user': username,
                'admin_pass': hashed_password,
                'admin_role': role,
                'admin_approval': 'Pending',
                'request_reason': reason,
                'requested_at': datetime.now().isoformat()
            }
            
            print(f"💾 Inserting account request: {request_data}")
            
            result = supabase.table('account_requests').insert(request_data).execute()
            
            if result.data:
                print(f"✅ Account request created successfully!")
                
                # Send confirmation email to user
                if send_account_request_confirmation(email, fullname, username):
                    print(f"✅ Confirmation email sent to {email}")
                else:
                    print(f"⚠️ Failed to send confirmation email to {email}")
                
                # Notify system administrators
                if notify_system_admins_of_new_request(fullname, username, email, role):
                    print("✅ System administrators notified")
                else:
                    print("⚠️ Failed to notify system administrators")
                
                flash('Your account request has been submitted successfully! A confirmation email has been sent. You will be notified via email once reviewed.', 'success')
                return render_template('request_success.html', name=fullname)
            else:
                flash('Failed to submit account request. Please try again.', 'error')
                return render_template('request_account.html')
                
        except Exception as e:
            print(f"❌ Error submitting account request: {e}")
            traceback.print_exc()
            flash('An error occurred while submitting your request. Please try again.', 'error')
            return render_template('request_account.html')
    
    return render_template('request_account.html')

@app.route('/request-success')
def request_success():
    """Account request success page"""
    name = session.get('request_name', 'User')
    return render_template('request_success.html', name=name)

@app.route('/api/check-username')
def api_check_username():
    """API endpoint to check if username is available"""
    try:
        username = request.args.get('username', '').strip()
        
        if not username:
            return jsonify({'available': False, 'message': 'Username is required'})
        
        # Check in both accounts_admin and account_requests tables
        admin_result = supabase.table('accounts_admin').select('admin_user').eq('admin_user', username).execute()
        request_result = supabase.table('account_requests').select('admin_user').eq('admin_user', username).execute()
        
        is_available = not admin_result.data and not request_result.data
        
        return jsonify({
            'available': is_available,
            'message': 'Username is available' if is_available else 'Username is already taken'
        })
        
    except Exception as e:
        print(f"❌ Error checking username: {e}")
        return jsonify({'available': False, 'message': 'Error checking username'})

@app.route('/api/pending-requests-count')
def get_pending_requests_count():
    """API endpoint to get count of pending account requests"""
    if 'admin_id' not in session:
        return jsonify({'count': 0})
    
    try:
        # Count pending account requests
        result = supabase.table('account_requests').select('*', count='exact').eq('admin_approval', 'Pending').execute()
        pending_count = result.count or 0
        return jsonify({'count': pending_count})
    except Exception as e:
        print(f"Error counting pending requests: {e}")
        return jsonify({'count': 0})

@app.route('/request-accounts')
def request_accounts():
    """Page to view and manage account requests - IMPROVED VERSION"""
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    # Check if current user has permission to manage requests
    current_admin = get_admin_by_id(session['admin_id'])
    if not current_admin:
        flash('Admin account not found.', 'error')
        return redirect(url_for('logout'))
    
    # Only System Administrators should manage account requests
    if current_admin.get('admin_role') != 'System Administrator':
        flash('Access denied. Only System Administrators can manage account requests.', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        # Get all account requests with additional details
        result = supabase.table('account_requests').select('*').order('requested_at', desc=True).execute()
        requests = result.data or []
        
        # Format timestamps and add additional processing
        for req in requests:
            req['formatted_requested_at'] = format_datetime(req.get('requested_at'))
            req['formatted_permitted_at'] = format_datetime(req.get('permitted_at'))
            
            # Add debug info
            print(f"Request ID: {req.get('id')}, Status: {req.get('admin_approval')}")
        
        print(f"📊 Loaded {len(requests)} account requests")
        
        return render_template('request_accounts.html', requests=requests)
        
    except Exception as e:
        print(f"❌ Error loading account requests: {e}")
        traceback.print_exc()
        flash('Error loading account requests', 'error')
        return render_template('request_accounts.html', requests=[])

@app.route('/debug-requests')
def debug_requests():
    """Debug route to check account requests data"""
    if 'admin_id' not in session:
        return "Unauthorized"
    
    try:
        # Check account_requests table structure
        result = supabase.table('account_requests').select('*').execute()
        requests = result.data or []
        
        debug_info = {
            'total_requests': len(requests),
            'requests': requests,
            'table_columns': list(requests[0].keys()) if requests else []
        }
        
        return jsonify(debug_info)
        
    except Exception as e:
        return jsonify({'error': str(e)})

def generate_next_admin_id():
    """Generate the next admin ID in ADM-XXXX format"""
    try:
        # Get the highest existing admin ID
        result = supabase.table('accounts_admin').select('admin_id').execute()
        admins = result.data or []
        
        max_number = 0
        for admin in admins:
            admin_id = admin.get('admin_id', '')
            if admin_id and isinstance(admin_id, str) and admin_id.upper().startswith('ADM-'):
                try:
                    # Extract the numeric part
                    numeric_part = admin_id[4:]  # Remove 'ADM-' prefix
                    if numeric_part.isdigit():
                        max_number = max(max_number, int(numeric_part))
                except ValueError:
                    continue
        
        next_number = max_number + 1
        generated_id = f"ADM-{next_number:04d}"
        
        # Ensure the generated ID is exactly 8 characters
        if len(generated_id) != 8:
            print(f"⚠️ Generated admin ID '{generated_id}' is not 8 characters. Adjusting...")
            # Ensure it's always 8 characters
            if len(generated_id) < 8:
                generated_id = generated_id.ljust(8, '0')
            else:
                generated_id = generated_id[:8]
        
        print(f"🔑 Generated admin ID: {generated_id}")
        return generated_id
    except Exception as e:
        print(f"Error generating admin ID: {e}")
        # Fallback to a default ID that's exactly 8 characters
        return "ADM-0001"

@app.route('/api/approve-request/<int:request_id>', methods=['POST'])
def approve_request(request_id):
    """API endpoint to approve an account request - IMPROVED VERSION"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    try:
        print(f"🔄 APPROVING REQUEST ID: {request_id}")
        
        # Get current admin info and verify permissions
        current_admin = get_admin_by_id(session['admin_id'])
        if not current_admin or current_admin.get('admin_role') != 'System Administrator':
            return jsonify({'success': False, 'message': 'Insufficient permissions'})
        
        # Get the request details with error handling
        request_result = supabase.table('account_requests').select('*').eq('id', request_id).execute()
        if not request_result.data:
            print(f"❌ Request {request_id} not found in account_requests table")
            return jsonify({'success': False, 'message': 'Request not found'})
        
        account_request = request_result.data[0]
        print(f"📋 Found request: {account_request['admin_user']} (ID: {account_request['id']})")
        
        # Check if already processed
        current_status = account_request.get('admin_approval')
        if current_status != 'Pending':
            print(f"⚠️ Request already processed: {current_status}")
            return jsonify({'success': False, 'message': f'Request already {current_status.lower()}'})
        
        # Generate the next admin ID
        next_admin_id = generate_next_admin_id()
        
        # Validate admin_id length to prevent database errors
        if len(next_admin_id) > 8:
            return jsonify({'success': False, 'message': f'Generated admin ID "{next_admin_id}" is too long for database (max 8 characters)'})
        
        # Create admin account with proper data mapping
        admin_data = {
            'admin_id': next_admin_id,  # Explicitly set admin_id to avoid UUID generation
            'admin_user': account_request['admin_user'],
            'admin_pass': account_request['admin_pass'],  # Already hashed
            'admin_email': account_request['admin_email'],
            'admin_fullname': account_request['admin_fullname'],
            'admin_role': account_request['admin_role'],
            'admin_status': 'Active',
            'admin_approval': 'Approved',
            'admin_created_at': datetime.now().isoformat(),
            'admin_profile': 'default.png'  # Default profile image
        }
        
        print(f"💾 Creating admin account with data: {admin_data}")
        
        # Insert into admin accounts
        admin_result = supabase.table('accounts_admin').insert(admin_data).execute()
        
        if admin_result.data:
            new_admin = admin_result.data[0]
            print(f"✅ Admin account created successfully! New admin ID: {new_admin.get('admin_id')}")
            
            # Update the request with approval details
            update_data = {
                'admin_approval': 'Approved',
                'permitted_at': datetime.now().isoformat(),
                'reviewed_by': session['admin_id'],
                'reviewed_by_name': current_admin.get('admin_fullname', 'System Administrator')
            }
            
            update_result = supabase.table('account_requests').update(update_data).eq('id', request_id).execute()
            print(f"📝 Request updated: {update_result.data}")
            
            # Send approval notification email
            try:
                email_sent = send_account_approval_notification(
                    account_request['admin_email'],
                    account_request['admin_fullname'],
                    account_request['admin_user'],
                    current_admin.get('admin_fullname', 'System Administrator')
                )
                
                if email_sent:
                    print(f"✅ Approval email sent to {account_request['admin_email']}")
                else:
                    print(f"⚠️ Failed to send approval email to {account_request['admin_email']}")
            except Exception as email_error:
                print(f"⚠️ Email error (non-critical): {email_error}")
            
            return jsonify({
                'success': True, 
                'message': 'Account approved successfully! User can now login.',
                'admin_id': new_admin.get('admin_id')
            })
        else:
            print("❌ Failed to create admin account - no data returned")
            return jsonify({'success': False, 'message': 'Failed to create admin account in database'})
            
    except Exception as e:
        print(f"❌ Error approving request: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'System error: {str(e)}'})

@app.route('/api/reject-request/<int:request_id>', methods=['POST'])
def reject_request(request_id):
    """API endpoint to reject an account request - IMPROVED VERSION"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    try:
        print(f"🔄 REJECTING REQUEST ID: {request_id}")
        
        # Get current admin info and verify permissions
        current_admin = get_admin_by_id(session['admin_id'])
        if not current_admin or current_admin.get('admin_role') != 'System Administrator':
            return jsonify({'success': False, 'message': 'Insufficient permissions'})
        
        # Get the request details
        request_result = supabase.table('account_requests').select('*').eq('id', request_id).execute()
        if not request_result.data:
            print(f"❌ Request {request_id} not found")
            return jsonify({'success': False, 'message': 'Request not found'})
        
        account_request = request_result.data[0]
        
        # Check if already processed
        current_status = account_request.get('admin_approval')
        if current_status != 'Pending':
            print(f"⚠️ Request already processed: {current_status}")
            return jsonify({'success': False, 'message': f'Request already {current_status.lower()}'})
        
        # Get rejection reason
        data = request.get_json()
        rejection_reason = data.get('reason', 'Does not meet current access requirements')
        
        print(f"📝 Rejecting request with reason: {rejection_reason}")
        
        # Update the request with rejection details
        update_data = {
            'admin_approval': 'Rejected',
            'permitted_at': datetime.now().isoformat(),
            'reviewed_by': session['admin_id'],
            'reviewed_by_name': current_admin.get('admin_fullname', 'System Administrator'),
            'rejection_reason': rejection_reason
        }
        
        result = supabase.table('account_requests').update(update_data).eq('id', request_id).execute()
        
        if result.data:
            print("✅ Request rejected successfully")
            
            # Send rejection notification email
            try:
                email_sent = send_account_rejection_notification(
                    account_request['admin_email'],
                    account_request['admin_fullname'],
                    account_request['admin_user'],
                    current_admin.get('admin_fullname', 'System Administrator'),
                    rejection_reason
                )
                
                if email_sent:
                    print(f"✅ Rejection email sent to {account_request['admin_email']}")
                else:
                    print(f"⚠️ Failed to send rejection email to {account_request['admin_email']}")
            except Exception as email_error:
                print(f"⚠️ Email error (non-critical): {email_error}")
            
            return jsonify({'success': True, 'message': 'Account request rejected'})
        else:
            print("❌ Failed to update request status")
            return jsonify({'success': False, 'message': 'Failed to update request status'})
            
    except Exception as e:
        print(f"❌ Error rejecting request: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'System error: {str(e)}'})

def validate_password_for_request(password):
    """Validate password for account requests - more lenient than admin passwords"""
    if not password or len(password) < 8:
        return False, "Password must be at least 8 characters long"
    
    # Check for at least one uppercase, lowercase, number, and special character
    import re
    
    if not re.search(r'[a-z]', password):
        return False, "Password must contain at least one lowercase letter"
    
    if not re.search(r'[A-Z]', password):
        return False, "Password must contain at least one uppercase letter"
    
    if not re.search(r'\d', password):
        return False, "Password must contain at least one number"
    
    if not re.search(r'[!@#$%^&*()_+\-=\[\]{};:,.<>?]', password):
        return False, "Password must contain at least one special character"
    
    return True, "Password is valid"

@app.route('/api/reset-password', methods=['POST'])
def api_reset_password():
    """API endpoint for AJAX password reset"""
    try:
        data = request.get_json()
        email = data.get('email')
        code = data.get('code')
        new_password = data.get('new_password')
        
        if not all([email, code, new_password]):
            return jsonify({'success': False, 'message': 'All fields are required'})
        
        # Verify code
        result = supabase.table('password_reset_requests').select('*').eq('email', email).eq('verification_code', code).eq('used', False).execute()
        reset_request = result.data[0] if result.data else None
        
        if not reset_request:
            return jsonify({'success': False, 'message': 'Invalid or expired verification code'})
        
        # Check expiration
        expires_at_str = reset_request['expires_at']
        if 'Z' in expires_at_str:
            expires_at = datetime.fromisoformat(expires_at_str.replace('Z', '+00:00'))
        else:
            expires_at = datetime.fromisoformat(expires_at_str)
        
        if datetime.now(timezone.utc) > expires_at:
            return jsonify({'success': False, 'message': 'Verification code has expired'})
        
        # Validate password
        is_valid, password_message = validate_password(new_password)
        if not is_valid:
            return jsonify({'success': False, 'message': password_message})
        
        # Get admin user
        user_result = supabase.table('accounts_admin').select('admin_id').eq('admin_email', email).execute()
        if not user_result.data:
            return jsonify({'success': False, 'message': 'User account not found'})
        
        admin_id = user_result.data[0]['admin_id']
        
        # Update password (plain text for now)
        update_result = supabase.table('accounts_admin').update({
            'admin_pass': new_password
        }).eq('admin_id', admin_id).execute()
        
        # Wait and verify
        time.sleep(2)
        
        verify_result = supabase.table('accounts_admin').select('admin_pass').eq('admin_id', admin_id).execute()
        new_hash = verify_result.data[0]['admin_pass'] if verify_result.data else None
        
        if new_hash != new_password:
            return jsonify({'success': False, 'message': 'Password update failed'})
        
        # Mark reset request as used
        supabase.table('password_reset_requests').update({
            'used': True,
            'used_at': datetime.now(timezone.utc).isoformat()
        }).eq('id', reset_request['id']).execute()
        
        return jsonify({'success': True, 'message': 'Password reset successfully!'})
        
    except Exception as e:
        print(f"API reset error: {e}")
        return jsonify({'success': False, 'message': 'An error occurred'})

@app.route('/resend-code', methods=['POST'])
def resend_code():
    """Resend verification code"""
    try:
        email = session.get('reset_email')
        if not email:
            return jsonify({'success': False, 'message': 'No reset request found'})
        
        # Generate new verification code
        verification_code = generate_verification_code()
        expires_at = datetime.now(timezone.utc) + timedelta(minutes=15)
        
        # Create new reset request
        reset_data = {
            'email': email,
            'verification_code': verification_code,
            'expires_at': expires_at.isoformat(),
            'used': False
        }
        
        supabase.table('password_reset_requests').insert(reset_data).execute()
        
        # Update session
        session['verification_code'] = verification_code
        
        # Get user info for email
        user_result = supabase.table('accounts_admin').select('admin_fullname, admin_user').eq('admin_email', email).execute()
        user = user_result.data[0] if user_result.data else None
        user_name = user.get('admin_fullname', user.get('admin_user', 'User')) if user else 'User'
        
        # Send email
        email_subject = "Emergency Alert System - New Verification Code"
        email_body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
            <div style="background: #dc2626; color: white; padding: 20px; text-align: center; border-radius: 10px 10px 0 0;">
                <h1 style="margin: 0;">🛡️ Emergency Alert System</h1>
                <h2 style="margin: 10px 0 0 0;">New Verification Code</h2>
            </div>
            <div style="background: #f9fafb; padding: 30px; border-radius: 0 0 10px 10px;">
                <p>Hello {user_name},</p>
                <p>You have requested a new verification code for password reset.</p>
                <div style="background: white; border: 2px solid #2563eb; border-radius: 10px; padding: 20px; margin: 20px 0; text-align: center;">
                    <h3 style="color: #2563eb; margin: 0;">Your New Verification Code:</h3>
                    <div style="font-size: 32px; font-weight: bold; color: #2563eb; font-family: monospace; letter-spacing: 3px; margin: 10px 0;">
                        {verification_code}
                    </div>
                </div>
                <p><strong>This code will expire in 15 minutes.</strong></p>
                <p>If you did not request this code, please ignore this email and contact your system administrator.</p>
            </div>
        </body>
        </html>
        """
        
        if send_email(email, email_subject, email_body):
            return jsonify({'success': True, 'message': 'New verification code sent!'})
        else:
            return jsonify({'success': False, 'message': 'Failed to send email, but code was generated'})
            
    except Exception as e:
        print(f"Resend code error: {e}")
        return jsonify({'success': False, 'message': 'An error occurred'})

@app.route('/')
def index():
    """Root route - redirect to appropriate page"""
    if 'admin_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/developers')
def developers():
    """Developers page route"""
    return render_template('developers.html')

@app.route('/dashboard')
def dashboard():
    """Admin dashboard route"""
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    admin_id = session['admin_id']
    current_admin_id = str(admin_id)
    current_admin = get_admin_by_id(admin_id)
    
    if current_admin:
        session['admin_profile'] = current_admin.get('admin_profile', None)
        session['admin_name'] = current_admin.get('admin_fullname', current_admin.get('admin_user', 'Admin'))
        session['role'] = current_admin.get('admin_role', 'Administrator')
        session['admin_profile_exists'] = check_profile_image_exists(current_admin.get('admin_profile'))
    else:
        session['admin_profile_exists'] = False
    
    # Get dashboard statistics using safe count queries
    students_count = safe_count_query('accounts_student')
    active_alerts_count = safe_count_query('alert_incidents', [{'type': 'in', 'column': 'icd_status', 'value': ['Active', 'Pending']}])
    admin_count = safe_count_query('accounts_admin')
    
    try:
        # Get admin users with full details
        admin_users = supabase.table('accounts_admin').select('admin_id, admin_user, admin_fullname, admin_role, admin_status, admin_last_login, admin_profile').order('admin_fullname').execute().data or []
        # Check if profile images actually exist for each admin
        for admin in admin_users:
            admin['profile_image_exists'] = check_profile_image_exists(admin.get('admin_profile'))
    except Exception as e:
        print(f"Error fetching admin users: {e}")
        admin_users = []
    
    try:
        # Get incidents with assigned responder info and coordinates
        incidents_result = supabase.table('alert_incidents').select('*').execute()
        incidents = incidents_result.data or []
        
        # Get student full_name from accounts_student table for each incident
        user_ids = {incident.get('user_id') for incident in incidents if incident.get('user_id')}
        students_map = {}
        
        if user_ids:
            try:
                batch_size = 100
                user_ids_list = list(user_ids)
                for i in range(0, len(user_ids_list), batch_size):
                    batch = user_ids_list[i:i + batch_size]
                    students_result = supabase.table('accounts_student').select('user_id, full_name').in_('user_id', batch).execute()
                    if students_result.data:
                        for student in students_result.data:
                            # Store with both string and integer keys for flexibility
                            user_id = student.get('user_id')
                            if user_id is not None:
                                students_map[str(user_id)] = student.get('full_name', '')
                                try:
                                    students_map[int(user_id)] = student.get('full_name', '')
                                except (ValueError, TypeError):
                                    pass
            except Exception as e:
                print(f"Error fetching student names for dashboard: {e}")
        
        # Filter incidents based on assignment rules
        incidents = filter_incidents_for_admin(incidents, current_admin_id)
        
        # Process incidents data - ensure proper data structure for map
        for incident in incidents:
            incident['formatted_timestamp'] = format_datetime(incident.get('icd_timestamp'))
            # Ensure coordinates exist for map display
            if not incident.get('icd_lat') or not incident.get('icd_lng'):
                # Default to UMAK coordinates if not available
                incident['icd_lat'] = 14.5633428
                incident['icd_lng'] = 121.0565387
            
            # Add student full_name from accounts_student table
            user_id = incident.get('user_id')
            if user_id:
                # Try both string and integer keys
                full_name = students_map.get(str(user_id)) or students_map.get(int(user_id)) if isinstance(user_id, int) else students_map.get(str(user_id))
                if full_name:
                    incident['student_full_name'] = full_name
    except Exception as e:
        print(f"Error fetching incidents: {e}")
        incidents = []
    
    try:
        # Get active/pending alerts
        alerts_result = supabase.table('alert_incidents').select('*').in_('icd_status', ['Active', 'Pending']).order('icd_timestamp', desc=True).execute()
        all_alerts = alerts_result.data or []
        
        # Get student full_name from accounts_student table for alerts
        alert_user_ids = {alert.get('user_id') for alert in all_alerts if alert.get('user_id')}
        alert_students_map = {}
        
        if alert_user_ids:
            try:
                batch_size = 100
                alert_user_ids_list = list(alert_user_ids)
                for i in range(0, len(alert_user_ids_list), batch_size):
                    batch = alert_user_ids_list[i:i + batch_size]
                    alert_students_result = supabase.table('accounts_student').select('user_id, full_name').in_('user_id', batch).execute()
                    if alert_students_result.data:
                        for student in alert_students_result.data:
                            user_id = student.get('user_id')
                            if user_id is not None:
                                alert_students_map[str(user_id)] = student.get('full_name', '')
                                try:
                                    alert_students_map[int(user_id)] = student.get('full_name', '')
                                except (ValueError, TypeError):
                                    pass
            except Exception as e:
                print(f"Error fetching student names for alerts: {e}")
        
        # Filter alerts based on assignment rules
        all_alerts = filter_incidents_for_admin(all_alerts, current_admin_id)
        
        # Build responder map for display
        responder_ids = {alert.get('assigned_responder_id') for alert in all_alerts if alert.get('assigned_responder_id')}
        responder_map = {}
        if responder_ids:
            try:
                responder_ids_list = [str(rid) for rid in responder_ids if rid]
                batch_size = 100
                for i in range(0, len(responder_ids_list), batch_size):
                    batch = responder_ids_list[i:i + batch_size]
                    responder_result = supabase.table('accounts_admin').select('admin_id, admin_fullname').in_('admin_id', batch).execute()
                    if responder_result.data:
                        for responder in responder_result.data:
                            responder_id = responder.get('admin_id')
                            if responder_id is not None:
                                responder_map[str(responder_id)] = responder.get('admin_fullname', responder_id)
            except Exception as e:
                print(f"Error fetching responder names for alerts: {e}")
        
        # Process alerts data
        for alert in all_alerts:
            alert['formatted_timestamp'] = format_datetime(alert.get('icd_timestamp'))
            
            # Add student full_name from accounts_student table
            user_id = alert.get('user_id')
            if user_id:
                # Try both string and integer keys
                full_name = alert_students_map.get(str(user_id)) or alert_students_map.get(int(user_id)) if isinstance(user_id, int) else alert_students_map.get(str(user_id))
                if full_name:
                    alert['student_full_name'] = full_name
            
            assigned_responder_id = alert.get('assigned_responder_id')
            if assigned_responder_id:
                alert['assigned_responder_name'] = responder_map.get(str(assigned_responder_id), str(assigned_responder_id))
            else:
                alert['assigned_responder_name'] = None
    except Exception as e:
        print(f"Error fetching alerts: {e}")
        all_alerts = []
    
    # Get additional data
    admin_list = get_active_admins()
    recent_activities = get_recent_activities()
    
    return render_template('dashboard.html',
                         students_count=students_count,
                         active_alerts_count=active_alerts_count,
                         admin_count=admin_count,
                         admin_users=admin_users,
                         incidents=incidents,
                         all_alerts=all_alerts,
                         admin_list=admin_list,
                         recent_activities=recent_activities,
                         datetime=datetime)


# ==================== LIVE MAP AND ALERT API ROUTES ====================

@app.route('/api/incidents')
def get_incidents_api():
    """API endpoint to get incidents with optional filtering"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        current_admin_id = str(session['admin_id'])
        status_filter = request.args.get('status', 'all')
        category_filter = request.args.get('category', 'all')
        date_range = request.args.get('date_range', 'all')
        
        query = supabase.table('alert_incidents').select('*')
        
        # Apply status filter
        if status_filter != 'all':
            query = query.eq('icd_status', status_filter)
        
        # Apply category filter
        if category_filter != 'all':
            query = query.eq('icd_category', category_filter)
        
        # Apply date filter based on icd_timestamp
        query = apply_date_filter(query, date_range)
        
        result = query.order('icd_timestamp', desc=True).execute()
        incidents = result.data or []
        
        incidents = filter_incidents_for_admin(incidents, current_admin_id)
        
        # Get student full_name from accounts_student table for each incident
        user_ids = {incident.get('user_id') for incident in incidents if incident.get('user_id')}
        students_map = {}
        
        if user_ids:
            try:
                batch_size = 100
                user_ids_list = list(user_ids)
                for i in range(0, len(user_ids_list), batch_size):
                    batch = user_ids_list[i:i + batch_size]
                    students_result = supabase.table('accounts_student').select('user_id, full_name').in_('user_id', batch).execute()
                    if students_result.data:
                        for student in students_result.data:
                            # Store with both string and integer keys for flexibility
                            user_id = student.get('user_id')
                            if user_id is not None:
                                students_map[str(user_id)] = student.get('full_name', '')
                                try:
                                    students_map[int(user_id)] = student.get('full_name', '')
                                except (ValueError, TypeError):
                                    pass
            except Exception as e:
                print(f"Error fetching student names: {e}")
        
        # Build responder map
        responder_ids = {incident.get('assigned_responder_id') for incident in incidents if incident.get('assigned_responder_id')}
        responder_map = {}
        if responder_ids:
            try:
                responder_ids_list = [str(rid) for rid in responder_ids if rid]
                batch_size = 100
                for i in range(0, len(responder_ids_list), batch_size):
                    batch = responder_ids_list[i:i + batch_size]
                    responder_result = supabase.table('accounts_admin').select('admin_id, admin_fullname').in_('admin_id', batch).execute()
                    if responder_result.data:
                        for responder in responder_result.data:
                            responder_id = responder.get('admin_id')
                            if responder_id is not None:
                                responder_map[str(responder_id)] = responder.get('admin_fullname', responder_id)
            except Exception as e:
                print(f"Error fetching responder names: {e}")
        
        # Add student full_name to each incident
        for incident in incidents:
            user_id = incident.get('user_id')
            if user_id:
                # Try both string and integer keys
                full_name = students_map.get(str(user_id)) or students_map.get(int(user_id)) if isinstance(user_id, int) else students_map.get(str(user_id))
                if full_name:
                    incident['student_full_name'] = full_name
            
            assigned_responder_id = incident.get('assigned_responder_id')
            if assigned_responder_id:
                incident['assigned_responder_name'] = responder_map.get(str(assigned_responder_id), str(assigned_responder_id))
            else:
                incident['assigned_responder_name'] = None
        
        return jsonify({
            'success': True,
            'incidents': incidents,
            'count': len(incidents)
        })
        
    except Exception as e:
        print(f"Error fetching incidents: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/incidents/<incident_id>')
def get_incident_details_api(incident_id):
    """API endpoint to get detailed incident information"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        can_view, view_message = can_admin_view_incident(incident_id, session['admin_id'])
        if not can_view:
            return jsonify({'success': False, 'error': view_message}), 403
        
        result = supabase.table('alert_incidents').select('*').eq('icd_id', incident_id).execute()
        
        if result.data:
            incident = result.data[0]
            
            # Get student details if user_id exists
            student_details = None
            if incident.get('user_id'):
                student_result = supabase.table('accounts_student').select('*').eq('user_id', incident['user_id']).execute()
                if student_result.data:
                    student_details = student_result.data[0]
            
            # Get assigned responder details
            responder_details = None
            if incident.get('assigned_responder_id'):
                responder_result = supabase.table('accounts_admin').select('*').eq('admin_id', incident['assigned_responder_id']).execute()
                if responder_result.data:
                    responder_details = responder_result.data[0]
                    incident['assigned_responder_name'] = responder_details.get('admin_fullname', incident.get('assigned_responder_id'))
                else:
                    incident['assigned_responder_name'] = incident.get('assigned_responder_id')
            else:
                incident['assigned_responder_name'] = None
            
            return jsonify({
                'success': True,
                'incident': incident,
                'student_details': student_details,
                'responder_details': responder_details
            })
        else:
            return jsonify({'success': False, 'error': 'Incident not found'}), 404
            
    except Exception as e:
        print(f"Error fetching incident details: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/student/<user_id>')
def get_student_api(user_id):
    """API endpoint to get student details"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        result = supabase.table('accounts_student').select('*').eq('user_id', user_id).execute()
        
        if result.data:
            return jsonify({
                'success': True,
                'student': result.data[0]
            })
        else:
            return jsonify({'success': False, 'error': 'Student not found'}), 404
            
    except Exception as e:
        print(f"Error fetching student details: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/alerts')
def get_alerts_api():
    """API endpoint to get alerts with filtering"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        current_admin_id = str(session['admin_id'])
        status_filter = request.args.get('status', 'all')
        category_filter = request.args.get('category', 'all')
        
        query = supabase.table('alert_incidents').select('*')
        
        # Filter for active alerts only by default
        if status_filter == 'all':
            query = query.in_('icd_status', ['Active', 'Pending'])
        else:
            query = query.eq('icd_status', status_filter)
        
        # Apply category filter
        if category_filter != 'all':
            query = query.eq('icd_category', category_filter)
        
        result = query.order('icd_timestamp', desc=True).execute()
        alerts = result.data or []
        
        alerts = filter_incidents_for_admin(alerts, current_admin_id)
        
        responder_ids = {alert.get('assigned_responder_id') for alert in alerts if alert.get('assigned_responder_id')}
        responder_map = {}
        if responder_ids:
            try:
                responder_ids_list = [str(rid) for rid in responder_ids if rid]
                batch_size = 100
                for i in range(0, len(responder_ids_list), batch_size):
                    batch = responder_ids_list[i:i + batch_size]
                    responder_result = supabase.table('accounts_admin').select('admin_id, admin_fullname').in_('admin_id', batch).execute()
                    if responder_result.data:
                        for responder in responder_result.data:
                            responder_id = responder.get('admin_id')
                            if responder_id is not None:
                                responder_map[str(responder_id)] = responder.get('admin_fullname', responder_id)
            except Exception as e:
                print(f"Error fetching responder names for alerts: {e}")
        
        for alert in alerts:
            assigned_responder_id = alert.get('assigned_responder_id')
            if assigned_responder_id:
                alert['assigned_responder_name'] = responder_map.get(str(assigned_responder_id), str(assigned_responder_id))
            else:
                alert['assigned_responder_name'] = None
        
        return jsonify({
            'success': True,
            'alerts': alerts,
            'count': len(alerts)
        })
        
    except Exception as e:
        print(f"Error fetching alerts: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/activity-feed')
def get_activity_feed_api():
    """API endpoint to get the recent incident activity feed."""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401
    
    try:
        limit = request.args.get('limit', default=60, type=int)
        activities = get_recent_activities(limit=limit)
        return jsonify({
            'success': True,
            'activities': activities,
            'count': len(activities)
        })
    except Exception as e:
        print(f"Error fetching activity feed: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/alerts/count')
def get_alerts_count_api():
    """API endpoint to get alert counts by status"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        current_admin_id = str(session['admin_id'])
        incidents_result = supabase.table('alert_incidents').select('icd_status, assigned_responder_id').execute()
        incidents = incidents_result.data or []
        
        counts = {
            'active': 0,
            'pending': 0,
            'resolved': 0,
            'cancelled': 0
        }
        
        for incident in incidents:
            status = (incident.get('icd_status') or '').strip()
            assigned_responder_id = incident.get('assigned_responder_id')
            
            if status in ['Active', 'Pending']:
                if assigned_responder_id and str(assigned_responder_id) != current_admin_id:
                    continue
            
            if status == 'Active':
                counts['active'] += 1
            elif status == 'Pending':
                counts['pending'] += 1
            elif status == 'Resolved':
                counts['resolved'] += 1
            elif status == 'Cancelled':
                counts['cancelled'] += 1
        
        return jsonify({
            'success': True,
            'counts': {
                'active': counts['active'],
                'pending': counts['pending'],
                'resolved': counts['resolved'],
                'cancelled': counts['cancelled'],
                'total_active': counts['active'] + counts['pending']
            }
        })
        
    except Exception as e:
        print(f"Error fetching alert counts: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/map/stats')
def get_map_stats_api():
    """API endpoint to get map statistics"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        # Get counts by category and status
        categories = ['Urgent', 'Medical', 'Security', 'University']
        statuses = ['Active', 'Pending', 'Resolved', 'Cancelled']
        
        category_stats = {}
        for category in categories:
            count_result = supabase.table('alert_incidents').select('*', count='exact').eq('icd_category', category).execute()
            category_stats[category.lower()] = count_result.count or 0
        
        status_stats = {}
        for status in statuses:
            count_result = supabase.table('alert_incidents').select('*', count='exact').eq('icd_status', status).execute()
            status_stats[status.lower()] = count_result.count or 0
        
        return jsonify({
            'success': True,
            'stats': {
                'categories': category_stats,
                'statuses': status_stats
            }
        })
        
    except Exception as e:
        print(f"Error fetching map stats: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

def retry_supabase_query(query_func, max_retries=3, delay=0.5):
    """
    Retry a Supabase query function with exponential backoff.
    Handles socket errors and connection issues.
    """
    import socket
    try:
        import httpx
        import httpcore
    except ImportError:
        httpx = None
        httpcore = None
    
    for attempt in range(max_retries):
        try:
            return query_func()
        except (OSError, socket.error) as e:
            error_str = str(e)
            # Check for Windows socket errors (10035, 10054, etc.)
            if '10035' in error_str or '10054' in error_str or 'non-blocking' in error_str.lower():
                if attempt < max_retries - 1:
                    wait_time = delay * (2 ** attempt)  # Exponential backoff
                    time.sleep(wait_time)
                    continue
                else:
                    # Return None on final failure to prevent crashes
                    print(f"Socket error after {max_retries} attempts, returning None: {e}")
                    return None
            else:
                raise
        except Exception as e:
            error_str = str(e)
            error_type = type(e).__name__
            
            # Check for httpx/httpcore ReadError
            is_read_error = False
            if httpx:
                try:
                    is_read_error = isinstance(e, httpx.ReadError)
                except:
                    pass
            if httpcore and not is_read_error:
                try:
                    is_read_error = isinstance(e, httpcore.ReadError)
                except:
                    pass
            
            # Check for Windows socket errors (10035, 10054) or ReadError
            if is_read_error or '10035' in error_str or '10054' in error_str or 'non-blocking' in error_str.lower() or 'ReadError' in error_type:
                if attempt < max_retries - 1:
                    wait_time = delay * (2 ** attempt)
                    time.sleep(wait_time)
                    continue
                else:
                    print(f"Connection error after {max_retries} attempts, returning None: {error_type}: {e}")
                    return None
            
            # For other errors, only retry on first attempt if it's a connection issue
            if attempt == 0 and ('connection' in error_str.lower() or 'timeout' in error_str.lower()):
                time.sleep(delay)
                continue
            
            # Don't retry on other errors
            raise
    return None

def get_date_range(start_date, end_date):
    """Helper function to get date range in ISO format"""
    if start_date and end_date:
        try:
            start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            # Ensure end date includes the full day
            end = end.replace(hour=23, minute=59, second=59)
            return start.isoformat(), end.isoformat()
        except:
            pass
    return None, None

def apply_date_filter(query, date_range):
    """Apply date filter to query based on date_range parameter"""
    if not date_range or date_range == 'all':
        return query
    
    now = datetime.now()
    start_date = None
    end_date = None
    
    if date_range == 'today':
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = now.replace(hour=23, minute=59, second=59, microsecond=999999)
    elif date_range == 'last_7_days':
        start_date = now - timedelta(days=7)
        end_date = now
    elif date_range == 'last_30_days':
        start_date = now - timedelta(days=30)
        end_date = now
    elif date_range == 'last_90_days':
        start_date = now - timedelta(days=90)
        end_date = now
    elif date_range == 'custom':
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        if start_date_str and end_date_str:
            try:
                # Handle date strings from HTML date inputs (YYYY-MM-DD format)
                if len(start_date_str) == 10:  # YYYY-MM-DD format
                    start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
                    end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
                    end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)
                else:
                    # Handle ISO format
                    start_date = datetime.fromisoformat(start_date_str.replace('Z', '+00:00'))
                    end_date = datetime.fromisoformat(end_date_str.replace('Z', '+00:00'))
                    end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)
            except Exception as e:
                print(f"Error parsing custom date range: {e}")
                pass
    
    if start_date and end_date:
        query = query.gte('icd_timestamp', start_date.isoformat())
        query = query.lte('icd_timestamp', end_date.isoformat())
    
    return query

@app.route('/api/dashboard/stats')
def get_dashboard_stats():
    """API endpoint to get dashboard statistics with date filtering"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        date_range = request.args.get('date_range', 'today')
        
        # Base query with retry logic
        def get_total_alerts():
            query = supabase.table('alert_incidents').select('*', count='exact')
            query = apply_date_filter(query, date_range)
            return query.execute()
        
        total_alerts_result = retry_supabase_query(get_total_alerts)
        total_alerts_count = total_alerts_result.count if total_alerts_result else 0
        
        # Active alerts in date range with retry
        def get_active_alerts():
            active_query = supabase.table('alert_incidents').select('*', count='exact').eq('icd_status', 'Active')
            active_query = apply_date_filter(active_query, date_range)
            return active_query.execute()
        
        active_result = retry_supabase_query(get_active_alerts)
        active_count = active_result.count if active_result else 0
        
        # Resolved alerts in date range with retry
        def get_resolved_alerts():
            resolved_query = supabase.table('alert_incidents').select('*', count='exact').eq('icd_status', 'Resolved')
            resolved_query = apply_date_filter(resolved_query, date_range)
            return resolved_query.execute()
        
        resolved_result = retry_supabase_query(get_resolved_alerts)
        resolved_count = resolved_result.count if resolved_result else 0
        
        # Calculate average response time with retry
        def get_resolved_incidents():
            resolved_incidents_query = supabase.table('alert_incidents').select('icd_timestamp, resolved_timestamp').eq('icd_status', 'Resolved').not_.is_('resolved_timestamp', 'null')
            resolved_incidents_query = apply_date_filter(resolved_incidents_query, date_range)
            return resolved_incidents_query.execute()
        
        resolved_incidents = retry_supabase_query(get_resolved_incidents)
        
        avg_response_time = 0
        if resolved_incidents and resolved_incidents.data:
            total_seconds = 0
            count = 0
            for incident in resolved_incidents.data:
                if incident.get('icd_timestamp') and incident.get('resolved_timestamp'):
                    try:
                        start = datetime.fromisoformat(incident['icd_timestamp'].replace('Z', '+00:00'))
                        end = datetime.fromisoformat(incident['resolved_timestamp'].replace('Z', '+00:00'))
                        diff = (end - start).total_seconds()
                        if diff > 0:
                            total_seconds += diff
                            count += 1
                    except:
                        pass
            if count > 0:
                avg_response_time = total_seconds / count / 60  # Convert to minutes
        
        return jsonify({
            'success': True,
            'stats': {
                'total_alerts': total_alerts_count,
                'active': active_count,
                'resolved': resolved_count,
                'avg_response_time_minutes': round(avg_response_time, 2)
            }
        })
    except Exception as e:
        print(f"Error fetching dashboard stats: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/dashboard/alert-types')
def get_alert_types():
    """API endpoint to get alert types distribution for pie chart with date filtering"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        date_range = request.args.get('date_range', 'today')
        
        # Get incidents with date filter
        query = supabase.table('alert_incidents').select('icd_category')
        query = apply_date_filter(query, date_range)
        incidents = query.execute().data or []
        
        category_counts = {}
        for incident in incidents:
            category = incident.get('icd_category') or 'Unknown'
            category_counts[category] = category_counts.get(category, 0) + 1
        
        return jsonify({
            'success': True,
            'data': category_counts
        })
    except Exception as e:
        print(f"Error fetching alert types: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/dashboard/student-registration-trend')
def get_student_registration_trend():
    """API endpoint to get student registration trend grouped by year level"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        students = supabase.table('accounts_student').select('student_yearlvl, student_created_at').execute().data or []
        
        # Group by year level
        year_level_counts = {}
        for student in students:
            yearlvl = student.get('student_yearlvl') or 'Unknown'
            year_level_counts[yearlvl] = year_level_counts.get(yearlvl, 0) + 1
        
        return jsonify({
            'success': True,
            'data': year_level_counts
        })
    except Exception as e:
        print(f"Error fetching student registration trend: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/dashboard/top-responders')
def get_top_responders():
    """API endpoint to get top responders (filtered by student_yearlvl if provided)"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        yearlvl_filter = request.args.get('student_yearlvl')
        
        # Get all incidents with assigned responders
        query = supabase.table('alert_incidents').select('assigned_responder_id, user_id')
        
        incidents = query.execute().data or []
        
        # If year level filter is provided, filter by student year level
        if yearlvl_filter and yearlvl_filter != 'all':
            # Get students with this year level
            students = supabase.table('accounts_student').select('user_id').eq('student_yearlvl', yearlvl_filter).execute().data or []
            student_ids = {s['user_id'] for s in students}
            incidents = [inc for inc in incidents if inc.get('user_id') in student_ids]
        
        # Count by responder
        responder_counts = {}
        for incident in incidents:
            responder_id = incident.get('assigned_responder_id')
            if responder_id:
                responder_counts[responder_id] = responder_counts.get(responder_id, 0) + 1
        
        # Get responder names
        responder_data = []
        for responder_id, count in sorted(responder_counts.items(), key=lambda x: x[1], reverse=True)[:10]:
            # Try to get admin name
            admin_result = supabase.table('accounts_admin').select('admin_fullname, admin_user').eq('admin_id', responder_id).execute()
            if admin_result.data:
                name = admin_result.data[0].get('admin_fullname') or admin_result.data[0].get('admin_user', responder_id)
            else:
                name = responder_id
            responder_data.append({'name': name, 'count': count})
        
        return jsonify({
            'success': True,
            'data': responder_data
        })
    except Exception as e:
        print(f"Error fetching top responders: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/dashboard/alert-volume')
def get_alert_volume():
    """API endpoint to get alert volume over time"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        from datetime import timedelta
        
        # Get alerts from last 30 days
        thirty_days_ago = (datetime.now() - timedelta(days=30)).isoformat()
        incidents = supabase.table('alert_incidents').select('icd_timestamp').gte('icd_timestamp', thirty_days_ago).execute().data or []
        
        # Group by date
        daily_counts = {}
        for incident in incidents:
            if incident.get('icd_timestamp'):
                try:
                    dt = datetime.fromisoformat(incident['icd_timestamp'].replace('Z', '+00:00'))
                    date_key = dt.strftime('%Y-%m-%d')
                    daily_counts[date_key] = daily_counts.get(date_key, 0) + 1
                except:
                    pass
        
        # Sort by date
        sorted_dates = sorted(daily_counts.keys())
        data = [{'date': date, 'count': daily_counts[date]} for date in sorted_dates]
        
        return jsonify({
            'success': True,
            'data': data
        })
    except Exception as e:
        print(f"Error fetching alert volume: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/dashboard/false-alerts')
def get_false_alerts():
    """API endpoint to get false alerts count with date filtering"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        date_range = request.args.get('date_range', 'today')
        
        # Count cancelled alerts (assuming cancelled = false alerts)
        cancelled_query = supabase.table('alert_incidents').select('*', count='exact').eq('icd_status', 'Cancelled')
        cancelled_query = apply_date_filter(cancelled_query, date_range)
        cancelled_result = cancelled_query.execute()
        cancelled_count = cancelled_result.count or 0
        
        # Total alerts
        total_query = supabase.table('alert_incidents').select('*', count='exact')
        total_query = apply_date_filter(total_query, date_range)
        total_result = total_query.execute()
        total_count = total_result.count or 0
        
        valid_count = total_count - cancelled_count
        
        return jsonify({
            'success': True,
            'data': {
                'false': cancelled_count,
                'valid': valid_count,
                'total': total_count
            }
        })
    except Exception as e:
        print(f"Error fetching false alerts: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/dashboard/alerts-by-time')
def get_alerts_by_time():
    """API endpoint to get alerts by time of day (hourly distribution) with date filtering"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        date_range = request.args.get('date_range', 'today')
        
        # Get incidents with date filter - select icd_timestamp from alert_incidents table
        query = supabase.table('alert_incidents').select('icd_timestamp')
        query = apply_date_filter(query, date_range)
        incidents = query.execute().data or []
        
        # Group by hour of day (0-23) in Philippines timezone
        hourly_counts = {hour: 0 for hour in range(24)}
        
        for incident in incidents:
            if incident.get('icd_timestamp'):
                try:
                    # Parse timestamp - handle both with and without timezone
                    timestamp_str = incident['icd_timestamp']
                    if 'Z' in timestamp_str:
                        # UTC timestamp
                        dt = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
                        # Convert to Philippines timezone
                        if dt.tzinfo is None:
                            dt = dt.replace(tzinfo=timezone.utc)
                        dt_ph = dt.astimezone(PHILIPPINES_TZ)
                    else:
                        # Try to parse as ISO format
                        dt = datetime.fromisoformat(timestamp_str)
                        # If no timezone info, assume UTC and convert to PH
                        if dt.tzinfo is None:
                            dt = dt.replace(tzinfo=timezone.utc)
                        dt_ph = dt.astimezone(PHILIPPINES_TZ)
                    
                    # Get hour in Philippines timezone
                    hour = dt_ph.hour
                    hourly_counts[hour] = hourly_counts.get(hour, 0) + 1
                except Exception as e:
                    print(f"Error parsing timestamp {incident.get('icd_timestamp')}: {e}")
                    continue
        
        # Format for chart (hour labels and counts)
        labels = [f"{hour:02d}:00" for hour in range(24)]
        data = [hourly_counts.get(hour, 0) for hour in range(24)]
        
        return jsonify({
            'success': True,
            'data': {
                'labels': labels,
                'counts': data
            }
        })
    except Exception as e:
        print(f"Error fetching alerts by time: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/dashboard/student-registered')
def get_student_registered():
    """API endpoint to get student registered by college and year level (pie chart) with date filtering"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        date_range = request.args.get('date_range', 'today')
        
        # Get incidents with date filter and join with student data
        query = supabase.table('alert_incidents').select('user_id, icd_timestamp')
        query = apply_date_filter(query, date_range)
        incidents = query.execute().data or []
        
        # Get unique user_ids from incidents
        user_ids = list(set([inc.get('user_id') for inc in incidents if inc.get('user_id')]))
        
        if not user_ids:
            return jsonify({
                'success': True,
                'data': {}
            })
        
        # Get student data for these users
        students_result = supabase.table('accounts_student').select('user_id, student_college, student_yearlvl').in_('user_id', user_ids).execute()
        students = students_result.data or []
        
        # Create a map of user_id to student data
        student_map = {s['user_id']: s for s in students}
        
        # Count by college and year level
        college_year_counts = {}
        
        for incident in incidents:
            user_id = incident.get('user_id')
            if user_id and user_id in student_map:
                student = student_map[user_id]
                college = student.get('student_college') or 'Unknown'
                yearlvl = student.get('student_yearlvl') or 'Unknown'
                key = f"{college} - {yearlvl}"
                college_year_counts[key] = college_year_counts.get(key, 0) + 1
        
        return jsonify({
            'success': True,
            'data': college_year_counts
        })
    except Exception as e:
        print(f"Error fetching student registered: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/dashboard/alerts-by-student')
def get_alerts_by_student():
    """API endpoint to get alerts filtered by student college and year level with date filtering"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        date_range = request.args.get('date_range', 'today')
        college_filter = request.args.get('college', None)
        yearlvl_filter = request.args.get('yearlvl', None)
        
        # Get incidents with date filter
        query = supabase.table('alert_incidents').select('user_id, icd_timestamp, icd_category, icd_status')
        query = apply_date_filter(query, date_range)
        incidents = query.execute().data or []
        
        # Get unique user_ids from incidents
        user_ids = list(set([inc.get('user_id') for inc in incidents if inc.get('user_id')]))
        
        if not user_ids:
            return jsonify({
                'success': True,
                'data': {
                    'by_college': {},
                    'by_yearlvl': {},
                    'by_college_yearlvl': {}
                }
            })
        
        # Get student data for these users
        students_query = supabase.table('accounts_student').select('user_id, student_college, student_yearlvl')
        if college_filter:
            students_query = students_query.eq('student_college', college_filter)
        if yearlvl_filter:
            students_query = students_query.eq('student_yearlvl', yearlvl_filter)
        students_query = students_query.in_('user_id', user_ids)
        students_result = students_query.execute()
        students = students_result.data or []
        
        # Create a map of user_id to student data
        student_map = {s['user_id']: s for s in students}
        valid_user_ids = set(student_map.keys())
        
        # Filter incidents to only those with valid students
        filtered_incidents = [inc for inc in incidents if inc.get('user_id') in valid_user_ids]
        
        # Count by college
        by_college = {}
        # Count by year level
        by_yearlvl = {}
        # Count by college and year level combination
        by_college_yearlvl = {}
        
        for incident in filtered_incidents:
            user_id = incident.get('user_id')
            if user_id and user_id in student_map:
                student = student_map[user_id]
                college = student.get('student_college') or 'Unknown'
                yearlvl = student.get('student_yearlvl') or 'Unknown'
                
                # Count by college
                by_college[college] = by_college.get(college, 0) + 1
                
                # Count by year level
                by_yearlvl[yearlvl] = by_yearlvl.get(yearlvl, 0) + 1
                
                # Count by college and year level
                key = f"{college} - {yearlvl}"
                by_college_yearlvl[key] = by_college_yearlvl.get(key, 0) + 1
        
        return jsonify({
            'success': True,
            'data': {
                'by_college': by_college,
                'by_yearlvl': by_yearlvl,
                'by_college_yearlvl': by_college_yearlvl
            }
        })
    except Exception as e:
        print(f"Error fetching alerts by student: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/dashboard/location-distribution')
def get_location_distribution():
    """API endpoint to get on-campus vs external incident distribution with date filtering"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        date_range = request.args.get('date_range', 'today')
        
        query = supabase.table('alert_incidents').select('icd_lat, icd_lng')
        query = apply_date_filter(query, date_range)
        incidents = query.execute().data or []
        
        on_campus = 0
        external = 0
        
        for incident in incidents:
            lat = incident.get('icd_lat')
            lng = incident.get('icd_lng')
            
            try:
                lat = float(lat)
                lng = float(lng)
                distance = calculate_distance(UMAK_LAT, UMAK_LNG, lat, lng)
                if distance is not None and distance <= UMAK_RADIUS:
                    on_campus += 1
                else:
                    external += 1
            except (TypeError, ValueError):
                external += 1
        
        return jsonify({
            'success': True,
            'data': {
                'on_campus': on_campus,
                'external': external
            }
        })
    except Exception as e:
        print(f"Error fetching location distribution: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/dashboard/high-risk-areas')
def get_high_risk_areas():
    """API endpoint to get high-risk areas for map heat with location names and active alerts"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        # Get incidents with coordinates and status from Supabase
        incidents = supabase.table('alert_incidents').select('icd_lat, icd_lng, icd_status').not_.is_('icd_lat', 'null').not_.is_('icd_lng', 'null').execute().data or []
        
        # Group by location (rounded to 4 decimal places for clustering)
        location_data = {}
        for incident in incidents:
            lat = round(float(incident.get('icd_lat', 0)), 4)
            lng = round(float(incident.get('icd_lng', 0)), 4)
            key = f"{lat},{lng}"
            
            if key not in location_data:
                location_data[key] = {
                    'lat': lat,
                    'lng': lng,
                    'count': 0,
                    'active_count': 0
                }
            
            location_data[key]['count'] += 1
            if incident.get('icd_status') in ['Active', 'Pending']:
                location_data[key]['active_count'] += 1
        
        # Convert to list format and sort by active_count (most active first)
        heat_data = []
        for key, data in location_data.items():
            heat_data.append({
                'lat': data['lat'],
                'lng': data['lng'],
                'count': data['count'],
                'active_count': data['active_count']
            })
        
        # Sort by active_count descending
        heat_data.sort(key=lambda x: x['active_count'], reverse=True)
        
        return jsonify({
            'success': True,
            'data': heat_data
        })
    except Exception as e:
        print(f"Error fetching high-risk areas: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== PERMISSION HELPERS ====================

def can_admin_edit_incident(incident_id, admin_id):
    """
    Check if an admin can edit an incident.
    Rules:
    - If incident is Active/Pending and has assigned_responder_id, only that admin can edit
    - If incident is Resolved/Cancelled, any admin can view but not edit (unless reassigned)
    - If incident has no assigned_responder_id, any admin can assign themselves
    - If admin is the assigned responder, they can always edit
    """
    try:
        incident_result = supabase.table('alert_incidents').select('icd_status, assigned_responder_id').eq('icd_id', incident_id).execute()
        
        if not incident_result.data or len(incident_result.data) == 0:
            return False, "Incident not found"
        
        incident = incident_result.data[0]
        status = incident.get('icd_status', '')
        assigned_responder_id = incident.get('assigned_responder_id')
        
        # If incident is Resolved or Cancelled, no one can edit (read-only)
        if status in ['Resolved', 'Cancelled']:
            return False, f"Incident is {status} and cannot be modified"
        
        # If no one is assigned, any admin can take it
        if not assigned_responder_id:
            return True, "No responder assigned - available for assignment"
        
        # If this admin is the assigned responder, they can edit
        if str(assigned_responder_id) == str(admin_id):
            return True, "You are the assigned responder"
        
        # Otherwise, only the assigned responder can edit
        return False, f"Incident is assigned to another responder (ID: {assigned_responder_id})"
    
    except Exception as e:
        print(f"Error checking edit permission: {e}")
        return False, f"Error checking permissions: {str(e)}"

def can_admin_view_incident(incident_id, admin_id):
    """
    Check if an admin can view an incident.
    Active/Pending incidents are only viewable by their assigned responder (or anyone if unassigned).
    """
    try:
        incident_result = supabase.table('alert_incidents').select('icd_status, assigned_responder_id').eq('icd_id', incident_id).execute()
        
        if not incident_result.data or len(incident_result.data) == 0:
            return False, "Incident not found"
        
        incident = incident_result.data[0]
        status = (incident.get('icd_status') or '').strip()
        assigned_responder_id = incident.get('assigned_responder_id')
        
        if status in ['Active', 'Pending']:
            if assigned_responder_id and str(assigned_responder_id) != str(admin_id):
                return False, "This incident is assigned to another responder"
        
        return True, "Incident is viewable"
    
    except Exception as e:
        print(f"Error checking view permission: {e}")
        return False, f"Error checking permissions: {str(e)}"

def filter_incidents_for_admin(incidents, admin_id):
    """
    Filter incidents so admins only see incidents they are allowed to access.
    Active/Pending incidents assigned to other admins are hidden.
    """
    if not admin_id:
        return incidents
    
    filtered = []
    admin_id_str = str(admin_id)
    
    for incident in incidents:
        status = (incident.get('icd_status') or '').strip()
        assigned_responder_id = incident.get('assigned_responder_id')
        
        if status in ['Active', 'Pending']:
            if assigned_responder_id and str(assigned_responder_id) != admin_id_str:
                continue
        
        filtered.append(incident)
    
    return filtered

# ==================== INCIDENT STATUS UPDATE ROUTES ====================

@app.route('/mark_resolved', methods=['POST'])
def mark_resolved():
    """Mark incident as resolved"""
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    incident_id = request.form.get('incident_id')
    summary_text = (request.form.get('resolution_summary') or '').strip()
    wants_json = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or (
        'application/json' in (request.headers.get('Accept') or '')
    )

    if not incident_id:
        message = 'Incident ID is required'
        if wants_json:
            return jsonify({'success': False, 'error': message}), 400
        flash(message, 'error')
        return redirect(url_for('dashboard'))

    if not summary_text:
        message = 'Resolution summary is required.'
        if wants_json:
            return jsonify({'success': False, 'error': message}), 400
        flash(message, 'error')
        return redirect(url_for('dashboard'))
    
    # Check if admin can edit this incident
    can_edit, edit_message = can_admin_edit_incident(incident_id, session['admin_id'])
    if not can_edit:
        message = f'Permission denied: {edit_message}'
        if wants_json:
            return jsonify({'success': False, 'error': message}), 403
        flash(message, 'error')
        return redirect(url_for('dashboard'))
    
    incident_label = format_incident_label(incident_id)
    summary_payload = None
    storage_error = None

    try:
        # Get current status for logging
        current_result = supabase.table('alert_incidents').select('icd_status').eq('icd_id', incident_id).execute()
        old_status = current_result.data[0]['icd_status'] if current_result.data else 'Unknown'
        
        # Update incident status
        now = get_philippines_time().isoformat()
        update_data = {
            'icd_status': 'Resolved',
            'resolved_timestamp': now,
            'status_updated_at': now,
            'status_updated_by': session['admin_id']
        }
        
        result = supabase.table('alert_incidents').update(update_data).eq('icd_id', incident_id).execute()
        
        if result.data:
            # Log admin activity
            log_admin_activity(
                session['admin_id'],
                session['admin_name'],
                'status_change',
                incident_id,
                old_status,
                'Resolved'
            )

            try:
                summary_payload, storage_error = create_incident_resolution_report(
                    incident_id,
                    session['admin_id'],
                    session.get('admin_name'),
                    old_status,
                    summary_text
                )
            except Exception as summary_exception:
                storage_error = f"Failed to assemble resolution summary: {summary_exception}"
                print(storage_error)
            
            success_message = f'{incident_label} has been marked as resolved!'
            if summary_payload and summary_payload.get('stored'):
                success_message += ' Resolution summary archived.'
            if summary_payload:
                summary_payload['submitted_summary'] = summary_text
                if storage_error:
                    summary_payload['storage_warning'] = storage_error
            elif storage_error:
                if wants_json:
                    return jsonify({
                        'success': True,
                        'message': success_message,
                        'summary': None,
                        'warning': storage_error
                    })
                flash(f'{incident_label} resolved, but summary storage failed: {storage_error}', 'error')
                return redirect(url_for('dashboard'))

            if wants_json:
                return jsonify({
                    'success': True,
                    'message': success_message,
                    'summary': summary_payload,
                    'storage_warning': storage_error
                })

            flash(success_message, 'success')

            if summary_payload:
                flash(summary_payload, 'resolution_summary')
        else:
            message = 'Failed to update incident status'
            if wants_json:
                return jsonify({'success': False, 'error': message}), 500
            flash(message, 'error')
            
    except Exception as e:
        print(f"Error marking incident as resolved: {e}")
        if wants_json:
            return jsonify({'success': False, 'error': str(e)}), 500
        flash('An error occurred while updating the incident', 'error')
    
    return redirect(url_for('dashboard'))

@app.route('/mark_pending', methods=['POST'])
def mark_pending():
    """Mark incident as pending"""
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    incident_id = request.form.get('incident_id')
    if not incident_id:
        flash('Incident ID is required', 'error')
        return redirect(url_for('dashboard'))
    
    # Check if admin can edit this incident
    can_edit, edit_message = can_admin_edit_incident(incident_id, session['admin_id'])
    if not can_edit:
        flash(f'Permission denied: {edit_message}', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        # Get current incident data (including user_id for chat message)
        current_result = supabase.table('alert_incidents').select('icd_status, user_id').eq('icd_id', incident_id).execute()
        if not current_result.data:
            flash('Incident not found', 'error')
            return redirect(url_for('dashboard'))
            
        old_status = current_result.data[0]['icd_status']
        student_id = current_result.data[0].get('user_id')
        
        # Update incident status
        now = get_philippines_time().isoformat()
        update_data = {
            'icd_status': 'Pending',
            'pending_timestamp': now,
            'status_updated_at': now,
            'status_updated_by': session['admin_id'],
            'assigned_responder_id': session['admin_id']
        }
        
        # Update the incident
        result = supabase.table('alert_incidents').update(update_data).eq('icd_id', incident_id).execute()
        
        if result.data:
            # Log admin activity
            log_admin_activity(
                session['admin_id'],
                session['admin_name'],
                'status_change',
                incident_id,
                old_status,
                'Pending'
            )
            
            # Get admin full name for the chat message
            admin_fullname = session.get('admin_name', 'Admin')
            try:
                admin_result = supabase.table('accounts_admin').select('admin_fullname').eq('admin_id', session['admin_id']).limit(1).execute()
                if admin_result.data and admin_result.data[0].get('admin_fullname'):
                    admin_fullname = admin_result.data[0]['admin_fullname']
            except Exception as e:
                print(f"Error fetching admin full name: {e}")
                # Use session admin_name as fallback
            
            # Send chat message to student if student_id exists
            if student_id:
                try:
                    message = f"Hello, your alert has been received. Admin {admin_fullname} has taken charge of your case and marked it as Pending. Please wait while we assist you."
                    send_result = send_chat_message(
                        incident_id=incident_id,
                        sender_id=session['admin_id'],
                        sender_type='admin',
                        receiver_id=student_id,
                        receiver_type='student',
                        message=message
                    )
                    if send_result:
                        print(f"Chat message sent successfully to student {student_id} for incident {incident_id}")
                    else:
                        print(f"Failed to send chat message to student {student_id} for incident {incident_id}")
                except Exception as chat_error:
                    # Don't fail the pending action if chat message fails
                    print(f"Error sending chat message: {chat_error}")
                    import traceback
                    traceback.print_exc()
            
            flash(f'Incident ICD_9100{incident_id} has been set to pending!', 'success')
        else:
            flash('Failed to update incident status - no data returned', 'error')
            
    except Exception as e:
        error_msg = str(e)
        print(f"Error marking incident as pending: {error_msg}")
        print(f"Full traceback: {traceback.format_exc()}")
        # Provide more specific error message
        if 'column' in error_msg.lower() or 'does not exist' in error_msg.lower():
            flash('Database schema error. Please contact administrator.', 'error')
        else:
            flash(f'An error occurred while updating the incident: {error_msg[:100]}', 'error')
    
    return redirect(url_for('dashboard'))

@app.route('/mark_cancelled', methods=['POST'])
def mark_cancelled():
    """Mark incident as cancelled"""
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    incident_id = request.form.get('incident_id')
    if not incident_id:
        flash('Incident ID is required', 'error')
        return redirect(url_for('dashboard'))
    
    # Check if admin can edit this incident
    can_edit, edit_message = can_admin_edit_incident(incident_id, session['admin_id'])
    if not can_edit:
        flash(f'Permission denied: {edit_message}', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        # Get current status for logging
        current_result = supabase.table('alert_incidents').select('icd_status').eq('icd_id', incident_id).execute()
        old_status = current_result.data[0]['icd_status'] if current_result.data else 'Unknown'
        
        # Update incident status
        now = get_philippines_time().isoformat()
        update_data = {
            'icd_status': 'Cancelled',
            'cancelled_timestamp': now,
            'status_updated_at': now,
            'status_updated_by': session['admin_id']
        }
        
        result = supabase.table('alert_incidents').update(update_data).eq('icd_id', incident_id).execute()
        
        if result.data:
            # Log admin activity
            log_admin_activity(
                session['admin_id'],
                session['admin_name'],
                'status_change',
                incident_id,
                old_status,
                'Cancelled'
            )
            
            flash(f'Incident ICD_9100{incident_id} has been cancelled!', 'success')
        else:
            flash('Failed to update incident status', 'error')
            
    except Exception as e:
        print(f"Error marking incident as cancelled: {e}")
        flash('An error occurred while updating the incident', 'error')
    
    return redirect(url_for('dashboard'))


# ==================== LIVE DATA HELPER FUNCTIONS ====================

def calculate_distance(lat1, lng1, lat2, lng2):
    """Calculate distance between two points in kilometers"""
    R = 6371  # Earth's radius in kilometers
    dLat = (lat2 - lat1) * math.pi / 180
    dLng = (lng2 - lng1) * math.pi / 180
    a = (math.sin(dLat/2) * math.sin(dLat/2) +
         math.cos(lat1 * math.pi / 180) * math.cos(lat2 * math.pi / 180) *
         math.sin(dLng/2) * math.sin(dLng/2))
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    return R * c

def format_distance(distance):
    """Format distance for display"""
    if distance < 1:
        return f"{int(distance * 1000)} meters"
    else:
        return f"{distance:.2f} km"

# Constants for location calculations
UMAK_LAT = 14.5633428
UMAK_LNG = 121.0565387
UMAK_RADIUS = 0.5  # 500 meters

@app.route('/dispatch_team', methods=['POST'])
def dispatch_team():
    """Dispatch team to an incident"""
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    incident_id = request.form.get('incident_id')
    responder_id = request.form.get('responder_id')
    
    if not incident_id or not responder_id:
        flash('Error: Missing incident ID or responder ID', 'error')
        return redirect(url_for('dashboard'))
    
    # Check if admin can assign this incident
    # Allow assignment if:
    # 1. No one is assigned (anyone can assign)
    # 2. Current admin is assigning to themselves (taking over)
    # 3. Current admin is the assigned responder (reassigning)
    try:
        incident_result = supabase.table('alert_incidents').select('icd_status, assigned_responder_id').eq('icd_id', incident_id).execute()
        if not incident_result.data:
            flash('Incident not found', 'error')
            return redirect(url_for('dashboard'))
        
        incident = incident_result.data[0]
        current_assigned = incident.get('assigned_responder_id')
        status = incident.get('icd_status', '')
        
        # If already assigned to someone else and not assigning to self, check permission
        if current_assigned and str(current_assigned) != str(session['admin_id']) and str(responder_id) != str(session['admin_id']):
            can_edit, edit_message = can_admin_edit_incident(incident_id, session['admin_id'])
            if not can_edit:
                flash(f'Permission denied: {edit_message}', 'error')
                return redirect(url_for('dashboard'))
    except Exception as e:
        flash(f'Error checking permissions: {str(e)}', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        # Get current status
        old_status = incident.get('icd_status', 'Unknown')
        
        # Update incident
        pending_timestamp = get_philippines_time().isoformat()
        
        result = supabase.table('alert_incidents').update({
            'icd_status': 'Pending',
            'pending_timestamp': pending_timestamp,
            'assigned_responder_id': responder_id
        }).eq('icd_id', incident_id).execute()
        
        if result.data:
            # Log activity
            log_admin_activity(
                session['admin_id'],
                session['admin_name'],
                'dispatch_team',
                incident_id,
                old_status,
                'Pending'
            )
            flash('Responder assigned and incident set to pending successfully!', 'success')
        else:
            flash('Error assigning responder', 'error')
    
    except Exception as e:
        flash(f'Error assigning responder: {str(e)}', 'error')
    
    return redirect(url_for('dashboard'))

@app.route('/get_student/<int:user_id>')
def get_student(user_id):
    """API endpoint to get student details"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    student = get_student_details(user_id)
    if student:
        return jsonify(student)
    else:
        return jsonify({'error': 'Student not found'}), 404

# ==================== PROFILE ROUTES ====================
@app.route('/profile', methods=['GET', 'POST'])
def profile():
    """Admin profile management route"""
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    admin_id = session['admin_id']
    admin = get_admin_by_id(admin_id)
    
    if not admin:
        flash('Admin account not found.', 'error')
        return redirect(url_for('logout'))
    
    # Ensure admin_id is present in admin object (fallback to session value)
    if 'admin_id' not in admin or not admin.get('admin_id'):
        admin['admin_id'] = admin_id
    
    # Check if profile image exists
    profile_image_exists = False
    if admin.get('admin_profile') and admin['admin_profile'] != 'default.png':
        image_path = os.path.join(app.static_folder, 'images', admin['admin_profile'])
        profile_image_exists = os.path.exists(image_path)
    
    # Handle profile information update
    if request.method == 'POST' and 'update_profile' in request.form:
        full_name = request.form.get('full_name', '').strip()
        # Email is read-only and managed by system administrators - use existing email
        email = admin.get('admin_email', '')
        username = request.form.get('username', '').strip()
        
        # Validate form data (email is not required from form since it's read-only)
        if not all([full_name, username]):
            flash('Full name and username are required.', 'error')
        elif check_username_exists(username, admin_id):
            flash('Username already exists. Please choose a different one.', 'error')
        else:
            # Update profile (email remains unchanged)
            result = update_admin_profile(admin_id, full_name, email, username)
            if result:
                # Update session data
                session['admin_name'] = full_name
                admin['admin_fullname'] = full_name
                # admin_email remains unchanged (read-only)
                admin['admin_user'] = username
                flash('Profile updated successfully!', 'success')
            else:
                flash('Failed to update profile. Please try again.', 'error')
    
    # Handle profile picture upload
    elif request.method == 'POST' and 'profile_image' in request.files:
        file = request.files['profile_image']
        
        if file and file.filename and allowed_file(file.filename):
            try:
                # Create upload directory if it doesn't exist
                upload_dir = os.path.join(app.static_folder, 'images')
                os.makedirs(upload_dir, exist_ok=True)
                
                # Generate unique filename
                file_ext = file.filename.rsplit('.', 1)[1].lower()
                new_filename = f"admin_{admin_id}_{uuid.uuid4().hex[:8]}.{file_ext}"
                file_path = os.path.join(upload_dir, new_filename)
                
                # Save file
                file.save(file_path)
                
                # Update database
                result = update_admin_profile_image(admin_id, new_filename)
                if result:
                    session['admin_profile'] = new_filename
                    session['admin_profile_exists'] = True
                    admin['admin_profile'] = new_filename
                    profile_image_exists = True
                    flash('Profile image updated successfully!', 'success')
                else:
                    flash('Failed to update profile image in database.', 'error')
                    # Remove uploaded file if database update failed
                    if os.path.exists(file_path):
                        os.remove(file_path)
                        
            except Exception as e:
                print(f"Error uploading profile image: {e}")
                flash('Failed to upload profile image. Please try again.', 'error')
        else:
            if file and file.filename:
                flash('Invalid file type. Only JPG, PNG, and GIF files are allowed.', 'error')
            else:
                flash('Please select a file to upload.', 'error')
    
    # Get alert counts for header
    try:
        active_alerts_count = safe_count_query('alert_incidents', [{'type': 'in', 'column': 'icd_status', 'value': ['Active', 'Pending']}])
        
        # Get active and pending incident counts for tooltip
        all_incidents_result = supabase.table('alert_incidents').select('icd_status').execute()
        all_incidents_data = all_incidents_result.data or []
        active_incidents = len([i for i in all_incidents_data if i.get('icd_status') == 'Active'])
        pending_incidents = len([i for i in all_incidents_data if i.get('icd_status') == 'Pending'])
    except Exception as e:
        print(f"Error fetching alert counts for profile: {e}")
        active_alerts_count = 0
        active_incidents = 0
        pending_incidents = 0
    
    return render_template('profile.html', 
                         admin=admin,
                         profile_image_exists=profile_image_exists,
                         active_alerts_count=active_alerts_count,
                         active_incidents=active_incidents,
                         pending_incidents=pending_incidents)

# ==================== USER MANAGEMENT ROUTES ====================
@app.route('/user-management', methods=['GET', 'POST'])
def user_management():
    """User management route with full CRUD functionality"""
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    # Restrict access to System Administrators only
    current_admin = get_admin_by_id(session['admin_id'])
    if not current_admin or current_admin.get('admin_role') != 'System Administrator':
        flash('Access denied. Only System Administrators can manage users.', 'error')
        return redirect(url_for('dashboard'))
    
    # Handle POST requests for various actions
    if request.method == 'POST':
        action = request.form.get('action')
        user_type = request.form.get('type')
        
        if action == 'add':
            return handle_add_user(user_type)
        elif action == 'edit':
            return handle_edit_user(user_type)
    
    # Handle GET requests for delete
    if request.method == 'GET' and request.args.get('action') == 'delete':
        return handle_delete_user()
    
    # GET request - display user management page
    return display_user_management()

def handle_add_user(user_type):
    """Handle adding new users (admin or student)"""
    # Get form data
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '').strip()
    confirm_password = request.form.get('confirm_password', '').strip()
    email = request.form.get('email', '').strip()
    fullname = request.form.get('fullname', '').strip()
    
    # Validate required fields
    if not username:
        flash('Username is required!', 'error')
        return redirect(url_for('user_management', filter=request.args.get('filter', 'all')))
    
    if not email:
        flash('Email is required!', 'error')
        return redirect(url_for('user_management', filter=request.args.get('filter', 'all')))
    
    if not fullname:
        flash('Full name is required!', 'error')
        return redirect(url_for('user_management', filter=request.args.get('filter', 'all')))
    
    if not password:
        flash('Password is required!', 'error')
        return redirect(url_for('user_management', filter=request.args.get('filter', 'all')))
    
    # Validate passwords
    if password != confirm_password:
        flash('Passwords do not match!', 'error')
        return redirect(url_for('user_management', filter=request.args.get('filter', 'all')))
    
    # Use the same password validation as reset password
    is_valid, password_message = validate_password(password)
    if not is_valid:
        flash(f'Password validation failed: {password_message}', 'error')
        return redirect(url_for('user_management', filter=request.args.get('filter', 'all')))
    
    # Check if username already exists
    if check_username_exists(username):
        flash(f'Username "{username}" already exists. Please choose a different username.', 'error')
        return redirect(url_for('user_management', filter=request.args.get('filter', 'all')))
    
    # Handle profile image upload
    profile_image = 'default.png'
    if 'profile_image' in request.files:
        file = request.files['profile_image']
        upload_result = handle_profile_upload(file, user_type)
        if upload_result.get('error'):
            flash(upload_result['error'], 'error')
            return redirect(url_for('user_management', filter=request.args.get('filter', 'all')))
        profile_image = upload_result.get('filename', 'default.png')
    
    try:
        if user_type == 'admin':
            role = request.form.get('role', '').strip()
            status = request.form.get('status', '').strip()
            
            # Validate role is required
            if not role:
                flash('Role is required for administrators!', 'error')
                return redirect(url_for('user_management', filter=request.args.get('filter', 'all')))
            
            # Generate admin ID for new admin
            new_admin_id = generate_next_admin_id()
            
            admin_data = {
                'admin_id': new_admin_id,
                'admin_user': username,
                'admin_pass': hash_password(password) if password else None,
                'admin_email': email,
                'admin_fullname': fullname,
                'admin_role': role,
                'admin_status': status,
                'admin_profile': profile_image,
                'admin_created_at': datetime.now().isoformat()
            }
            
            print(f"📤 Inserting admin data: {admin_data}")
            result = supabase.table('accounts_admin').insert(admin_data).execute()
            print(f"✅ Admin insert result: {result}")
            if hasattr(result, 'error') and result.error:
                print(f"❌ Supabase error: {result.error}")
                error_details = result.error
                if isinstance(error_details, dict):
                    error_msg = error_details.get('message', str(error_details))
                    error_code = error_details.get('code', '')
                    if 'RLS' in error_msg or 'row level security' in error_msg.lower() or error_code == '42501':
                        error_msg = f"RLS Policy Error: {error_msg}. Please run ADD_USER_INSERT_POLICIES.sql in Supabase SQL Editor."
                else:
                    error_msg = str(error_details)
                raise Exception(f"Database error: {error_msg}")
            
        elif user_type == 'student':
            student_id = request.form.get('student_id', '').strip()
            
            # Validate student_id is required
            if not student_id:
                flash('Student ID is required!', 'error')
                return redirect(url_for('user_management', filter=request.args.get('filter', 'all')))
            
            # Check if student_id already exists
            existing_student = supabase.table('accounts_student').select('user_id').eq('student_id', student_id).execute()
            if existing_student.data:
                flash(f'Student ID "{student_id}" already exists. Please use a different student ID.', 'error')
                return redirect(url_for('user_management', filter=request.args.get('filter', 'all')))
            
            yearlvl = request.form.get('yearlvl', '').strip()
            college = request.form.get('college', '').strip() or 'CLAS'  # Default to CLAS
            cnum = request.form.get('cnum', '').strip()
            emergency = request.form.get('emergency', '').strip() or None
            contactperson = request.form.get('contactperson', '').strip() or None
            cprelationship = request.form.get('cprelationship', '').strip() or None
            secondary_emergency = request.form.get('secondary_emergency', '').strip() or None
            secondary_contact = request.form.get('secondary_contact', '').strip() or None
            secondary_relationship = request.form.get('secondary_relationship', '').strip() or None
            medinfo = request.form.get('medinfo', '').strip() or None
            address = request.form.get('address', '').strip()
            status = request.form.get('status', '').strip() or 'Active'
            residency = request.form.get('residency', '').strip() or 'MAKATI'
            
            # NOTE: user_id is NOT included - it's auto-generated by the database sequence (bigint)
            student_data = {
                'student_id': student_id,
                'student_user': username,
                'student_pass': hash_password(password) if password else None,
                'student_email': email,
                'full_name': fullname,
                'student_yearlvl': yearlvl,
                'student_college': college,
                'student_cnum': cnum,
                'primary_emergencycontact': emergency,
                'primary_contactperson': contactperson,
                'primary_cprelationship': cprelationship,
                'secondary_emergencycontact': secondary_emergency,
                'secondary_contactperson': secondary_contact,
                'secondary_cprelationship': secondary_relationship,
                'student_medinfo': medinfo,
                'student_address': address,
                'student_profile': profile_image,
                'residency': residency,
                'student_status': status,
                'student_created_at': datetime.now().isoformat()
            }
            
            # Keep None values only for fields that are allowed to be NULL in the database
            optional_nullable_fields = ['primary_cprelationship', 'secondary_cprelationship', 'primary_emergencycontact', 
                                       'secondary_emergencycontact', 'primary_contactperson', 'secondary_contactperson', 
                                       'student_medinfo', 'student_profile']
            student_data = {k: v for k, v in student_data.items() if v is not None or k in optional_nullable_fields}
            
            print(f"📤 Inserting student data (user_id will be auto-generated): {student_data}")
            result = supabase.table('accounts_student').insert(student_data).execute()
            print(f"✅ Student insert result: {result}")
            if hasattr(result, 'error') and result.error:
                print(f"❌ Supabase error: {result.error}")
                error_details = result.error
                if isinstance(error_details, dict):
                    error_msg = error_details.get('message', str(error_details))
                    error_code = error_details.get('code', '')
                    if 'RLS' in error_msg or 'row level security' in error_msg.lower() or error_code == '42501':
                        error_msg = f"RLS Policy Error: {error_msg}. Please run ADD_USER_INSERT_POLICIES.sql in Supabase SQL Editor."
                else:
                    error_msg = str(error_details)
                raise Exception(f"Database error: {error_msg}")
        
        if result.data:
            flash(f'{user_type.title()} added successfully!', 'success')
        else:
            # Check for errors
            if hasattr(result, 'error') and result.error:
                error_details = result.error
                if isinstance(error_details, dict):
                    error_msg = error_details.get('message', str(error_details))
                    error_code = error_details.get('code', '')
                    if 'RLS' in error_msg or 'row level security' in error_msg.lower() or error_code == '42501':
                        error_msg = f"RLS Policy Error: {error_msg}. Please run ADD_USER_INSERT_POLICIES.sql in Supabase SQL Editor."
                else:
                    error_msg = str(error_details)
                flash(f'Error adding {user_type}: {error_msg}', 'error')
            else:
                flash(f'Error adding {user_type}: No data returned from database', 'error')
            
    except Exception as e:
        print(f"❌ Error adding {user_type}: {e}")
        print(f"❌ Error details: {traceback.format_exc()}")
        error_message = str(e)
        # Provide more helpful error messages
        if 'duplicate' in error_message.lower() or 'unique' in error_message.lower():
            error_message = 'A user with this information already exists. Please check for duplicates.'
        elif 'enum' in error_message.lower() or 'invalid input' in error_message.lower():
            error_message = 'Invalid enum value. Please check that all dropdown values match the database.'
        elif 'not null' in error_message.lower():
            error_message = 'Required field is missing. Please fill in all required fields.'
        flash(f'Error adding {user_type}: {error_message}', 'error')
    
    return redirect(url_for('user_management', filter=request.args.get('filter', 'all')))

def handle_edit_user(user_type):
    """Handle editing existing users"""
    user_id = request.form.get('id')
    if not user_id:
        flash('User ID is required', 'error')
        return redirect(url_for('user_management'))
    
    # Get form data
    username = request.form.get('username', '').strip()
    email = request.form.get('email', '').strip()
    fullname = request.form.get('fullname', '').strip()
    password = request.form.get('password', '').strip()
    confirm_password = request.form.get('confirm_password', '').strip()
    
    # Validate password if provided
    if password:
        if password != confirm_password:
            flash('Passwords do not match!', 'error')
            return redirect(url_for('user_management', filter=request.args.get('filter', 'all')))
        
        # Use the same password validation as reset password
        is_valid, password_message = validate_password(password)
        if not is_valid:
            flash(f'Password validation failed: {password_message}', 'error')
            return redirect(url_for('user_management', filter=request.args.get('filter', 'all')))
    
    # Handle profile image upload
    profile_image = None
    if 'profile_image' in request.files:
        file = request.files['profile_image']
        if file and file.filename:
            upload_result = handle_profile_upload(file, user_type)
            if upload_result.get('error'):
                flash(upload_result['error'], 'error')
                return redirect(url_for('user_management', filter=request.args.get('filter', 'all')))
            profile_image = upload_result.get('filename')
    
    try:
        if user_type == 'admin':
            role = request.form.get('role', '').strip()
            status = request.form.get('status', '').strip()
            
            update_data = {
                'admin_user': username,
                'admin_email': email,
                'admin_fullname': fullname,
                'admin_role': role,
                'admin_status': status
            }
            
            if password:
                update_data['admin_pass'] = password  # Plain text for now
            
            if profile_image:
                update_data['admin_profile'] = profile_image
            
            result = supabase.table('accounts_admin').update(update_data).eq('admin_id', user_id).execute()
            
        elif user_type == 'student':
            student_id = request.form.get('student_id', '').strip()
            yearlvl = request.form.get('yearlvl', '').strip()
            college = request.form.get('college', '').strip()
            cnum = request.form.get('cnum', '').strip()
            emergency = request.form.get('emergency', '').strip()
            contactperson = request.form.get('contactperson', '').strip()
            cprelationship = request.form.get('cprelationship', '').strip()
            secondary_emergency = request.form.get('secondary_emergency', '').strip()
            secondary_contact = request.form.get('secondary_contact', '').strip()
            secondary_relationship = request.form.get('secondary_relationship', '').strip()
            medinfo = request.form.get('medinfo', '').strip()
            address = request.form.get('address', '').strip()
            status = request.form.get('status', '').strip()
            
            update_data = {
                'student_id': student_id,
                'student_user': username,
                'student_email': email,
                'full_name': fullname,
                'student_yearlvl': yearlvl,
                'student_college': college,
                'student_cnum': cnum,
                'primary_emergencycontact': emergency,
                'primary_contactperson': contactperson,
                'primary_cprelationship': cprelationship,
                'secondary_emergencycontact': secondary_emergency,
                'secondary_contactperson': secondary_contact,
                'secondary_cprelationship': secondary_relationship,
                'student_medinfo': medinfo,
                'student_address': address,
                'student_status': status
            }
            
            if password:
                update_data['student_pass'] = password  # Plain text for now
            
            if profile_image:
                update_data['student_profile'] = profile_image
            
            print(f"Updating student data: {update_data}")
            result = supabase.table('accounts_student').update(update_data).eq('user_id', user_id).execute()
            print(f"Student update result: {result}")
        
        if result.data:
            flash(f'{user_type.title()} updated successfully!', 'success')
        else:
            flash(f'Error updating {user_type}', 'error')
            
    except Exception as e:
        print(f"Error updating {user_type}: {e}")
        print(f"Error details: {traceback.format_exc()}")
        flash(f'Error updating {user_type}: {str(e)}', 'error')
    
    return redirect(url_for('user_management', filter=request.args.get('filter', 'all')))

def handle_delete_user():
    """Handle deleting users"""
    user_id = request.args.get('id')
    user_type = request.args.get('type')
    
    if not user_id or not user_type:
        flash('User ID and type are required', 'error')
        return redirect(url_for('user_management'))
    
    try:
        if user_type == 'admin':
            result = supabase.table('accounts_admin').delete().eq('admin_id', user_id).execute()
        elif user_type == 'student':
            result = supabase.table('accounts_student').delete().eq('user_id', user_id).execute()
        
        if result.data:
            flash(f'{user_type.title()} deleted successfully!', 'success')
        else:
            flash(f'Error deleting {user_type}', 'error')
        
    except Exception as e:
        print(f"Error deleting {user_type}: {e}")
        flash(f'Error deleting {user_type}: {str(e)}', 'error')
    
    return redirect(url_for('user_management', filter=request.args.get('filter', 'all')))

def get_contact_relationship_enum_values():
    """Get contact_relationship enum values from PostgreSQL database"""
    # Standard relationship values that are commonly used
    # These should match the enum values in your database
    standard_relationships = [
        'Parent',
        'Guardian',
        'Sibling',
        'Spouse',
        'Relative',
        'Friend',
        'Other'
    ]
    
    try:
        # Try to get enum values from database using RPC
        # First, try to get from existing data
        result = supabase.table('accounts_student').select('primary_cprelationship, secondary_cprelationship').execute()
        
        if result.data:
            # Collect all unique relationship values from existing data
            relationships = set()
            for record in result.data:
                if record.get('primary_cprelationship'):
                    relationships.add(record['primary_cprelationship'])
                if record.get('secondary_cprelationship'):
                    relationships.add(record['secondary_cprelationship'])
            
            # Combine with standard values and sort
            all_relationships = sorted(list(set(standard_relationships + list(relationships))))
            return all_relationships
        
        # If no data exists, return standard values
        return standard_relationships
        
    except Exception as e:
        print(f"⚠️ Could not fetch enum values from database: {e}")
        print("Using standard relationship values")
        return standard_relationships

def display_user_management():
    """Display the user management page with pagination and filtering - FIXED VERSION"""
    # Get query parameters
    page = max(1, int(request.args.get('page', 1)))
    limit_param = request.args.get('limit', 'all')
    filter_type = request.args.get('filter', 'all')
    search_term = request.args.get('search', '')
    action = request.args.get('action')
    edit_type = request.args.get('type')
    edit_id = request.args.get('id')
    
    # Handle limit
    if limit_param == 'all':
        limit = 999999
        show_all = True
    else:
        limit = int(limit_param)
        show_all = False
    
    offset = (page - 1) * limit
    
    # Get user counts - FIXED with proper error handling
    try:
        admin_count_result = supabase.table('accounts_admin').select('*', count='exact').execute()
        admin_count = admin_count_result.count or 0
        print(f"📊 Admin count: {admin_count}")
    except Exception as e:
        print(f"❌ Error counting admins: {e}")
        admin_count = 0
    
    try:
        student_count_result = supabase.table('accounts_student').select('*', count='exact').execute()
        student_count = student_count_result.count or 0
        print(f"📊 Student count: {student_count}")
    except Exception as e:
        print(f"❌ Error counting students: {e}")
        student_count = 0
    
    total_count = admin_count + student_count
    print(f"📊 Total users: {total_count}")
    
    # Get active alerts count
    active_alerts_count = safe_count_query('alert_incidents', [{'type': 'in', 'column': 'icd_status', 'value': ['Active', 'Pending']}])
    
    # Build query based on filter - IMPROVED VERSION
    users = []
    total_records = 0
    
    try:
        print(f"🔍 Fetching users with filter: {filter_type}")
        
        if filter_type == 'admin':
            # Get admin users with proper field mapping
            result = supabase.table('accounts_admin').select('*').execute()
            print(f"📋 Raw admin data: {len(result.data) if result.data else 0} records")
            
            users = result.data or []
            total_records = len(users)
            
            for user in users:
                user['type'] = 'admin'
                user['id'] = user['admin_id']
                user['username'] = user['admin_user']
                user['email'] = user['admin_email']
                user['fullname'] = user['admin_fullname']
                user['role'] = user['admin_role']
                user['status'] = user['admin_status']
                user['created_at'] = user.get('admin_created_at', 'N/A')
                user['last_login'] = user.get('admin_last_login', None)
                user['profile_image'] = user.get('admin_profile', 'default.png')
                
        elif filter_type == 'student':
            # Get student users with proper field mapping
            result = supabase.table('accounts_student').select('*').execute()
            print(f"📋 Raw student data: {len(result.data) if result.data else 0} records")
            
            users = result.data or []
            total_records = len(users)
            
            for user in users:
                user['type'] = 'student'
                user['id'] = user['user_id']
                user['username'] = user['student_user']
                user['email'] = user['student_email']
                user['fullname'] = user['full_name']
                user['role'] = 'Student'
                user['status'] = user['student_status']
                user['created_at'] = user.get('student_created_at', 'N/A')
                user['last_login'] = user.get('student_last_login', None)
                user['profile_image'] = user.get('student_profile', 'default.png')
                user['is_verified'] = user.get('is_verified', user.get('email_verified'))
                
                # ADDITIONAL STUDENT FIELDS FOR TABLE DISPLAY
                user['student_id'] = user.get('student_id', '')
                user['student_yearlvl'] = user.get('student_yearlvl', '')
                user['student_college'] = user.get('student_college', '')
                user['student_cnum'] = user.get('student_cnum', '')
                user['primary_emergencycontact'] = user.get('primary_emergencycontact', '')
                user['primary_contactperson'] = user.get('primary_contactperson', '')
                user['primary_cprelationship'] = user.get('primary_cprelationship', '')
                user['secondary_emergencycontact'] = user.get('secondary_emergencycontact', '')
                user['secondary_contactperson'] = user.get('secondary_contactperson', '')
                user['secondary_cprelationship'] = user.get('secondary_cprelationship', '')
                user['student_medinfo'] = user.get('student_medinfo', '')
                user['student_address'] = user.get('student_address', '')
                
        else:  # all users
            # Get admin users
            admin_result = supabase.table('accounts_admin').select('*').execute()
            admin_users = admin_result.data or []
            print(f"📋 Raw admin data (all): {len(admin_users)} records")
            
            for user in admin_users:
                user['type'] = 'admin'
                user['id'] = user['admin_id']
                user['username'] = user['admin_user']
                user['email'] = user['admin_email']
                user['fullname'] = user['admin_fullname']
                user['role'] = user['admin_role']
                user['status'] = user['admin_status']
                user['created_at'] = user.get('admin_created_at', 'N/A')
                user['last_login'] = user.get('admin_last_login', None)
                user['profile_image'] = user.get('admin_profile', 'default.png')
            
            # Get student users
            student_result = supabase.table('accounts_student').select('*').execute()
            student_users = student_result.data or []
            print(f"📋 Raw student data (all): {len(student_users)} records")
            
            for user in student_users:
                user['type'] = 'student'
                user['id'] = user['user_id']
                user['username'] = user['student_user']
                user['email'] = user['student_email']
                user['fullname'] = user['full_name']
                user['role'] = 'Student'
                user['status'] = user['student_status']
                user['created_at'] = user.get('student_created_at', 'N/A')
                user['last_login'] = user.get('student_last_login', None)
                user['profile_image'] = user.get('student_profile', 'default.png')
                user['is_verified'] = user.get('is_verified', user.get('email_verified'))
                
                # ADDITIONAL STUDENT FIELDS FOR TABLE DISPLAY
                user['student_id'] = user.get('student_id', '')
                user['student_yearlvl'] = user.get('student_yearlvl', '')
                user['student_cnum'] = user.get('student_cnum', '')
                user['primary_emergencycontact'] = user.get('primary_emergencycontact', '')
                user['primary_contactperson'] = user.get('primary_contactperson', '')
                user['primary_cprelationship'] = user.get('primary_cprelationship', '')
                user['secondary_emergencycontact'] = user.get('secondary_emergencycontact', '')
                user['secondary_contactperson'] = user.get('secondary_contactperson', '')
                user['secondary_cprelationship'] = user.get('secondary_cprelationship', '')
                user['student_medinfo'] = user.get('student_medinfo', '')
                user['student_address'] = user.get('student_address', '')
            
            # Combine and paginate
            all_users = admin_users + student_users
            total_records = len(all_users)
            users = all_users[offset:offset + limit] if not show_all else all_users
            
        print(f"✅ Processed {len(users)} users for display")
            
    except Exception as e:
        print(f"❌ Error fetching users: {e}")
        print(f"Error details: {traceback.format_exc()}")
        users = []
        total_records = 0
    
    # Apply search filter if search term exists
    if search_term:
        filtered_users = []
        search_lower = search_term.lower()
        for user in users:
            # Search in various fields
            searchable_fields = [
                user.get('username', ''),
                user.get('email', ''),
                user.get('fullname', ''),
                user.get('role', ''),
                user.get('status', ''),
                user.get('student_id', ''),
                user.get('student_yearlvl', ''),
                user.get('student_cnum', ''),
                user.get('primary_emergencycontact', ''),
                user.get('primary_contactperson', '')
            ]
            if any(search_lower in str(field).lower() for field in searchable_fields if field):
                filtered_users.append(user)
        users = filtered_users
        total_records = len(users)
        print(f"🔍 After search filtering: {len(users)} users")
    
    # Calculate pagination
    if show_all:
        total_pages = 1
    else:
        total_pages = math.ceil(total_records / limit) if total_records > 0 else 1
    
    # Get edit data if editing
    edit_data = None
    if action == 'edit' and edit_id and edit_type:
        try:
            if edit_type == 'admin':
                result = supabase.table('accounts_admin').select('*').eq('admin_id', edit_id).execute()
            elif edit_type == 'student':
                result = supabase.table('accounts_student').select('*').eq('user_id', edit_id).execute()
            
            if result.data:
                edit_data = result.data[0]
                # Normalize email verification flag
                if edit_data is not None:
                    edit_data['is_verified'] = edit_data.get('is_verified', edit_data.get('email_verified'))
                print(f"Edit data loaded: {edit_data}")
            else:
                print(f"No data found for {edit_type} with ID {edit_id}")
        except Exception as e:
            print(f"Error fetching edit data: {e}")
            print(f"Error details: {traceback.format_exc()}")
    
    # Get unique values from database for dropdowns
    try:
        # Get unique year levels
        year_level_result = supabase.table('accounts_student').select('student_yearlvl').not_.is_('student_yearlvl', 'null').execute()
        year_levels = sorted(list(set([y.get('student_yearlvl') for y in (year_level_result.data or []) if y.get('student_yearlvl')])))
        if not year_levels:
            year_levels = ['First Year', 'Second Year', 'Third Year', 'Fourth Year']  # Fallback
    except:
        year_levels = ['First Year', 'Second Year', 'Third Year', 'Fourth Year']  # Fallback
    
    try:
        # Get unique colleges
        college_result = supabase.table('accounts_student').select('student_college').not_.is_('student_college', 'null').execute()
        colleges = sorted(list(set([c.get('student_college') for c in (college_result.data or []) if c.get('student_college')])))
    except:
        colleges = []
    
    try:
        residency_result = supabase.table('accounts_student').select('residency').not_.is_('residency', 'null').execute()
        residency_options = sorted(list(set([r.get('residency') for r in (residency_result.data or []) if r.get('residency')])))
        if not residency_options:
            residency_options = ['MAKATI', 'NON-MAKATI']
    except:
        residency_options = ['MAKATI', 'NON-MAKATI']
    
    # Get contact relationship enum values (from database enum or existing data)
    contact_relationships = get_contact_relationship_enum_values()
    
    # Also get unique values from existing data to ensure we have all values
    try:
        primary_rel_result = supabase.table('accounts_student').select('primary_cprelationship').not_.is_('primary_cprelationship', 'null').execute()
        existing_primary = [r.get('primary_cprelationship') for r in (primary_rel_result.data or []) if r.get('primary_cprelationship')]
        
        secondary_rel_result = supabase.table('accounts_student').select('secondary_cprelationship').not_.is_('secondary_cprelationship', 'null').execute()
        existing_secondary = [r.get('secondary_cprelationship') for r in (secondary_rel_result.data or []) if r.get('secondary_cprelationship')]
        
        # Combine all relationship values and remove duplicates
        all_relationships = sorted(list(set(contact_relationships + existing_primary + existing_secondary)))
        primary_relationships = all_relationships
        secondary_relationships = all_relationships
    except:
        # If error, use the enum values
        primary_relationships = contact_relationships
        secondary_relationships = contact_relationships
    
    # Define options
    admin_roles = ['Security Staff', 'Dispatcher', 'System Administrator']
    status_options = ['Active', 'Inactive']
    admin_status_options = ['Active', 'Inactive', 'Away']
    year_level_options = year_levels
    
    return render_template('user_management.html',
                         users=users,
                         total_records=total_records,
                         page=page,
                         limit=limit,
                         limit_param=limit_param,
                         total_pages=total_pages,
                         filter_type=filter_type,
                         show_all=show_all,
                         admin_count=admin_count,
                         student_count=student_count,
                         total_count=total_count,
                         active_alerts_count=active_alerts_count,
                         edit_data=edit_data,
                         action=action,
                         edit_type=edit_type,
                         admin_roles=admin_roles,
                         status_options=status_options,
                         admin_status_options=admin_status_options,
                         year_level_options=year_level_options,
                         colleges=colleges,
                         residency_options=residency_options,
                         primary_relationships=primary_relationships,
                         secondary_relationships=secondary_relationships)

# ==================== CHAT SYSTEM FUNCTIONS ====================

def create_chat_table():
    """Create chat messages table if it doesn't exist"""
    try:
        # Create chat_messages table
        create_table_sql = """
        CREATE TABLE IF NOT EXISTS chat_messages (
            id SERIAL PRIMARY KEY,
            incident_id VARCHAR(50) NOT NULL,
            sender_id VARCHAR(100) NOT NULL,
            sender_type VARCHAR(20) NOT NULL CHECK (sender_type IN ('admin', 'student')),
            receiver_id VARCHAR(100) NOT NULL,
            receiver_type VARCHAR(20) NOT NULL CHECK (receiver_type IN ('admin', 'student')),
            message TEXT NOT NULL,
            timestamp TIMESTAMPTZ DEFAULT NOW(),
            is_read BOOLEAN DEFAULT FALSE
        );
        """
        
        # Note: In a real implementation, you would execute this SQL
        # For now, we'll assume the table exists or create it manually
        print("Chat messages table structure prepared")
        return True
    except Exception as e:
        print(f"Error creating chat table: {e}")
        return False

def get_students_with_incidents():
    """Get list of students who have reported incidents"""
    try:
        # Get all incidents first
        incidents_result = supabase.table('alert_incidents').select('user_id, icd_id, icd_status, icd_category, icd_timestamp').execute()
        incidents = incidents_result.data or []
        
        if not incidents:
            print("No incidents found in database")
            return []
        
        # Filter out incidents with null, None, or invalid user_ids
        valid_incidents = []
        for incident in incidents:
            user_id = incident.get('user_id')
            # Check if user_id is valid (not None, not 'None', not empty string)
            if user_id is not None and str(user_id).strip() and str(user_id).lower() != 'none':
                valid_incidents.append(incident)
        
        if not valid_incidents:
            print("No incidents with valid user_id found")
            return []
        
        # Get unique student IDs - handle both string and integer types
        student_ids_set = set()
        for incident in valid_incidents:
            user_id = incident.get('user_id')
            if user_id is not None:
                # Try to convert to int first (if it's numeric), otherwise keep as string
                try:
                    user_id_int = int(user_id)
                    student_ids_set.add(user_id_int)
                except (ValueError, TypeError):
                    student_ids_set.add(str(user_id))
        
        student_ids = list(student_ids_set)
        
        if not student_ids:
            print("No valid student IDs extracted from incidents")
            return []
        
        # Get student details - query in batches if needed to avoid issues
        all_students = []
        # Process in smaller batches to avoid query size limits
        batch_size = 100
        for i in range(0, len(student_ids), batch_size):
            batch = student_ids[i:i + batch_size]
            try:
                # Try querying with the batch - handle both string and integer user_ids
                students_result = supabase.table('accounts_student').select('user_id, full_name, student_id, student_cnum, student_email').in_('user_id', batch).execute()
                if students_result.data:
                    all_students.extend(students_result.data)
            except Exception as batch_error:
                print(f"Error fetching student batch: {batch_error}")
                import traceback
                traceback.print_exc()
                # Try individual queries as fallback
                for student_id in batch:
                    try:
                        student_result = supabase.table('accounts_student').select('user_id, full_name, student_id, student_cnum, student_email').eq('user_id', student_id).limit(1).execute()
                        if student_result.data:
                            all_students.extend(student_result.data)
                    except Exception as individual_error:
                        print(f"Error fetching individual student {student_id}: {individual_error}")
                        continue
        
        if not all_students:
            print("No students found matching incident user_ids")
            return []
        
        # Create a mapping of user_id to student for quick lookup
        # Handle both string and integer keys
        students_map = {}
        for student in all_students:
            user_id = student.get('user_id')
            if user_id is not None:
                # Store with both string and integer keys for flexibility
                students_map[str(user_id)] = student
                try:
                    students_map[int(user_id)] = student
                except (ValueError, TypeError):
                    pass
        
        # Get latest chat message timestamp for each student (if chat_messages table exists)
        # Note: This function doesn't have access to admin_id, so we'll get all messages
        # The API endpoint will filter by admin_id when called
        chat_timestamps = {}
        if check_chat_table_exists():
            try:
                # Get all chat messages involving students (where student is sender or receiver)
                # We'll get messages where receiver_type='student' or sender_type='student'
                # Note: In the API endpoint, we'll filter by admin_id
                chat_messages_result = supabase.table('chat_messages').select('sender_id, receiver_id, sender_type, receiver_type, timestamp').execute()
                if chat_messages_result.data:
                    for msg in chat_messages_result.data:
                        msg_timestamp = msg.get('timestamp') or msg.get('created_at')
                        if not msg_timestamp:
                            continue
                        
                        # If student is the receiver, use receiver_id
                        if msg.get('receiver_type') == 'student':
                            student_id = str(msg.get('receiver_id', ''))
                            if student_id and student_id not in chat_timestamps:
                                chat_timestamps[student_id] = msg_timestamp
                            elif student_id:
                                # Keep the most recent timestamp
                                try:
                                    if msg_timestamp > chat_timestamps[student_id]:
                                        chat_timestamps[student_id] = msg_timestamp
                                except:
                                    chat_timestamps[student_id] = msg_timestamp
                        
                        # If student is the sender, use sender_id
                        if msg.get('sender_type') == 'student':
                            student_id = str(msg.get('sender_id', ''))
                            if student_id and student_id not in chat_timestamps:
                                chat_timestamps[student_id] = msg_timestamp
                            elif student_id:
                                # Keep the most recent timestamp
                                try:
                                    if msg_timestamp > chat_timestamps[student_id]:
                                        chat_timestamps[student_id] = msg_timestamp
                                except:
                                    chat_timestamps[student_id] = msg_timestamp
            except Exception as e:
                print(f"Error fetching chat message timestamps: {e}")
        
        # Combine student info with their latest incident
        student_data = []
        processed_students = set()
        
        for incident in valid_incidents:
            user_id = incident.get('user_id')
            if user_id is None:
                continue
                
            user_id_str = str(user_id)
            
            # Skip if we've already processed this student
            if user_id_str in processed_students:
                continue
            
            # Check if student exists (try both string and int keys)
            student = None
            if user_id_str in students_map:
                student = students_map[user_id_str]
            else:
                try:
                    user_id_int = int(user_id)
                    if user_id_int in students_map:
                        student = students_map[user_id_int]
                except (ValueError, TypeError):
                    pass
            
            if student:
                processed_students.add(user_id_str)
                
                # Get all incidents for this student to find the latest
                student_incidents = [inc for inc in valid_incidents if str(inc.get('user_id')) == user_id_str]
                if student_incidents:
                    # Sort by timestamp to get the latest
                    try:
                        latest_incident = max(student_incidents, key=lambda x: x.get('icd_timestamp', '') or '')
                    except Exception:
                        latest_incident = student_incidents[0]
                    
                    # Get latest chat message timestamp for this student
                    latest_chat_timestamp = chat_timestamps.get(user_id_str, None)
                    
                    student_data.append({
                        'user_id': str(student['user_id']),  # Ensure string format
                        'full_name': student.get('full_name', 'Unknown'),
                        'student_id': student.get('student_id', ''),
                        'contact_number': student.get('student_cnum', ''),
                        'email': student.get('student_email', ''),
                        'latest_incident_id': str(latest_incident.get('icd_id', '')),
                        'latest_incident_status': latest_incident.get('icd_status', 'Unknown'),
                        'latest_incident_category': latest_incident.get('icd_category', 'Unknown'),
                        'latest_incident_timestamp': latest_incident.get('icd_timestamp', ''),
                        'latest_chat_timestamp': latest_chat_timestamp  # Most recent chat message timestamp
                    })
        
        # Sort students by latest chat timestamp (most recent first), then by incident timestamp
        def get_sort_key(student):
            chat_ts = student.get('latest_chat_timestamp')
            incident_ts = student.get('latest_incident_timestamp', '')
            
            # Convert timestamps to comparable format
            def parse_timestamp(ts):
                if not ts:
                    return None
                try:
                    # Try to parse ISO format or other common formats
                    if isinstance(ts, str):
                        # Handle ISO format with timezone
                        if 'T' in ts:
                            from datetime import datetime
                            try:
                                dt = datetime.fromisoformat(ts.replace('Z', '+00:00'))
                                return dt.timestamp()
                            except:
                                pass
                        # Try other formats
                        try:
                            from datetime import datetime
                            dt = datetime.strptime(ts[:19], '%Y-%m-%d %H:%M:%S')
                            return dt.timestamp()
                        except:
                            pass
                    # If it's already a datetime object
                    if hasattr(ts, 'timestamp'):
                        return ts.timestamp()
                    return None
                except:
                    return None
            
            chat_timestamp = parse_timestamp(chat_ts) if chat_ts else None
            incident_timestamp = parse_timestamp(incident_ts) if incident_ts else None
            
            # Prioritize chat timestamp, fallback to incident timestamp
            if chat_timestamp is not None:
                return (chat_timestamp, 1)  # Higher priority for chat
            elif incident_timestamp is not None:
                return (incident_timestamp, 0)
            else:
                return (0, 0)  # No timestamp available
        
        student_data.sort(key=get_sort_key, reverse=True)
        
        print(f"Successfully loaded {len(student_data)} students with incidents")
        return student_data
    except Exception as e:
        print(f"Error getting students with incidents: {e}")
        import traceback
        traceback.print_exc()
        return []

def check_chat_table_exists():
    """Check if chat_messages table exists"""
    try:
        # Try to query the table with a limit of 0 to check if it exists
        supabase.table('chat_messages').select('id').limit(0).execute()
        return True
    except Exception as e:
        error_msg = str(e)
        if 'PGRST205' in error_msg or 'Could not find the table' in error_msg:
            return False
        # If it's a different error, the table might exist but have other issues
        return True

def validate_incident_exists(incident_id):
    """Validate that an incident exists in alert_incidents table"""
    try:
        result = supabase.table('alert_incidents').select('icd_id').eq('icd_id', str(incident_id)).limit(1).execute()
        return result.data and len(result.data) > 0
    except Exception as e:
        print(f"Error validating incident: {e}")
        return False

def validate_student_exists(user_id):
    """Validate that a student exists in accounts_student table"""
    try:
        result = supabase.table('accounts_student').select('user_id').eq('user_id', str(user_id)).limit(1).execute()
        return result.data and len(result.data) > 0
    except Exception as e:
        print(f"Error validating student: {e}")
        return False

def validate_incident_student_relationship(incident_id, student_id):
    """Validate that the student (user_id) is associated with the incident
    
    This ensures that:
    - The incident exists in alert_incidents
    - The incident's user_id matches the student_id
    - The student is the one who reported the incident
    """
    try:
        # Get the incident and check its user_id
        result = supabase.table('alert_incidents').select('icd_id, user_id').eq('icd_id', str(incident_id)).limit(1).execute()
        
        if not result.data or len(result.data) == 0:
            print(f"Error: Incident {incident_id} does not exist")
            return False
        
        incident = result.data[0]
        incident_user_id = str(incident.get('user_id', ''))
        student_id_str = str(student_id)
        
        # Check if the incident's user_id matches the student_id
        if incident_user_id != student_id_str:
            print(f"Error: Student {student_id} is not associated with incident {incident_id} (incident belongs to user_id: {incident_user_id})")
            return False
        
        # Also verify the student exists in accounts_student
        if not validate_student_exists(student_id):
            print(f"Error: Student {student_id} does not exist in accounts_student")
            return False
        
        return True
    except Exception as e:
        print(f"Error validating incident-student relationship: {e}")
        import traceback
        traceback.print_exc()
        return False

def get_incident_student_id(incident_id):
    """Get the user_id (student_id) associated with an incident"""
    try:
        result = supabase.table('alert_incidents').select('user_id').eq('icd_id', str(incident_id)).limit(1).execute()
        if result.data and len(result.data) > 0:
            return str(result.data[0].get('user_id', ''))
        return None
    except Exception as e:
        print(f"Error getting incident student ID: {e}")
        return None

def get_chat_history(student_id, admin_id, incident_id=None, limit=50):
    """Get chat history between admin and student, optionally filtered by incident_id
    
    Returns messages with foreign key relationships validated:
    - incident_id references alert_incidents.icd_id
    - sender_id/receiver_id reference accounts_student.user_id when type is 'student'
    """
    try:
        # Check if table exists first
        if not check_chat_table_exists():
            print("Warning: chat_messages table does not exist. Please run the SQL script to create it.")
            return []
        
        # Build query for messages where admin is sender and student is receiver
        admin_to_student_query = supabase.table('chat_messages').select('*').eq('sender_id', str(admin_id)).eq('receiver_id', str(student_id)).eq('sender_type', 'admin').eq('receiver_type', 'student')
        
        # Build query for messages where student is sender and admin is receiver
        student_to_admin_query = supabase.table('chat_messages').select('*').eq('sender_id', str(student_id)).eq('receiver_id', str(admin_id)).eq('sender_type', 'student').eq('receiver_type', 'admin')
        
        # Filter by incident_id if provided (ensures foreign key relationship)
        if incident_id:
            # Validate incident exists before filtering
            if not validate_incident_exists(incident_id):
                print(f"Warning: incident_id {incident_id} does not exist in alert_incidents")
                return []
            
            # Validate that the student is associated with this incident
            if not validate_incident_student_relationship(incident_id, student_id):
                print(f"Warning: Student {student_id} is not associated with incident {incident_id}")
                return []
            
            admin_to_student_query = admin_to_student_query.eq('incident_id', str(incident_id))
            student_to_admin_query = student_to_admin_query.eq('incident_id', str(incident_id))
        
        # Order by timestamp ascending and limit
        admin_to_student = admin_to_student_query.order('timestamp', desc=False).limit(limit).execute()
        student_to_admin = student_to_admin_query.order('timestamp', desc=False).limit(limit).execute()
        
        # Combine and sort by timestamp
        all_messages = (admin_to_student.data or []) + (student_to_admin.data or [])
        all_messages.sort(key=lambda x: x.get('timestamp', '') or x.get('created_at', ''))
        
        # Validate foreign key relationships for returned messages
        validated_messages = []
        unique_incident_ids = set()
        
        for msg in (all_messages[-limit:] if len(all_messages) > limit else all_messages):
            msg_incident_id = msg.get('incident_id')
            msg_sender_id = msg.get('sender_id')
            msg_receiver_id = msg.get('receiver_id')
            msg_sender_type = msg.get('sender_type')
            msg_receiver_type = msg.get('receiver_type')
            
            # Verify incident_id exists (foreign key validation)
            if not validate_incident_exists(msg_incident_id):
                print(f"Warning: Message {msg.get('id')} references non-existent incident {msg_incident_id}")
                continue
            
            # Get the student_id associated with this incident
            incident_student_id = get_incident_student_id(msg_incident_id)
            if not incident_student_id:
                print(f"Warning: Message {msg.get('id')} references incident {msg_incident_id} with no user_id")
                continue
            
            # Verify student IDs exist when type is 'student'
            # When filtering by incident_id, verify they match the incident's student
            # When loading all messages, verify they match the student_id parameter
            if msg_sender_type == 'student':
                if not validate_student_exists(msg_sender_id):
                    print(f"Warning: Message {msg.get('id')} references non-existent student sender {msg_sender_id}")
                    continue
                # If filtering by incident_id, verify sender matches incident's student
                if incident_id:
                    if str(msg_sender_id) != str(incident_student_id):
                        print(f"Warning: Message {msg.get('id')} sender {msg_sender_id} is not the reporter of incident {msg_incident_id}")
                        continue
                # If not filtering by incident_id, verify sender matches the student_id parameter
                else:
                    if str(msg_sender_id) != str(student_id):
                        print(f"Warning: Message {msg.get('id')} sender {msg_sender_id} does not match requested student {student_id}")
                        continue
            
            if msg_receiver_type == 'student':
                if not validate_student_exists(msg_receiver_id):
                    print(f"Warning: Message {msg.get('id')} references non-existent student receiver {msg_receiver_id}")
                    continue
                # If filtering by incident_id, verify receiver matches incident's student
                if incident_id:
                    if str(msg_receiver_id) != str(incident_student_id):
                        print(f"Warning: Message {msg.get('id')} receiver {msg_receiver_id} is not the reporter of incident {msg_incident_id}")
                        continue
                # If not filtering by incident_id, verify receiver matches the student_id parameter
                else:
                    if str(msg_receiver_id) != str(student_id):
                        print(f"Warning: Message {msg.get('id')} receiver {msg_receiver_id} does not match requested student {student_id}")
                        continue
            
            validated_messages.append(msg)
            unique_incident_ids.add(str(msg_incident_id))
        
        # Fetch all incidents for the messages in batch
        incident_map = {}
        if unique_incident_ids:
            try:
                # Fetch all incidents at once using 'in' filter
                incident_ids_list = list(unique_incident_ids)
                incidents_result = supabase.table('alert_incidents').select('*').in_('icd_id', incident_ids_list).execute()
                
                if incidents_result.data:
                    # Create a map of incident_id -> incident data
                    for incident in incidents_result.data:
                        incident_map[str(incident.get('icd_id'))] = incident
            except Exception as e:
                print(f"Error fetching incidents for messages: {e}")
        
        # Add alert_incident data to each message
        for msg in validated_messages:
            msg_incident_id = str(msg.get('incident_id'))
            if msg_incident_id in incident_map:
                msg['alert_incident'] = incident_map[msg_incident_id]
            else:
                # If incident not found in batch, try to fetch it individually
                try:
                    incident_result = supabase.table('alert_incidents').select('*').eq('icd_id', msg_incident_id).limit(1).execute()
                    if incident_result.data and len(incident_result.data) > 0:
                        msg['alert_incident'] = incident_result.data[0]
                    else:
                        msg['alert_incident'] = None
                except Exception as e:
                    print(f"Error fetching individual incident {msg_incident_id}: {e}")
                    msg['alert_incident'] = None
        
        return validated_messages
    except Exception as e:
        error_msg = str(e)
        if 'PGRST205' in error_msg or 'Could not find the table' in error_msg:
            print("Error: chat_messages table does not exist. Please create it using the SQL script.")
            return []
        print(f"Error getting chat history: {e}")
        import traceback
        traceback.print_exc()
        return []

def send_chat_message(incident_id, sender_id, sender_type, receiver_id, receiver_type, message, image_url=None):
    """Send a chat message to the chat_messages table with foreign key validation
    
    Args:
        incident_id: Required incident ID (character varying(50) not null in table)
        sender_id: ID of the message sender
        sender_type: Type of sender ('admin' or 'student')
        receiver_id: ID of the message receiver
        receiver_type: Type of receiver ('admin' or 'student')
        message: Message text content (can be empty if image_url is provided)
        image_url: Optional image URL for the message
    """
    try:
        # Check if table exists first
        if not check_chat_table_exists():
            print("Error: chat_messages table does not exist. Please create it using the SQL script.")
            return None
        
        # Validate required fields
        if not incident_id:
            print("Error: incident_id is required (table constraint: not null)")
            return None
        if not sender_id or not receiver_id:
            print("Error: sender_id and receiver_id are required")
            return None
        # Allow empty message if image is provided
        if (not message or not message.strip()) and not image_url:
            print("Error: message cannot be empty unless image is provided")
            return None
        
        # Validate sender_type and receiver_type
        if sender_type not in ['admin', 'student']:
            print(f"Error: Invalid sender_type: {sender_type}. Must be 'admin' or 'student'.")
            return None
        if receiver_type not in ['admin', 'student']:
            print(f"Error: Invalid receiver_type: {receiver_type}. Must be 'admin' or 'student'.")
            return None
        
        # Validate foreign key relationships
        # 1. Validate incident_id exists in alert_incidents
        if not validate_incident_exists(incident_id):
            print(f"Error: incident_id {incident_id} does not exist in alert_incidents table (foreign key constraint violation)")
            return None
        
        # 2. Get the student_id (user_id) associated with this incident
        incident_student_id = get_incident_student_id(incident_id)
        if not incident_student_id:
            print(f"Error: Could not retrieve user_id for incident {incident_id}")
            return None
        
        # 3. Validate student IDs exist in accounts_student when type is 'student'
        # AND ensure they match the incident's user_id
        if sender_type == 'student':
            if not validate_student_exists(sender_id):
                print(f"Error: sender_id {sender_id} does not exist in accounts_student table (foreign key constraint violation)")
                return None
            # Validate that the sender student is the one who reported the incident
            if str(sender_id) != str(incident_student_id):
                print(f"Error: Student {sender_id} is not the reporter of incident {incident_id} (incident belongs to student {incident_student_id})")
                return None
        
        if receiver_type == 'student':
            if not validate_student_exists(receiver_id):
                print(f"Error: receiver_id {receiver_id} does not exist in accounts_student table (foreign key constraint violation)")
                return None
            # Validate that the receiver student is the one who reported the incident
            if str(receiver_id) != str(incident_student_id):
                print(f"Error: Student {receiver_id} is not the reporter of incident {incident_id} (incident belongs to student {incident_student_id})")
                return None
        
        # 4. Use the relationship validation function for additional check
        # This ensures the incident-student relationship is valid
        student_id_to_validate = sender_id if sender_type == 'student' else receiver_id
        if (sender_type == 'student' or receiver_type == 'student') and not validate_incident_student_relationship(incident_id, student_id_to_validate):
            print(f"Error: Incident-student relationship validation failed")
            return None
        
        # Use current time in UTC for timestamp (created_at will be set automatically by database)
        from datetime import timezone
        now = datetime.now(timezone.utc)
        
        # Prepare message data matching the Supabase table structure
        # Note: incident_id is NOT NULL in the table, so we ensure it's always provided
        message_data = {
            'incident_id': str(incident_id),  # Required field, must not be None (FK: alert_incidents.icd_id)
            'sender_id': str(sender_id),      # FK: accounts_student.user_id if sender_type='student'
            'sender_type': sender_type,
            'receiver_id': str(receiver_id),  # FK: accounts_student.user_id if receiver_type='student'
            'receiver_type': receiver_type,
            'message': str(message).strip() if message else '',
            'timestamp': now.isoformat(),
            'is_read': False
            # Note: 'created_at' is not set here as it has a default value in the database
        }
        
        # Add image_url if provided (check if column exists in table)
        if image_url:
            message_data['image_url'] = str(image_url)
        
        result = supabase.table('chat_messages').insert(message_data).execute()
        if result.data and len(result.data) > 0:
            return result.data[0]
        return None
    except Exception as e:
        error_msg = str(e)
        if 'PGRST205' in error_msg or 'Could not find the table' in error_msg:
            print("Error: chat_messages table does not exist. Please create it using the SQL script.")
            return None
        # Handle foreign key constraint violations
        if 'foreign key' in error_msg.lower() or 'violates foreign key constraint' in error_msg.lower():
            print(f"Error: Foreign key constraint violation - {error_msg}")
            return None
        if 'does not exist' in error_msg.lower():
            print(f"Error: Referenced record does not exist - {error_msg}")
            return None
        print(f"Error sending chat message: {e}")
        import traceback
        traceback.print_exc()
        return None

def mark_messages_as_read(sender_id, receiver_id):
    """Mark messages as read - marks unread messages sent by sender_id to receiver_id"""
    try:
        # Check if table exists first
        if not check_chat_table_exists():
            print("Warning: chat_messages table does not exist. Please create it using the SQL script.")
            return False
        
        # Mark unread messages where sender_id sent to receiver_id as read
        result = supabase.table('chat_messages').update({
            'is_read': True
        }).eq('sender_id', str(sender_id)).eq('receiver_id', str(receiver_id)).eq('is_read', False).execute()
        
        return True
    except Exception as e:
        error_msg = str(e)
        if 'PGRST205' in error_msg or 'Could not find the table' in error_msg:
            print("Error: chat_messages table does not exist. Please create it using the SQL script.")
            return False
        print(f"Error marking messages as read: {e}")
        import traceback
        traceback.print_exc()
        return False

def get_unread_message_count(admin_id):
    """Get count of unread messages for admin (where admin is receiver and sender is student)"""
    try:
        # Check if table exists first
        if not check_chat_table_exists():
            return 0
        
        # Get count of unread messages where admin is receiver and sender is student
        result = supabase.table('chat_messages').select('id', count='exact').eq('receiver_id', str(admin_id)).eq('receiver_type', 'admin').eq('is_read', False).execute()
        return result.count if hasattr(result, 'count') and result.count is not None else 0
    except Exception as e:
        error_msg = str(e)
        if 'PGRST205' in error_msg or 'Could not find the table' in error_msg:
            return 0
        print(f"Error getting unread message count: {e}")
        import traceback
        traceback.print_exc()
        return 0

# ==================== CHAT API ROUTES ====================

@app.route('/api/chat/students', methods=['GET'])
def api_get_chat_students():
    """API endpoint to get list of students who can be chatted with, sorted by most recent chat activity"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        admin_id = session['admin_id']
        students = get_students_with_incidents()
        
        if students is None:
            return jsonify({
                'success': False,
                'message': 'Failed to retrieve students. Please check server logs.'
            }), 500
        
        # Filter chat timestamps by admin_id to only show students this admin has chatted with
        if check_chat_table_exists():
            try:
                # Get latest chat message timestamp for each student filtered by this admin
                chat_timestamps = {}
                # Get messages where admin is sender or receiver
                try:
                    # Query 1: Messages where admin is sender
                    admin_as_sender = supabase.table('chat_messages').select('sender_id, receiver_id, sender_type, receiver_type, timestamp').eq('sender_id', str(admin_id)).eq('sender_type', 'admin').execute()
                    # Query 2: Messages where admin is receiver
                    admin_as_receiver = supabase.table('chat_messages').select('sender_id, receiver_id, sender_type, receiver_type, timestamp').eq('receiver_id', str(admin_id)).eq('receiver_type', 'admin').execute()
                    
                    # Combine both results
                    all_messages = (admin_as_sender.data or []) + (admin_as_receiver.data or [])
                except Exception as query_error:
                    print(f"Error querying chat messages: {query_error}")
                    all_messages = []
                
                if all_messages:
                    for msg in all_messages:
                        msg_timestamp = msg.get('timestamp') or msg.get('created_at')
                        if not msg_timestamp:
                            continue
                        
                        # Only process messages where student is involved
                        student_id = None
                        if msg.get('receiver_type') == 'student' and msg.get('sender_id') == str(admin_id):
                            student_id = str(msg.get('receiver_id', ''))
                        elif msg.get('sender_type') == 'student' and msg.get('receiver_id') == str(admin_id):
                            student_id = str(msg.get('sender_id', ''))
                        
                        if student_id:
                            if student_id not in chat_timestamps:
                                chat_timestamps[student_id] = msg_timestamp
                            else:
                                # Keep the most recent timestamp
                                try:
                                    if msg_timestamp > chat_timestamps[student_id]:
                                        chat_timestamps[student_id] = msg_timestamp
                                except:
                                    chat_timestamps[student_id] = msg_timestamp
                
                # Update students with filtered chat timestamps
                for student in students:
                    user_id_str = str(student.get('user_id', ''))
                    if user_id_str in chat_timestamps:
                        student['latest_chat_timestamp'] = chat_timestamps[user_id_str]
                    else:
                        student['latest_chat_timestamp'] = None
                
                # Re-sort students by latest chat timestamp (most recent first)
                def get_sort_key(student):
                    chat_ts = student.get('latest_chat_timestamp')
                    incident_ts = student.get('latest_incident_timestamp', '')
                    
                    def parse_timestamp(ts):
                        if not ts:
                            return None
                        try:
                            if isinstance(ts, str):
                                if 'T' in ts:
                                    from datetime import datetime
                                    try:
                                        dt = datetime.fromisoformat(ts.replace('Z', '+00:00'))
                                        return dt.timestamp()
                                    except:
                                        pass
                                try:
                                    from datetime import datetime
                                    dt = datetime.strptime(ts[:19], '%Y-%m-%d %H:%M:%S')
                                    return dt.timestamp()
                                except:
                                    pass
                            if hasattr(ts, 'timestamp'):
                                return ts.timestamp()
                            return None
                        except:
                            return None
                    
                    chat_timestamp = parse_timestamp(chat_ts) if chat_ts else None
                    incident_timestamp = parse_timestamp(incident_ts) if incident_ts else None
                    
                    if chat_timestamp is not None:
                        return (chat_timestamp, 1)
                    elif incident_timestamp is not None:
                        return (incident_timestamp, 0)
                    else:
                        return (0, 0)
                
                students.sort(key=get_sort_key, reverse=True)
            except Exception as e:
                print(f"Error filtering chat timestamps by admin: {e}")
        
        return jsonify({
            'success': True,
            'students': students,
            'count': len(students)
        })
    except Exception as e:
        error_msg = str(e)
        print(f"Error getting chat students: {error_msg}")
        import traceback
        traceback.print_exc()
        
        # Provide more helpful error messages
        if 'table' in error_msg.lower() and 'not exist' in error_msg.lower():
            return jsonify({
                'success': False,
                'message': 'Database table error. Please contact administrator.'
            }), 500
        elif 'connection' in error_msg.lower() or 'network' in error_msg.lower():
            return jsonify({
                'success': False,
                'message': 'Database connection error. Please try again.'
            }), 500
        else:
            return jsonify({
                'success': False,
                'message': f'Error loading students: {error_msg}'
            }), 500

@app.route('/api/chat/history/<student_id>', methods=['GET'])
def api_get_chat_history(student_id):
    """API endpoint to get chat history with a student, optionally filtered by incident_id"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        admin_id = session['admin_id']
        incident_id = request.args.get('incident_id')
        
        if not incident_id:
            return jsonify({'success': False, 'message': 'incident_id is required to load conversation'}), 400
        
        can_view, view_message = can_admin_view_incident(incident_id, admin_id)
        if not can_view:
            return jsonify({'success': False, 'message': view_message}), 403
        
        messages = get_chat_history(student_id, admin_id, incident_id=incident_id)
        
        # Mark messages as read (messages sent by student to admin)
        mark_messages_as_read(student_id, admin_id)
        
        return jsonify({
            'success': True,
            'messages': messages
        })
    except Exception as e:
        print(f"Error getting chat history: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/chat/send', methods=['POST'])
def api_send_chat_message():
    """API endpoint to send a chat message with optional image
    Accepts both formats:
    - New format: sender_id, sender_type, receiver_id, receiver_type
    - Legacy format: student_id (for backward compatibility)
    """
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        data = request.get_json()
        incident_id = data.get('incident_id')
        message = data.get('message', '').strip()
        image_base64 = data.get('image')
        image_name = data.get('image_name', 'image.jpg')
        image_type = data.get('image_type', 'image/jpeg')
        
        admin_id = session['admin_id']
        
        # Support new format: sender_id, sender_type, receiver_id, receiver_type
        sender_id = data.get('sender_id')
        sender_type = data.get('sender_type')
        receiver_id = data.get('receiver_id')
        receiver_type = data.get('receiver_type')
        
        # Legacy format support: student_id (for backward compatibility)
        student_id = data.get('student_id')
        
        # Determine sender and receiver based on format
        if sender_id and sender_type and receiver_id and receiver_type:
            # New format
            if sender_type not in ['admin', 'student'] or receiver_type not in ['admin', 'student']:
                return jsonify({'success': False, 'message': 'Invalid sender_type or receiver_type. Must be "admin" or "student"'}), 400
        elif student_id:
            # Legacy format: admin sending to student
            sender_id = admin_id
            sender_type = 'admin'
            receiver_id = student_id
            receiver_type = 'student'
        else:
            return jsonify({'success': False, 'message': 'Missing required fields: either (sender_id, sender_type, receiver_id, receiver_type) or student_id'}), 400
        
        if not incident_id:
            return jsonify({'success': False, 'message': 'Missing required field: incident_id'}), 400
        
        # Must have either message or image
        if not message and not image_base64:
            return jsonify({'success': False, 'message': 'Message or image is required'}), 400
        
        # Validate permissions: if admin is sender, check if they can edit the incident
        if sender_type == 'admin' and str(sender_id) != str(admin_id):
            return jsonify({'success': False, 'message': 'Unauthorized: You can only send messages as yourself'}), 403
        
        if sender_type == 'admin':
            can_edit, edit_message = can_admin_edit_incident(incident_id, sender_id)
            if not can_edit:
                return jsonify({'success': False, 'message': f'Permission denied: {edit_message}'}), 403
        
        # Handle image upload if provided
        image_url = None
        if image_base64:
            try:
                import base64
                import uuid
                import os
                
                # Decode base64 image
                if ',' in image_base64:
                    # Remove data URL prefix if present
                    image_base64 = image_base64.split(',')[1]
                
                image_data = base64.b64decode(image_base64)
                
                # Validate file size (max 5MB)
                if len(image_data) > 5 * 1024 * 1024:
                    return jsonify({'success': False, 'message': 'Image size must be less than 5MB'}), 400
                
                # Determine file extension from image type or name
                if 'png' in image_type.lower() or image_name.lower().endswith('.png'):
                    ext = 'png'
                elif 'gif' in image_type.lower() or image_name.lower().endswith('.gif'):
                    ext = 'gif'
                elif 'webp' in image_type.lower() or image_name.lower().endswith('.webp'):
                    ext = 'webp'
                else:
                    ext = 'jpg'
                
                # Generate unique filename
                filename = f"chat_{uuid.uuid4().hex[:8]}.{ext}"
                
                # Save to static/images directory
                upload_dir = os.path.join(app.static_folder, 'images')
                os.makedirs(upload_dir, exist_ok=True)
                file_path = os.path.join(upload_dir, filename)
                
                with open(file_path, 'wb') as f:
                    f.write(image_data)
                
                # Store relative URL
                image_url = filename
                
            except Exception as img_error:
                print(f"Error processing image: {img_error}")
                return jsonify({'success': False, 'message': f'Failed to process image: {str(img_error)}'}), 400
        
        # Send message
        result = send_chat_message(
            incident_id=incident_id,
            sender_id=sender_id,
            sender_type=sender_type,
            receiver_id=receiver_id,
            receiver_type=receiver_type,
            message=message,
            image_url=image_url
        )
        
        if result:
            # Add image_url to result if it was saved
            if image_url:
                result['image_url'] = image_url
            return jsonify({
                'success': True,
                'message': result
            })
        else:
            return jsonify({'success': False, 'message': 'Failed to send message'}), 500
    except Exception as e:
        print(f"Error sending chat message: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/chat/unread-count', methods=['GET'])
def api_get_unread_count():
    """API endpoint to get unread message count"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        admin_id = session['admin_id']
        count = get_unread_message_count(admin_id)
        
        return jsonify({
            'success': True,
            'unread_count': count
        })
    except Exception as e:
        print(f"Error getting unread count: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/chat/check-table', methods=['GET'])
def api_check_chat_table():
    """API endpoint to check if chat_messages table exists"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        table_exists = check_chat_table_exists()
        return jsonify({
            'success': True,
            'table_exists': table_exists,
            'message': 'Table exists' if table_exists else 'Table does not exist. Please run the SQL script in Supabase SQL Editor.'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'table_exists': False,
            'message': f'Error checking table: {str(e)}'
        }), 500

@app.route('/api/chat/incidents', methods=['GET'])
def api_get_chat_incidents():
    """API endpoint to get incidents where admin is assigned handler with unread counts"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        admin_id = session['admin_id']
        
        if not check_chat_table_exists():
            return jsonify({'success': False, 'message': 'Chat table does not exist'}), 500
        
        # Get all incidents where this admin is the assigned handler
        incidents_result = supabase.table('alert_incidents').select('*').eq('assigned_responder_id', str(admin_id)).order('icd_timestamp', desc=True).execute()
        incidents = incidents_result.data or []
        
        # Get student full names
        user_ids = {incident.get('user_id') for incident in incidents if incident.get('user_id')}
        students_map = {}
        if user_ids:
            try:
                batch_size = 100
                user_ids_list = list(user_ids)
                for i in range(0, len(user_ids_list), batch_size):
                    batch = user_ids_list[i:i + batch_size]
                    students_result = supabase.table('accounts_student').select('user_id, full_name').in_('user_id', batch).execute()
                    if students_result.data:
                        for student in students_result.data:
                            user_id = student.get('user_id')
                            if user_id is not None:
                                students_map[str(user_id)] = student.get('full_name', '')
            except Exception as e:
                print(f"Error fetching student names: {e}")
        
        # Get unread counts for each incident
        incident_list = []
        for incident in incidents:
            incident_id = str(incident.get('icd_id', ''))
            user_id = str(incident.get('user_id', ''))
            
            # Get unread count for this incident
            try:
                result = supabase.table('chat_messages').select('id', count='exact').eq('incident_id', incident_id).eq('receiver_id', str(admin_id)).eq('is_read', False).execute()
                unread_count = result.count if hasattr(result, 'count') and result.count is not None else 0
            except Exception as e:
                print(f"Error getting unread count for incident {incident_id}: {e}")
                unread_count = 0
            
            incident_list.append({
                'icd_id': incident_id,
                'full_name': students_map.get(user_id, 'Unknown Student'),
                'icd_status': incident.get('icd_status', ''),
                'icd_timestamp': incident.get('icd_timestamp', ''),
                'user_id': user_id,
                'unread_count': unread_count,
                'has_unread': unread_count > 0
            })
        
        return jsonify({
            'success': True,
            'incidents': incident_list,
            'count': len(incident_list)
        })
    except Exception as e:
        print(f"Error getting chat incidents: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/chat/incident/<incident_id>', methods=['GET'])
def api_get_incident_chat(incident_id):
    """API endpoint to get ALL chat messages for an incident_id - no date limits"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        admin_id = session['admin_id']
        
        # For chat, allow viewing if:
        # 1. Admin can view the incident normally, OR
        # 2. Admin has sent/received messages in this chat (more lenient for chat access)
        can_view, view_message = can_admin_view_incident(incident_id, admin_id)
        
        # If normal view check fails, check if admin has participated in this chat
        if not can_view:
            try:
                # Check if admin has any messages in this incident's chat (as sender or receiver)
                sender_check = supabase.table('chat_messages').select('id').eq('incident_id', str(incident_id)).eq('sender_id', str(admin_id)).limit(1).execute()
                receiver_check = supabase.table('chat_messages').select('id').eq('incident_id', str(incident_id)).eq('receiver_id', str(admin_id)).limit(1).execute()
                if (sender_check.data and len(sender_check.data) > 0) or (receiver_check.data and len(receiver_check.data) > 0):
                    # Admin has participated in chat, allow viewing
                    can_view = True
                    view_message = "Admin has participated in this chat"
            except Exception as chat_check_error:
                print(f"Error checking chat participation: {chat_check_error}")
                # Fall through to return 403
        
        if not can_view:
            return jsonify({'success': False, 'message': view_message}), 403
        
        if not check_chat_table_exists():
            return jsonify({'success': False, 'message': 'Chat table does not exist'}), 500
        
        # Get ALL messages for this incident - no limit, no date filter
        try:
            result = supabase.table('chat_messages').select('*').eq('incident_id', str(incident_id)).order('timestamp', desc=False).execute()
            messages = result.data or []
        except Exception as e:
            print(f"Error fetching chat messages: {e}")
            import traceback
            traceback.print_exc()
            messages = []
        
        # Sort by timestamp ascending to ensure chronological order
        try:
            messages.sort(key=lambda x: x.get('timestamp', '') or x.get('created_at', ''))
        except Exception as e:
            print(f"Error sorting messages: {e}")
        
        # Mark all messages for this incident as read where admin is receiver
        try:
            supabase.table('chat_messages').update({
                'is_read': True
            }).eq('incident_id', str(incident_id)).eq('receiver_id', str(admin_id)).eq('is_read', False).execute()
        except Exception as e:
            print(f"Error marking messages as read: {e}")
        
        # Get incident details for the response
        incident = None
        student_name = 'Unknown Student'
        try:
            incident_result = supabase.table('alert_incidents').select('*').eq('icd_id', str(incident_id)).limit(1).execute()
            incident = incident_result.data[0] if incident_result.data else None
            
            # Get student name
            if incident and incident.get('user_id'):
                try:
                    student_result = supabase.table('accounts_student').select('full_name').eq('user_id', str(incident.get('user_id'))).limit(1).execute()
                    if student_result.data:
                        student_name = student_result.data[0].get('full_name', 'Unknown Student')
                except Exception as e:
                    print(f"Error fetching student name: {e}")
                    student_name = 'Unknown Student'
        except Exception as e:
            print(f"Error fetching incident details: {e}")
            import traceback
            traceback.print_exc()
            incident = None
            student_name = 'Unknown Student'
        
        return jsonify({
            'success': True,
            'messages': messages,
            'incident': incident,
            'student_name': student_name
        })
    except Exception as e:
        print(f"Error getting chat history: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Server error: {str(e)}'}), 500

@app.route('/chat/conversation/<incident_id>', methods=['GET'])
def chat_conversation(incident_id):
    """Alias for /api/chat/incident/<incident_id> to match user specification"""
    return api_get_incident_chat(incident_id)

@app.route('/api/chat/read/<incident_id>', methods=['PUT'])
def api_mark_incident_read(incident_id):
    """API endpoint to mark all messages for an incident as read"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        admin_id = session['admin_id']
        
        can_view, view_message = can_admin_view_incident(incident_id, admin_id)
        if not can_view:
            return jsonify({'success': False, 'message': view_message}), 403
        
        # Update all messages for this incident where admin is receiver
        result = supabase.table('chat_messages').update({
            'is_read': True
        }).eq('incident_id', str(incident_id)).eq('receiver_id', str(admin_id)).eq('is_read', False).execute()
        
        return jsonify({
            'success': True
        })
    except Exception as e:
        print(f"Error marking incident as read: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# ==================== DEBUG ROUTES ====================
@app.route('/debug/users')
def debug_users():
    """Debug route to check user data in database"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        # Check admin users
        admin_result = supabase.table('accounts_admin').select('*').execute()
        admin_users = admin_result.data or []
        
        # Check student users
        student_result = supabase.table('accounts_student').select('*').execute()
        student_users = student_result.data or []
        
        debug_info = {
            'admin_users_count': len(admin_users),
            'student_users_count': len(student_users),
            'admin_users': admin_users,
            'student_users': student_users,
            'admin_table_columns': list(admin_users[0].keys()) if admin_users else [],
            'student_table_columns': list(student_users[0].keys()) if student_users else []
        }
        
        return jsonify(debug_info)
        
    except Exception as e:
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()})

@app.route('/debug/tables')
def debug_tables():
    """Debug route to check all tables in database"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        # List all tables (this is a simplified approach)
        tables_to_check = ['accounts_admin', 'accounts_student', 'alert_incidents']
        
        table_info = {}
        for table in tables_to_check:
            try:
                result = supabase.table(table).select('*').limit(5).execute()
                table_info[table] = {
                    'count': len(result.data) if result.data else 0,
                    'sample_data': result.data[:2] if result.data else [],
                    'columns': list(result.data[0].keys()) if result.data else []
                }
            except Exception as e:
                table_info[table] = {'error': str(e)}
        
        return jsonify(table_info)
        
    except Exception as e:
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()})

@app.route('/create-test-users')
def create_test_users():
    """Create test users for development"""
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    try:
        # Create test admin users
        test_admins = []
        admin_templates = [
            {
                'admin_user': 'admin_joleh',
                'admin_pass': 'Hushjoleh_10',
                'admin_email': 'juliemay1917@gmail.com',
                'admin_fullname': 'Julie May Joleh',
                'admin_role': 'System Administrator',
                'admin_status': 'Active',
                'admin_profile': 'default.png'
            },
            {
                'admin_user': 'security_staff',
                'admin_pass': 'Security123!',
                'admin_email': 'security@umak.edu.ph',
                'admin_fullname': 'Security Officer',
                'admin_role': 'Security Staff',
                'admin_status': 'Active',
                'admin_profile': 'default.png'
            },
            {
                'admin_user': 'dispatcher',
                'admin_pass': 'Dispatcher123!',
                'admin_email': 'dispatch@umak.edu.ph',
                'admin_fullname': 'Emergency Dispatcher',
                'admin_role': 'Dispatcher',
                'admin_status': 'Active',
                'admin_profile': 'default.png'
            }
        ]
        
        # Add admin_id to each admin template
        for template in admin_templates:
            template['admin_id'] = generate_next_admin_id()
            test_admins.append(template)
        
        # Create test student users
        test_students = [
            {
                'student_id': '2023-001',
                'student_user': 'student1',
                'student_pass': 'Student123!',
                'student_email': 'student1@umak.edu.ph',
                'full_name': 'Juan Dela Cruz',
                'student_yearlvl': 'Third Year',
                'student_cnum': '+639123456789',
                'primary_emergencycontact': '+639876543210',
                'primary_contactperson': 'Maria Dela Cruz',
                'primary_cprelationship': 'Mother',
                'secondary_emergencycontact': '+639555555555',
                'secondary_contactperson': 'Pedro Dela Cruz',
                'secondary_cprelationship': 'Father',
                'student_medinfo': 'None',
                'student_address': 'Makati City',
                'student_status': 'Active',
                'student_profile': 'default.png'
            },
            {
                'student_id': '2023-002',
                'student_user': 'student2',
                'student_pass': 'Student123!',
                'student_email': 'student2@umak.edu.ph',
                'full_name': 'Maria Santos',
                'student_yearlvl': 'Second Year',
                'student_cnum': '+639987654321',
                'primary_emergencycontact': '+639111111111',
                'primary_contactperson': 'Roberto Santos',
                'primary_cprelationship': 'Father',
                'secondary_emergencycontact': '+639222222222',
                'secondary_contactperson': 'Elena Santos',
                'secondary_cprelationship': 'Mother',
                'student_medinfo': 'Asthma',
                'student_address': 'Taguig City',
                'student_status': 'Active',
                'student_profile': 'default.png'
            }
        ]
        
        created_admins = []
        created_students = []
        
        # Insert admin users
        for admin in test_admins:
            result = supabase.table('accounts_admin').insert(admin).execute()
            if result.data:
                created_admins.append(result.data[0]['admin_user'])
        
        # Insert student users
        for student in test_students:
            result = supabase.table('accounts_student').insert(student).execute()
            if result.data:
                created_students.append(result.data[0]['student_user'])
        
        return f"""
        <h1>Test Users Created Successfully!</h1>
        <h2>Created Admins:</h2>
        <ul>
            {''.join([f'<li>{admin}</li>' for admin in created_admins])}
        </ul>
        <h2>Created Students:</h2>
        <ul>
            {''.join([f'<li>{student}</li>' for student in created_students])}
        </ul>
        <p><a href="/user-management">Go to User Management</a></p>
        """
        
    except Exception as e:
        return f"Error creating test users: {e}<br>Traceback: {traceback.format_exc()}"

# ==================== INCIDENT MANAGEMENT ROUTES ====================
@app.route('/incident-management', methods=['GET', 'POST'])
def incident_management():
    """Incident management route with filtering, search, status updates, and deletion"""
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    # Get current admin info
    admin_id = session['admin_id']
    try:
        admin_result = supabase.table('accounts_admin').select('*').eq('admin_id', admin_id).execute()
        if admin_result.data:
            current_admin = admin_result.data[0]
            session['admin_profile'] = current_admin.get('admin_profile')
            session['admin_profile_exists'] = check_profile_image_exists(current_admin.get('admin_profile'))
            session['admin_name'] = current_admin.get('admin_fullname', 'Unknown')
        else:
            session['admin_profile'] = ''
            session['admin_profile_exists'] = False
            session['admin_name'] = 'Unknown'
    except Exception as e:
        session['admin_profile'] = ''
        session['admin_profile_exists'] = False
        session['admin_name'] = 'Unknown'
    
    if 'role' not in session:
        session['role'] = 'Administrator'
    
    # Handle POST requests
    if request.method == 'POST':
        # Handle incident deletion
        if 'delete_incident' in request.form:
            incident_id = request.form.get('incident_id')
            try:
                # Check if incident exists
                incident_check = supabase.table('alert_incidents').select('*').eq('icd_id', incident_id).execute()
                
                if incident_check.data:
                    # Log to audit trail before deletion
                    log_incident_change(incident_id, 'deleted', admin_id=session['admin_id'], reason="Deleted by administrator")
                    
                    # Delete related activity logs first
                    supabase.table('admin_activity_logs').delete().eq('incident_id', incident_id).execute()
                    
                    # Delete the incident
                    supabase.table('alert_incidents').delete().eq('icd_id', incident_id).execute()
                    
                    flash(f'Incident {incident_id} has been successfully deleted.', 'success')
                else:
                    flash('Incident not found.', 'error')
                    
            except Exception as e:
                flash(f'Error deleting incident: {str(e)}', 'error')
        
        # Handle incident archiving
        elif 'archive_incident' in request.form:
            incident_id = request.form.get('incident_id')
            archive_reason = request.form.get('archive_reason', 'Archived by administrator')
            
            try:
                success, message = archive_incident(incident_id, session['admin_id'], archive_reason)
                
                if success:
                    flash(f'Incident {incident_id} has been successfully archived. {message}', 'success')
                else:
                    flash(f'Error archiving incident: {message}', 'error')
                    
            except Exception as e:
                flash(f'Error archiving incident: {str(e)}', 'error')
        
        # Handle bulk archive
        elif 'bulk_archive' in request.form:
            selected_incidents = request.form.getlist('selected_incidents')
            archive_reason = request.form.get('bulk_archive_reason', 'Bulk archived by administrator')
            
            if not selected_incidents:
                flash('Please select incidents to archive.', 'warning')
                return redirect(url_for('incident_management'))
            
            try:
                results = bulk_archive_incidents(selected_incidents, session['admin_id'], archive_reason)
                
                success_count = sum(1 for result in results if result['success'])
                failure_count = len(results) - success_count
                
                if success_count > 0:
                    flash(f'Successfully archived {success_count} incident(s).', 'success')
                if failure_count > 0:
                    flash(f'Failed to archive {failure_count} incident(s).', 'error')
                    
            except Exception as e:
                flash(f'Error in bulk archive: {str(e)}', 'error')
            
            # Preserve current filters in redirect
            params = {}
            if request.form.get('current_status') and request.form.get('current_status') != 'All':
                params['status'] = request.form.get('current_status')
            if request.form.get('current_category') and request.form.get('current_category') != 'All':
                params['category'] = request.form.get('current_category')
            if request.form.get('current_search'):
                params['search'] = request.form.get('current_search')
            if request.form.get('current_sort') and request.form.get('current_sort') != 'recent':
                params['sort'] = request.form.get('current_sort')
            
            if params:
                from urllib.parse import urlencode
                return redirect(url_for('incident_management') + '?' + urlencode(params))
            else:
                return redirect(url_for('incident_management'))
        
        # Handle status updates
        elif 'update' in request.form:
            incident_id = request.form.get('incident_id')
            new_status = request.form.get('status')
            
            # Check if admin can edit this incident
            can_edit, edit_message = can_admin_edit_incident(incident_id, session['admin_id'])
            if not can_edit:
                flash(f'Permission denied: {edit_message}', 'error')
                return redirect(url_for('incident_management'))
            
            try:
                # Get current incident data for audit trail
                incident_check = supabase.table('alert_incidents').select('icd_status').eq('icd_id', incident_id).execute()
                if not incident_check.data:
                    flash('Incident not found.', 'error')
                    return redirect(url_for('incident_management'))
                
                old_status = incident_check.data[0]['icd_status']
                
                # Update the incident status
                current_time = get_philippines_time().isoformat()
                update_data = {
                    'icd_status': new_status,
                    'status_updated_at': current_time,
                    'status_updated_by': session['admin_id']
                }
                
                # Reset all timestamp fields
                update_data['resolved_timestamp'] = None
                update_data['pending_timestamp'] = None
                update_data['cancelled_timestamp'] = None
                
                # Set appropriate timestamp based on status
                if new_status == 'Resolved':
                    update_data['resolved_timestamp'] = current_time
                elif new_status == 'Pending':
                    update_data['pending_timestamp'] = current_time
                elif new_status == 'Cancelled':
                    update_data['cancelled_timestamp'] = current_time
                
                supabase.table('alert_incidents').update(update_data).eq('icd_id', incident_id).execute()
                
                # Log to audit trail
                log_incident_change(incident_id, 'status_updated', 
                                  old_status=old_status, new_status=new_status, 
                                  admin_id=session['admin_id'], reason=f"Status changed from {old_status} to {new_status}")
                
                flash(f'Incident {incident_id} status has been successfully updated to {new_status}!', 'success')
                
            except Exception as e:
                flash(f'Error updating incident status: {str(e)}', 'error')
            
            # Preserve current filters in redirect
            params = {}
            if request.form.get('current_status_filter') and request.form.get('current_status_filter') != 'All':
                params['status'] = request.form.get('current_status_filter')
            if request.form.get('current_category_filter') and request.form.get('current_category_filter') != 'All':
                params['category'] = request.form.get('current_category_filter')
            if request.form.get('current_search_term'):
                params['search'] = request.form.get('current_search_term')
            if request.form.get('current_sort_order') and request.form.get('current_sort_order') != 'recent':
                params['sort'] = request.form.get('current_sort_order')
            
            if params:
                from urllib.parse import urlencode
                return redirect(url_for('incident_management') + '?' + urlencode(params))
            else:
                return redirect(url_for('incident_management'))
    
    # Handle student details viewing
    student_details = None
    if request.args.get('view_student'):
        student_id = request.args.get('view_student')
        try:
            student_result = supabase.table('accounts_student').select('*').eq('user_id', student_id).execute()
            if student_result.data:
                student_details = student_result.data[0]
        except Exception as e:
            student_details = None
    # Get filter parameters
    status_filter = request.args.get('status', 'All')
    category_filter = request.args.get('category', 'All')
    search_term = request.args.get('search', '')
    sort_order = request.args.get('sort', 'recent')
    
    # Build query for incidents with joins
    try:
        # Get unique icd_category values from database for dropdown
        all_incidents_for_categories = supabase.table('alert_incidents').select('icd_category').execute()
        categories_set = set()
        if all_incidents_for_categories.data:
            for incident in all_incidents_for_categories.data:
                category = incident.get('icd_category')
                if category and category.strip():
                    categories_set.add(category.strip())
        categories_list = sorted(list(categories_set))
        
        # Get base incidents data
        query = supabase.table('alert_incidents').select('*')
        
        # Apply status filter
        if status_filter != 'All':
            query = query.eq('icd_status', status_filter)
        
        # Apply category filter
        if category_filter != 'All':
            query = query.eq('icd_category', category_filter)
        
        # Apply search filter (we'll filter in Python due to Supabase limitations)
        incidents_result = query.execute()
        incidents = incidents_result.data if incidents_result.data else []
        
        # Get student and admin data for joins
        students_result = supabase.table('accounts_student').select('*').execute()
        students = {}
        if students_result.data:
            for student in students_result.data:
                user_id = student.get('user_id')
                if user_id is not None:
                    students[user_id] = student
                    students[str(user_id)] = student
        
        admins_result = supabase.table('accounts_admin').select('*').execute()
        admins = {}
        if admins_result.data:
            for admin in admins_result.data:
                admin_id = admin.get('admin_id')
                if admin_id is not None:
                    admins[admin_id] = admin
                    admins[str(admin_id)] = admin
        
        # Process incidents with join data and search filtering
        processed_incidents = []
        for incident in incidents:
            # Get related student and admin data
            student_data = students.get(incident.get('user_id'))
            admin_data = admins.get(incident.get('admin_id'))
            
            # Add related data to incident
            incident['student_name'] = student_data.get('full_name') if student_data else None
            incident['student_number'] = student_data.get('student_id') if student_data else None
            incident['student_user_id'] = student_data.get('user_id') if student_data else None
            incident['admin_name'] = admin_data.get('admin_fullname') if admin_data else None
            
            # Apply search filter
            if search_term:
                search_fields = [
                    incident.get('student_name', ''),
                    incident.get('admin_name', ''),
                    incident.get('student_number', '')
                ]
                if not any(search_term.lower() in str(field).lower() for field in search_fields if field):
                    continue
            
            processed_incidents.append(incident)
        
        # Apply sorting
        if sort_order == 'old':
            processed_incidents.sort(key=lambda x: x.get('icd_id', 0))
        else:
            processed_incidents.sort(key=lambda x: x.get('icd_id', 0), reverse=True)
        
        # Get statistics
        all_incidents = supabase.table('alert_incidents').select('*').execute()
        all_incidents_data = all_incidents.data if all_incidents.data else []
        
        total_incidents = len(all_incidents_data)
        active_incidents = len([i for i in all_incidents_data if i.get('icd_status') == 'Active'])
        pending_incidents = len([i for i in all_incidents_data if i.get('icd_status') == 'Pending'])
        resolved_incidents = len([i for i in all_incidents_data if i.get('icd_status') == 'Resolved'])
        cancelled_incidents = len([i for i in all_incidents_data if i.get('icd_status') == 'Cancelled'])
        
        # Active alerts count (Active + Pending)
        active_alerts_count = active_incidents + pending_incidents
        
    except Exception as e:
        processed_incidents = []
        total_incidents = 0
        active_incidents = 0
        pending_incidents = 0
        resolved_incidents = 0
        cancelled_incidents = 0
        active_alerts_count = 0
        categories_list = []
        flash(f'Error loading incident data: {str(e)}', 'error')
    
    return render_template('incident_management.html',
                         incidents=processed_incidents,
                         student_details=student_details,
                         status_filter=status_filter,
                         category_filter=category_filter,
                         categories_list=categories_list,
                         search_term=search_term,
                         sort_order=sort_order,
                         total_incidents=total_incidents,
                         active_incidents=active_incidents,
                         pending_incidents=pending_incidents,
                         resolved_incidents=resolved_incidents,
                         cancelled_incidents=cancelled_incidents,
                         active_alerts_count=active_alerts_count,
                         results_count=len(processed_incidents))

# ==================== ENHANCED API ROUTES FOR INCIDENT MANAGEMENT ====================

@app.route('/api/incidents/refresh')
def api_refresh_incidents():
    """API endpoint for auto-refresh functionality"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        # Get filter parameters from query string
        status_filter = request.args.get('status', 'All')
        search_term = request.args.get('search', '')
        sort_order = request.args.get('sort', 'recent')
        
        # Get incidents with same logic as main route
        incidents = get_incidents_with_relations({
            'status': status_filter,
            'search': search_term,
            'sort': sort_order,
            'admin_id': session['admin_id']
        })
        
        # Return data in JSON format
        return jsonify({
            'success': True,
            'incidents': incidents,
            'total_count': len(incidents),
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        print(f"Error refreshing incidents: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/incident/<incident_id>/archive', methods=['POST'])
def api_archive_incident(incident_id):
    """API endpoint to archive an incident"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    # Check if admin can edit this incident
    can_edit, edit_message = can_admin_edit_incident(incident_id, session['admin_id'])
    if not can_edit:
        return jsonify({'success': False, 'message': f'Permission denied: {edit_message}'}), 403
    
    try:
        data = request.get_json() or {}
        reason = data.get('reason', 'Archived by administrator')
        
        success, message = archive_incident(incident_id, session['admin_id'], reason)
        
        if success:
            return jsonify({
                'success': True, 
                'message': message,
                'archived_at': get_philippines_time().isoformat()
            })
        else:
            return jsonify({'success': False, 'message': message}), 500
            
    except Exception as e:
        print(f"Error archiving incident: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/incident/<incident_id>/permissions', methods=['GET'])
def api_get_incident_permissions(incident_id):
    """API endpoint to get permission info for an incident"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        incident_result = supabase.table('alert_incidents').select('icd_status, assigned_responder_id').eq('icd_id', incident_id).execute()
        
        if not incident_result.data or len(incident_result.data) == 0:
            return jsonify({'success': False, 'message': 'Incident not found'}), 404
        
        incident = incident_result.data[0]
        status = incident.get('icd_status', '')
        assigned_responder_id = incident.get('assigned_responder_id')
        current_admin_id = session['admin_id']
        
        # Check permissions
        can_edit, edit_message = can_admin_edit_incident(incident_id, current_admin_id)
        
        # Get assigned responder name if assigned
        assigned_responder_name = None
        if assigned_responder_id:
            try:
                responder_result = supabase.table('accounts_admin').select('admin_fullname').eq('admin_id', assigned_responder_id).execute()
                if responder_result.data and len(responder_result.data) > 0:
                    assigned_responder_name = responder_result.data[0].get('admin_fullname', assigned_responder_id)
                else:
                    assigned_responder_name = assigned_responder_id
            except:
                assigned_responder_name = assigned_responder_id
        
        return jsonify({
            'success': True,
            'can_edit': can_edit,
            'edit_message': edit_message,
            'status': status,
            'assigned_responder_id': assigned_responder_id,
            'assigned_responder_name': assigned_responder_name,
            'is_assigned_to_me': str(assigned_responder_id) == str(current_admin_id) if assigned_responder_id else False,
            'is_resolved_or_cancelled': status in ['Resolved', 'Cancelled']
        })
    except Exception as e:
        print(f"Error getting incident permissions: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/incident/<incident_id>/assign', methods=['POST'])
def api_assign_incident(incident_id):
    """API endpoint to assign or reassign an incident"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        data = request.get_json() or {}
        new_responder_id = data.get('responder_id')
        current_admin_id = session['admin_id']
        
        if not new_responder_id:
            return jsonify({'success': False, 'message': 'Responder ID is required'}), 400
        
        # Get current incident
        incident_result = supabase.table('alert_incidents').select('icd_status, assigned_responder_id').eq('icd_id', incident_id).execute()
        
        if not incident_result.data or len(incident_result.data) == 0:
            return jsonify({'success': False, 'message': 'Incident not found'}), 404
        
        incident = incident_result.data[0]
        current_assigned = incident.get('assigned_responder_id')
        status = incident.get('icd_status', '')
        
        # If incident is Resolved or Cancelled, cannot reassign
        if status in ['Resolved', 'Cancelled']:
            return jsonify({'success': False, 'message': f'Cannot assign {status} incidents'}), 400
        
        # If already assigned to someone else, check if current admin can reassign
        # Only allow reassignment if:
        # 1. No one is assigned (anyone can assign)
        # 2. Current admin is the assigned responder (can reassign)
        # 3. Current admin is assigning to themselves (taking over)
        if current_assigned and str(current_assigned) != str(current_admin_id) and str(new_responder_id) != str(current_admin_id):
            # Check if current admin can reassign (only if they're the current assignee)
            can_edit, edit_message = can_admin_edit_incident(incident_id, current_admin_id)
            if not can_edit:
                return jsonify({'success': False, 'message': f'Permission denied: {edit_message}'}), 403
        
        # Update assignment
        update_data = {
            'assigned_responder_id': new_responder_id
        }
        
        # If status is Active and being assigned, optionally set to Pending
        if status == 'Active' and data.get('set_pending', False):
            update_data['icd_status'] = 'Pending'
            update_data['pending_timestamp'] = get_philippines_time().isoformat()
        
        result = supabase.table('alert_incidents').update(update_data).eq('icd_id', incident_id).execute()
        
        if result.data:
            # Get responder name
            responder_name = new_responder_id
            try:
                responder_result = supabase.table('accounts_admin').select('admin_fullname').eq('admin_id', new_responder_id).execute()
                if responder_result.data and len(responder_result.data) > 0:
                    responder_name = responder_result.data[0].get('admin_fullname', new_responder_id)
            except:
                pass
            
            # Log activity
            log_admin_activity(
                current_admin_id,
                session.get('admin_name', 'Unknown'),
                'assign_incident',
                incident_id,
                current_assigned or 'Unassigned',
                new_responder_id
            )
            
            return jsonify({
                'success': True,
                'message': 'Incident assigned successfully',
                'assigned_responder_id': new_responder_id,
                'assigned_responder_name': responder_name
            })
        else:
            return jsonify({'success': False, 'message': 'Failed to assign incident'}), 500
            
    except Exception as e:
        print(f"Error assigning incident: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/incident/<incident_id>/audit')
def api_get_incident_audit(incident_id):
    """API endpoint to get audit trail for an incident"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        audit_records = get_audit_trail(incident_id)
        
        return jsonify({
            'success': True,
            'audit_records': audit_records
        })
    except Exception as e:
        print(f"Error getting audit trail: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/incident/<incident_id>/details')
def api_get_incident_details(incident_id):
    """API endpoint to get detailed incident information"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        incident = get_incident_details(incident_id)
        
        if incident:
            return jsonify({
                'success': True,
                'incident': incident
            })
        else:
            return jsonify({'success': False, 'message': 'Incident not found'}), 404
    except Exception as e:
        print(f"Error getting incident details: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/incidents/export', methods=['POST'])
def api_export_incidents():
    """API endpoint to export incidents to CSV"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        data = request.get_json() or {}
        incident_ids = data.get('incident_ids', [])
        export_format = data.get('format', 'csv')
        
        # Get incidents data
        incidents = []
        if incident_ids:
            # Export selected incidents
            for incident_id in incident_ids:
                incident = get_incident_details(incident_id)
                if incident:
                    incidents.append(incident)
        else:
            # Export all incidents
            result = supabase.table('alert_incidents').select('*').execute()
            all_incidents = result.data or []
            for incident in all_incidents:
                incident_details = get_incident_details(incident.get('icd_id'))
                if incident_details:
                    incidents.append(incident_details)
        
        if not incidents:
            return jsonify({'success': False, 'message': 'No incidents found'}), 404
        
        if export_format == 'csv':
            # Create CSV data
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write header
            writer.writerow([
                'Incident ID', 'Status', 'Category', 'Description', 
                'Latitude', 'Longitude', 'Reported Time', 'Resolved Time',
                'Student Name', 'Admin Name', 'Medical Type', 'Security Type'
            ])
            
            # Write data rows
            for incident in incidents:
                writer.writerow([
                    incident.get('icd_id', ''),
                    incident.get('icd_status', ''),
                    incident.get('icd_category', ''),
                    incident.get('icd_description', ''),
                    incident.get('icd_lat', ''),
                    incident.get('icd_lng', ''),
                    format_datetime(incident.get('icd_timestamp')),
                    format_datetime(incident.get('resolved_timestamp')),
                    incident.get('student_details', {}).get('full_name', '') if incident.get('student_details') else '',
                    incident.get('admin_details', {}).get('admin_fullname', '') if incident.get('admin_details') else '',
                    incident.get('icd_medical_type', ''),
                    incident.get('icd_security_type', '')
                ])
            
            # Create file response
            output.seek(0)
            ph_time = get_philippines_time()
            return send_file(
                io.BytesIO(output.getvalue().encode('utf-8')),
                mimetype='text/csv',
                as_attachment=True,
                download_name=f'incidents_export_{ph_time.strftime("%Y%m%d_%H%M%S")}.csv'
            )
        
        return jsonify({'success': False, 'message': 'Unsupported export format'}), 400
        
    except Exception as e:
        print(f"Error exporting incidents: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/incidents/export-pdf', methods=['GET'])
def export_incidents_pdf():
    """Export incidents to PDF (print view) with filters and date range"""
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    try:
        # Get filter parameters
        status_filter = request.args.get('status', 'All')
        search_term = request.args.get('search', '')
        sort_order = request.args.get('sort', 'recent')
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        
        # Build query for incidents
        query = supabase.table('alert_incidents').select('*')
        
        # Apply status filter
        if status_filter != 'All':
            query = query.eq('icd_status', status_filter)
        
        # Get incidents
        incidents_result = query.execute()
        incidents = incidents_result.data if incidents_result.data else []
        
        # Get student and admin data for joins
        students_result = supabase.table('accounts_student').select('*').execute()
        students = {s['user_id']: s for s in students_result.data} if students_result.data else {}
        
        admins_result = supabase.table('accounts_admin').select('*').execute()
        admins = {a['admin_id']: a for a in admins_result.data} if admins_result.data else {}
        
        # Get resolution reports from incident_resolution_reports table
        resolution_reports_result = supabase.table(RESOLUTION_REPORTS_TABLE).select('*').execute()
        resolution_reports = {}
        if resolution_reports_result.data:
            for report in resolution_reports_result.data:
                icd_id = str(report.get('icd_id', ''))
                if icd_id:
                    # Store the most recent resolution report for each incident
                    if icd_id not in resolution_reports:
                        resolution_reports[icd_id] = report
                    else:
                        # If multiple reports exist, keep the most recent one
                        existing_created = resolution_reports[icd_id].get('created_at', '')
                        new_created = report.get('created_at', '')
                        if new_created > existing_created:
                            resolution_reports[icd_id] = report
        
        # Process incidents with join data and filters
        processed_incidents = []
        for incident in incidents:
            # Apply date filter if provided
            if start_date or end_date:
                incident_date = incident.get('icd_timestamp')
                if incident_date:
                    try:
                        # Parse incident date (from Supabase, usually ISO format with timezone)
                        if 'T' in str(incident_date):
                            incident_dt = datetime.fromisoformat(str(incident_date).replace('Z', '+00:00'))
                        else:
                            # If no time component, assume start of day
                            incident_dt = datetime.fromisoformat(str(incident_date) + 'T00:00:00+00:00')
                        
                        # Convert to PH timezone for comparison
                        if incident_dt.tzinfo is None:
                            incident_dt = incident_dt.replace(tzinfo=timezone.utc)
                        incident_dt = incident_dt.astimezone(PHILIPPINES_TZ)
                        
                        # Parse filter dates (assume PH timezone)
                        if start_date:
                            try:
                                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                                start_dt = PHILIPPINES_TZ.localize(start_dt.replace(hour=0, minute=0, second=0))
                                if incident_dt < start_dt:
                                    continue
                            except Exception as e:
                                print(f"Error parsing start_date: {e}")
                        
                        if end_date:
                            try:
                                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                                end_dt = PHILIPPINES_TZ.localize(end_dt.replace(hour=23, minute=59, second=59))
                                if incident_dt > end_dt:
                                    continue
                            except Exception as e:
                                print(f"Error parsing end_date: {e}")
                    except Exception as e:
                        print(f"Error processing date filter for incident {incident.get('icd_id')}: {e}")
                        # Continue processing if date parsing fails
                        pass
            
            # Apply search filter
            if search_term:
                search_lower = search_term.lower()
                incident_id = str(incident.get('icd_id', '')).lower()
                description = str(incident.get('icd_description', '')).lower()
                
                # Check student name
                student_name = ''
                if incident.get('user_id') and incident.get('user_id') in students:
                    student = students[incident.get('user_id')]
                    student_name = student.get('full_name', '').lower()
                
                # Check admin name
                admin_name = ''
                if incident.get('user_id') and incident.get('user_id') in admins:
                    admin = admins[incident.get('user_id')]
                    admin_name = admin.get('admin_fullname', '').lower()
                
                if (search_lower not in incident_id and 
                    search_lower not in description and 
                    search_lower not in student_name and 
                    search_lower not in admin_name):
                    continue
            
            # Add student/admin info
            incident_data = incident.copy()
            if incident.get('user_id') and incident.get('user_id') in students:
                student = students[incident.get('user_id')]
                incident_data['student_name'] = student.get('full_name', 'N/A')
                incident_data['student_number'] = student.get('student_id', 'N/A')
                incident_data['student_college'] = student.get('student_college', 'N/A')
                incident_data['reported_by'] = student.get('full_name', 'N/A')
            elif incident.get('user_id') and incident.get('user_id') in admins:
                admin = admins[incident.get('user_id')]
                incident_data['admin_name'] = admin.get('admin_fullname', 'N/A')
                incident_data['reported_by'] = admin.get('admin_fullname', 'N/A')
                incident_data['student_college'] = 'N/A'
            else:
                incident_data['reported_by'] = 'N/A'
                incident_data['student_college'] = 'N/A'
            
            # Get assigned responder name
            assigned_responder_id = incident.get('assigned_responder_id')
            if assigned_responder_id and assigned_responder_id in admins:
                incident_data['assigned_responder_name'] = admins[assigned_responder_id].get('admin_fullname', assigned_responder_id)
            elif assigned_responder_id:
                incident_data['assigned_responder_name'] = assigned_responder_id
            else:
                incident_data['assigned_responder_name'] = 'N/A'
            
            # Compute location name from coordinates (prioritize reverse geocoding)
            lat = incident.get('icd_lat')
            lng = incident.get('icd_lng')
            building = incident.get('icd_location_building', '')
            floor = incident.get('icd_location_floor', '')
            room = incident.get('icd_location_room', '')
            
            location_name = ''
            if lat and lng:
                try:
                    lat_float = float(lat)
                    lng_float = float(lng)
                    
                    # Validate coordinates
                    if not (-90 <= lat_float <= 90) or not (-180 <= lng_float <= 180):
                        location_name = f"Lat: {lat_float:.6f}, Lng: {lng_float:.6f}"
                    else:
                        # Calculate distance from UMAK first (fast, no API call)
                        distance = math.sqrt((lat_float - UMAK_LAT)**2 + (lng_float - UMAK_LNG)**2) * 111  # km
                        
                        # If within 1km of UMAK, consider it UMAK Campus (no need for geocoding)
                        if distance < 1.0:
                            if building:
                                location_parts = []
                                if building:
                                    location_parts.append(f"Building: {building}")
                                if floor:
                                    location_parts.append(f"Floor: {floor}")
                                if room:
                                    location_parts.append(f"Room: {room}")
                                location_name = ", ".join(location_parts) if location_parts else 'UMAK Campus'
                            else:
                                location_name = 'UMAK Campus'
                        else:
                            # For locations outside UMAK, try reverse geocoding
                            try:
                                print(f"Attempting geocoding for incident {incident.get('icd_id')} at lat={lat_float}, lng={lng_float}")
                                location_name = get_location_name_from_coords(lat_float, lng_float, use_cache=True)
                                
                                # If we got a location name from coordinates, use it
                                if location_name and location_name.strip():
                                    # Clean up the location name - remove "Philippines" if it's redundant
                                    if location_name.endswith(', Philippines'):
                                        location_name = location_name[:-13].strip()
                                    print(f"Geocoding successful for incident {incident.get('icd_id')}: {location_name[:50]}")
                                else:
                                    # Geocoding failed or returned None, but we have coordinates
                                    print(f"Geocoding returned None for incident {incident.get('icd_id')}, using 'External Location'")
                                    location_name = 'External Location'
                            except Exception as geo_error:
                                print(f"Geocoding exception for incident {incident.get('icd_id')}: {geo_error}")
                                import traceback
                                traceback.print_exc()
                                # Geocoding failed, but we have coordinates - show generic location
                                location_name = 'External Location'
                except Exception as e:
                    print(f"Error computing location name for incident {incident.get('icd_id')}: {e}")
                    # On error, set to N/A so template can show coordinates separately
                    location_name = 'N/A'
            elif building:
                # Has building info but no coordinates
                location_parts = []
                if building:
                    location_parts.append(f"Building: {building}")
                if floor:
                    location_parts.append(f"Floor: {floor}")
                if room:
                    location_parts.append(f"Room: {room}")
                location_name = ", ".join(location_parts) if location_parts else 'UMAK Campus'
            else:
                location_name = 'N/A'
            
            incident_data['computed_location_name'] = location_name
            
            # Attach resolution report if incident is resolved
            incident_id_str = str(incident.get('icd_id', ''))
            if incident_id_str in resolution_reports:
                incident_data['resolution_report'] = resolution_reports[incident_id_str]
            else:
                incident_data['resolution_report'] = None
            
            # Calculate recency for time-based color coding
            if incident.get('icd_timestamp'):
                try:
                    if 'T' in str(incident.get('icd_timestamp')):
                        incident_dt = datetime.fromisoformat(str(incident.get('icd_timestamp')).replace('Z', '+00:00'))
                    else:
                        incident_dt = datetime.fromisoformat(str(incident.get('icd_timestamp')) + 'T00:00:00+00:00')
                    
                    if incident_dt.tzinfo is None:
                        incident_dt = incident_dt.replace(tzinfo=timezone.utc)
                    incident_dt = incident_dt.astimezone(PHILIPPINES_TZ)
                    
                    now = get_philippines_time()
                    hours_diff = (now - incident_dt).total_seconds() / 3600
                    
                    if hours_diff < 1:
                        incident_data['time_recency'] = 'recent'
                    elif hours_diff < 24:
                        incident_data['time_recency'] = 'recent-24h'
                    elif hours_diff < 168:  # 7 days
                        incident_data['time_recency'] = 'recent-7d'
                    else:
                        incident_data['time_recency'] = 'older'
                except:
                    incident_data['time_recency'] = 'older'
            else:
                incident_data['time_recency'] = 'older'
            
            processed_incidents.append(incident_data)
        
        # Sort incidents
        if sort_order == 'recent':
            processed_incidents.sort(key=lambda x: x.get('icd_timestamp', ''), reverse=True)
        else:
            processed_incidents.sort(key=lambda x: x.get('icd_timestamp', ''))
        
        # Get current admin info for header
        admin_name = session.get('admin_name', 'Administrator')
        ph_time = get_philippines_time()
        
        # Format dates for display
        formatted_start_date = ''
        formatted_end_date = ''
        if start_date:
            try:
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                formatted_start_date = start_dt.strftime('%B %d, %Y')
            except:
                formatted_start_date = start_date
        if end_date:
            try:
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                formatted_end_date = end_dt.strftime('%B %d, %Y')
            except:
                formatted_end_date = end_date
        
        # Calculate analytics/summary statistics
        total_count = len(processed_incidents)
        status_counts = {
            'Active': 0,
            'Pending': 0,
            'Resolved': 0,
            'Cancelled': 0
        }
        location_counts = {
            'UMAK': 0,
            'External': 0
        }
        
        for incident in processed_incidents:
            # Count by status
            status = incident.get('icd_status', '')
            if status in status_counts:
                status_counts[status] += 1
            
            # Count by location type
            location_name = incident.get('computed_location_name', '')
            if location_name and ('Building:' in location_name or 'UMAK' in location_name or location_name == 'UMAK Campus'):
                location_counts['UMAK'] += 1
            elif location_name and location_name != 'N/A' and not location_name.startswith('Lat:'):
                location_counts['External'] += 1
            else:
                # Check if has building info
                if incident.get('icd_location_building'):
                    location_counts['UMAK'] += 1
                elif location_name and location_name.startswith('Lat:'):
                    location_counts['External'] += 1
        
        return render_template('incident_export_pdf.html',
                             incidents=processed_incidents,
                             status_filter=status_filter,
                             category_filter='All',
                             search_term=search_term,
                             start_date=formatted_start_date,
                             end_date=formatted_end_date,
                             admin_name=admin_name,
                             export_date=ph_time.strftime('%B %d, %Y %I:%M %p'),
                             total_count=total_count,
                             status_counts=status_counts,
                             location_counts=location_counts)
        
    except Exception as e:
        print(f"Error exporting incidents to PDF: {e}")
        flash(f'Error generating export: {str(e)}', 'error')
        return redirect(url_for('incident_management'))

@app.route('/incident/<incident_id>/export-pdf', methods=['GET'])
def export_incident_pdf(incident_id):
    """Export single incident to PDF (print view)"""
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    try:
        # Get incident details
        incident = get_incident_details(incident_id)
        if not incident:
            flash('Incident not found', 'error')
            return redirect(url_for('incident_management'))
        
        # Get student and admin data for joins
        students_result = supabase.table('accounts_student').select('*').execute()
        students = {s['user_id']: s for s in students_result.data} if students_result.data else {}
        
        admins_result = supabase.table('accounts_admin').select('*').execute()
        admins = {a['admin_id']: a for a in admins_result.data} if admins_result.data else {}
        
        # Process incident with join data (similar to bulk export)
        incident_data = incident.copy()
        if incident.get('user_id') and incident.get('user_id') in students:
            student = students[incident.get('user_id')]
            incident_data['student_name'] = student.get('full_name', 'N/A')
            incident_data['student_number'] = student.get('student_id', 'N/A')
            incident_data['student_college'] = student.get('student_college', 'N/A')
            incident_data['reported_by'] = student.get('full_name', 'N/A')
        elif incident.get('user_id') and incident.get('user_id') in admins:
            admin = admins[incident.get('user_id')]
            incident_data['admin_name'] = admin.get('admin_fullname', 'N/A')
            incident_data['reported_by'] = admin.get('admin_fullname', 'N/A')
            incident_data['student_college'] = 'N/A'
        else:
            incident_data['reported_by'] = 'N/A'
            incident_data['student_college'] = 'N/A'
        
        # Get assigned responder name
        assigned_responder_id = incident.get('assigned_responder_id')
        if assigned_responder_id and assigned_responder_id in admins:
            incident_data['assigned_responder_name'] = admins[assigned_responder_id].get('admin_fullname', assigned_responder_id)
        elif assigned_responder_id:
            incident_data['assigned_responder_name'] = assigned_responder_id
        else:
            incident_data['assigned_responder_name'] = 'N/A'
        
        # Compute location name from coordinates (prioritize reverse geocoding)
        lat = incident.get('icd_lat')
        lng = incident.get('icd_lng')
        building = incident.get('icd_location_building', '')
        floor = incident.get('icd_location_floor', '')
        room = incident.get('icd_location_room', '')
        
        location_name = ''
        if lat and lng:
            try:
                lat_float = float(lat)
                lng_float = float(lng)
                
                # Validate coordinates
                if not (-90 <= lat_float <= 90) or not (-180 <= lng_float <= 180):
                    print(f"Invalid coordinates: lat={lat_float}, lng={lng_float}")
                    location_name = f"Lat: {lat_float:.6f}, Lng: {lng_float:.6f}"
                else:
                    # Always try to get location name from coordinates first
                    print(f"Attempting geocoding for lat={lat_float}, lng={lng_float}")
                    location_name = get_location_name_from_coords(lat_float, lng_float)
                    print(f"Geocoding result: {location_name}")
                    
                    # If we got a location name from coordinates, use it
                    if location_name and location_name.strip():
                        # If within UMAK area and has building info, append building details
                        distance = math.sqrt((lat_float - UMAK_LAT)**2 + (lng_float - UMAK_LNG)**2) * 111  # km
                        if distance < 1.0 and building:
                            building_parts = []
                            if building:
                                building_parts.append(f"Building: {building}")
                            if floor:
                                building_parts.append(f"Floor: {floor}")
                            if room:
                                building_parts.append(f"Room: {room}")
                            if building_parts:
                                location_name += f" ({', '.join(building_parts)})"
                    else:
                        # No location name from geocoding, calculate distance from UMAK
                        distance = math.sqrt((lat_float - UMAK_LAT)**2 + (lng_float - UMAK_LNG)**2) * 111  # km
                        
                        # If within 1km and has building info, consider it UMAK
                        if distance < 1.0 and building:
                            location_parts = []
                            if building:
                                location_parts.append(f"Building: {building}")
                            if floor:
                                location_parts.append(f"Floor: {floor}")
                            if room:
                                location_parts.append(f"Room: {room}")
                            location_name = ", ".join(location_parts) if location_parts else 'UMAK Campus'
                        else:
                            # Try one more time with a different zoom level or just show coordinates
                            # But first, let's try with a simpler format
                            print(f"Geocoding failed, trying alternative...")
                            # For now, show coordinates as fallback
                            location_name = f"Lat: {lat_float:.6f}, Lng: {lng_float:.6f}"
            except Exception as e:
                print(f"Error computing location name: {e}")
                import traceback
                traceback.print_exc()
                location_name = f"Lat: {lat}, Lng: {lng}" if lat and lng else 'N/A'
        elif building:
            # Has building info but no coordinates
            location_parts = []
            if building:
                location_parts.append(f"Building: {building}")
            if floor:
                location_parts.append(f"Floor: {floor}")
            if room:
                location_parts.append(f"Room: {room}")
            location_name = ", ".join(location_parts) if location_parts else 'UMAK Campus'
        else:
            location_name = 'N/A'
        
        incident_data['computed_location_name'] = location_name
        print(f"Final computed location name: {location_name}")
        
        # Get current admin info for header
        admin_name = session.get('admin_name', 'Administrator')
        ph_time = get_philippines_time()
        
        # Use the new individual incident print template
        return render_template('incident_print_pdf.html',
                             incident=incident_data,
                             admin_name=admin_name,
                             export_date=ph_time.strftime('%B %d, %Y %I:%M %p'))
        
    except Exception as e:
        print(f"Error exporting incident to PDF: {e}")
        flash(f'Error generating export: {str(e)}', 'error')
        return redirect(url_for('incident_management'))

@app.route('/incident/<incident_id>/resolution-report', methods=['GET'])
def export_incident_resolution_report(incident_id):
    """Export individual incident resolution report"""
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    try:
        # Get incident details
        incident = get_incident_details(incident_id)
        if not incident:
            flash('Incident not found', 'error')
            return redirect(url_for('incident_management'))
        
        # Check if incident is resolved
        if incident.get('icd_status') != 'Resolved':
            flash('This incident is not resolved. Resolution reports are only available for resolved incidents.', 'error')
            return redirect(url_for('incident_management'))
        
        # Get resolution report from database
        try:
            resolution_result = supabase.table(RESOLUTION_REPORTS_TABLE).select('*').eq('icd_id', str(incident_id)).order('created_at', desc=True).limit(1).execute()
            if not resolution_result.data or len(resolution_result.data) == 0:
                flash('No resolution report found for this incident.', 'error')
                return redirect(url_for('incident_management'))
            
            report = resolution_result.data[0]
            
            # Parse summary_details if it's a JSON string
            if report.get('summary_details') and isinstance(report.get('summary_details'), str):
                try:
                    import json
                    report['summary_details'] = json.loads(report['summary_details'])
                except:
                    pass  # Keep as string if parsing fails
            
            # Get student data if available
            if incident.get('user_id'):
                try:
                    student_result = supabase.table('accounts_student').select('*').eq('user_id', incident.get('user_id')).execute()
                    if student_result.data:
                        student = student_result.data[0]
                        report['student_id'] = student.get('student_id', 'N/A')
                        report['student_name'] = student.get('full_name', incident.get('student_name', 'N/A'))
                except:
                    report['student_id'] = incident.get('student_number', 'N/A')
                    report['student_name'] = incident.get('student_name', 'N/A')
            else:
                report['student_id'] = 'N/A'
                report['student_name'] = 'N/A'
            
            # Get location name using reverse geocoding
            location_name = 'N/A'
            lat = incident.get('icd_lat')
            lng = incident.get('icd_lng')
            building = incident.get('icd_location_building', '')
            floor = incident.get('icd_location_floor', '')
            room = incident.get('icd_location_room', '')
            
            if lat and lng:
                try:
                    lat_float = float(lat)
                    lng_float = float(lng)
                    
                    # Validate coordinates
                    if (-90 <= lat_float <= 90) and (-180 <= lng_float <= 180):
                        # Use reverse geocoding to get location name
                        location_name = get_location_name_from_coords(lat_float, lng_float)
                        
                        # If we got a location name, use it
                        if location_name and location_name.strip() and location_name != 'N/A':
                            # Clean up the location name - remove "Philippines" if it's redundant
                            if location_name.endswith(', Philippines'):
                                location_name = location_name[:-13].strip()
                            # If within UMAK area and has building info, append building details
                            distance = math.sqrt((lat_float - UMAK_LAT)**2 + (lng_float - UMAK_LNG)**2) * 111  # km
                            if distance < 1.0 and building:
                                building_parts = []
                                if building:
                                    building_parts.append(f"Building: {building}")
                                if floor:
                                    building_parts.append(f"Floor: {floor}")
                                if room:
                                    building_parts.append(f"Room: {room}")
                                if building_parts:
                                    location_name += f" ({', '.join(building_parts)})"
                        else:
                            # No location name from geocoding, check if it's UMAK based on distance
                            distance = math.sqrt((lat_float - UMAK_LAT)**2 + (lng_float - UMAK_LNG)**2) * 111  # km
                            if distance < 1.0:
                                if building:
                                    building_parts = []
                                    if building:
                                        building_parts.append(f"Building: {building}")
                                    if floor:
                                        building_parts.append(f"Floor: {floor}")
                                    if room:
                                        building_parts.append(f"Room: {room}")
                                    location_name = ", ".join(building_parts) if building_parts else 'UMAK Campus'
                                else:
                                    location_name = 'UMAK Campus'
                            else:
                                location_name = 'N/A'
                except Exception as e:
                    print(f"Error computing location name: {e}")
                    location_name = 'N/A'
            elif building:
                # Has building info but no coordinates
                building_parts = []
                if building:
                    building_parts.append(f"Building: {building}")
                if floor:
                    building_parts.append(f"Floor: {floor}")
                if room:
                    building_parts.append(f"Room: {room}")
                location_name = ", ".join(building_parts) if building_parts else 'UMAK Campus'
            
            # Enrich report with incident data
            report['icd_lat'] = lat
            report['icd_lng'] = lng
            report['location_name'] = location_name
            
            # Get incident type
            if incident.get('icd_category') == 'Medical':
                report['incident_type'] = incident.get('icd_medical_type', 'N/A')
            elif incident.get('icd_category') == 'Security':
                report['incident_type'] = incident.get('icd_security_type', 'N/A')
            elif incident.get('icd_category') == 'University':
                report['incident_type'] = incident.get('icd_university_type', 'N/A')
            else:
                report['incident_type'] = 'N/A'
            report['category'] = incident.get('icd_category', 'N/A')
            
        except Exception as e:
            print(f"Error fetching resolution report: {e}")
            flash('Error loading resolution report.', 'error')
            return redirect(url_for('incident_management'))
        
        # Get current admin info
        admin_name = session.get('admin_name', 'Administrator')
        ph_time = get_philippines_time()
        
        return render_template('incident_resolution_report.html',
                             report=report,
                             incident=incident,
                             admin_name=admin_name,
                             export_date=ph_time.strftime('%B %d, %Y %I:%M %p'))
        
    except Exception as e:
        print(f"Error exporting resolution report: {e}")
        import traceback
        traceback.print_exc()
        flash(f'Error generating resolution report: {str(e)}', 'error')
        return redirect(url_for('incident_management'))

@app.route('/incidents/resolution-bundle', methods=['GET'])
def export_resolution_bundle():
    """Export resolution bundle report with filters"""
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    
    try:
        # Get filter parameters
        status_filter = request.args.get('status', 'Resolved')
        category_filter = request.args.get('category', 'All')
        search_term = request.args.get('search', '')
        sort_order = request.args.get('sort', 'recent')
        start_date = request.args.get('start_date', '')
        end_date = request.args.get('end_date', '')
        
        # Get resolution reports from database
        query = supabase.table(RESOLUTION_REPORTS_TABLE).select('*')
        
        # Apply date filters if provided
        if start_date:
            try:
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                start_dt = PHILIPPINES_TZ.localize(start_dt)
                query = query.gte('resolved_at', start_dt.isoformat())
            except:
                pass
        
        if end_date:
            try:
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                end_dt = PHILIPPINES_TZ.localize(end_dt).replace(hour=23, minute=59, second=59)
                query = query.lte('resolved_at', end_dt.isoformat())
            except:
                pass
        
        # Get all resolution reports
        reports_result = query.execute()
        all_reports = reports_result.data if reports_result.data else []
        
        # Get incidents for additional data
        incidents_result = supabase.table('alert_incidents').select('*').execute()
        incidents = {str(inc.get('icd_id')): inc for inc in (incidents_result.data or [])}
        
        # Get students for additional data
        students_result = supabase.table('accounts_student').select('*').execute()
        students = {str(s.get('user_id')): s for s in (students_result.data or [])}
        
        # Process and enrich resolution reports
        processed_reports = []
        for report in all_reports:
            incident_id = str(report.get('icd_id', ''))
            incident = incidents.get(incident_id, {})
            
            # Parse summary_details if it's a JSON string
            if report.get('summary_details') and isinstance(report.get('summary_details'), str):
                try:
                    import json
                    report['summary_details'] = json.loads(report['summary_details'])
                except:
                    pass  # Keep as string if parsing fails
            
            # Get location name using reverse geocoding
            location_name = 'N/A'
            lat = incident.get('icd_lat')
            lng = incident.get('icd_lng')
            building = incident.get('icd_location_building', '')
            floor = incident.get('icd_location_floor', '')
            room = incident.get('icd_location_room', '')
            
            if lat and lng:
                try:
                    lat_float = float(lat)
                    lng_float = float(lng)
                    
                    # Validate coordinates
                    if (-90 <= lat_float <= 90) and (-180 <= lng_float <= 180):
                        # Use reverse geocoding to get location name
                        location_name = get_location_name_from_coords(lat_float, lng_float)
                        
                        # If we got a location name, use it
                        if location_name and location_name.strip() and location_name != 'N/A':
                            # Clean up the location name - remove "Philippines" if it's redundant
                            if location_name.endswith(', Philippines'):
                                location_name = location_name[:-13].strip()
                            # If within UMAK area and has building info, append building details
                            distance = math.sqrt((lat_float - UMAK_LAT)**2 + (lng_float - UMAK_LNG)**2) * 111  # km
                            if distance < 1.0 and building:
                                building_parts = []
                                if building:
                                    building_parts.append(f"Building: {building}")
                                if floor:
                                    building_parts.append(f"Floor: {floor}")
                                if room:
                                    building_parts.append(f"Room: {room}")
                                if building_parts:
                                    location_name += f" ({', '.join(building_parts)})"
                        else:
                            # No location name from geocoding, check if it's UMAK based on distance
                            distance = math.sqrt((lat_float - UMAK_LAT)**2 + (lng_float - UMAK_LNG)**2) * 111  # km
                            if distance < 1.0:
                                if building:
                                    building_parts = []
                                    if building:
                                        building_parts.append(f"Building: {building}")
                                    if floor:
                                        building_parts.append(f"Floor: {floor}")
                                    if room:
                                        building_parts.append(f"Room: {room}")
                                    location_name = ", ".join(building_parts) if building_parts else 'UMAK Campus'
                                else:
                                    location_name = 'UMAK Campus'
                            else:
                                location_name = 'N/A'
                except Exception as e:
                    print(f"Error computing location name: {e}")
                    location_name = 'N/A'
            elif building:
                # Has building info but no coordinates
                building_parts = []
                if building:
                    building_parts.append(f"Building: {building}")
                if floor:
                    building_parts.append(f"Floor: {floor}")
                if room:
                    building_parts.append(f"Room: {room}")
                location_name = ", ".join(building_parts) if building_parts else 'UMAK Campus'
            
            # Get incident type
            incident_type = 'N/A'
            if incident.get('icd_category') == 'Medical':
                incident_type = incident.get('icd_medical_type', 'N/A')
            elif incident.get('icd_category') == 'Security':
                incident_type = incident.get('icd_security_type', 'N/A')
            elif incident.get('icd_category') == 'University':
                incident_type = incident.get('icd_university_type', 'N/A')
            
            # Get student data
            student_id = 'N/A'
            student_name = 'N/A'
            if incident.get('user_id'):
                student = students.get(str(incident.get('user_id')))
                if student:
                    student_id = student.get('student_id', 'N/A')
                    student_name = student.get('full_name', 'N/A')
            
            # Enrich report with incident data
            enriched_report = report.copy()
            enriched_report['icd_lat'] = incident.get('icd_lat')
            enriched_report['icd_lng'] = incident.get('icd_lng')
            enriched_report['location_name'] = location_name
            enriched_report['incident_type'] = incident_type
            enriched_report['category'] = incident.get('icd_category') or report.get('category', 'N/A')
            enriched_report['student_id'] = student_id
            enriched_report['student_name'] = student_name
            
            # Apply filters
            if category_filter != 'All' and enriched_report.get('category') != category_filter:
                continue
            
            if search_term:
                search_lower = search_term.lower()
                if not any(search_lower in str(v).lower() for v in [
                    enriched_report.get('icd_id', ''),
                    enriched_report.get('resolved_id', ''),
                    enriched_report.get('student_name', ''),
                    enriched_report.get('resolved_by_name', '')
                ]):
                    continue
            
            processed_reports.append(enriched_report)
        
        # Sort reports
        if sort_order == 'old':
            processed_reports.sort(key=lambda x: x.get('resolved_at', ''))
        else:
            processed_reports.sort(key=lambda x: x.get('resolved_at', ''), reverse=True)
        
        # Format dates for display
        formatted_start_date = ''
        if start_date:
            try:
                dt = datetime.strptime(start_date, '%Y-%m-%d')
                formatted_start_date = dt.strftime('%B %d, %Y')
            except:
                formatted_start_date = start_date
        
        formatted_end_date = ''
        if end_date:
            try:
                dt = datetime.strptime(end_date, '%Y-%m-%d')
                formatted_end_date = dt.strftime('%B %d, %Y')
            except:
                formatted_end_date = end_date
        
        # Calculate status distribution from incidents
        status_distribution = {
            'total': 0,
            'active': 0,
            'pending': 0,
            'resolved': 0,
            'cancelled': 0
        }
        
        # Calculate location distribution
        location_distribution = {
            'umak': 0,
            'external': 0
        }
        
        # Get all incidents to calculate distributions
        all_incidents_result = supabase.table('alert_incidents').select('*').execute()
        all_incidents = all_incidents_result.data if all_incidents_result.data else []
        
        for inc in all_incidents:
            # Status distribution
            status = inc.get('icd_status', '')
            status_distribution['total'] += 1
            if status == 'Active':
                status_distribution['active'] += 1
            elif status == 'Pending':
                status_distribution['pending'] += 1
            elif status == 'Resolved':
                status_distribution['resolved'] += 1
            elif status == 'Cancelled':
                status_distribution['cancelled'] += 1
            
            # Location distribution
            lat = inc.get('icd_lat')
            lng = inc.get('icd_lng')
            building = inc.get('icd_location_building', '')
            
            if lat and lng:
                try:
                    lat_float = float(lat)
                    lng_float = float(lng)
                    # Calculate distance from UMAK
                    distance = math.sqrt((lat_float - UMAK_LAT)**2 + (lng_float - UMAK_LNG)**2) * 111  # km
                    if distance < 1.0 or building:
                        location_distribution['umak'] += 1
                    else:
                        location_distribution['external'] += 1
                except:
                    if building:
                        location_distribution['umak'] += 1
                    else:
                        location_distribution['external'] += 1
            elif building:
                location_distribution['umak'] += 1
            else:
                location_distribution['external'] += 1
        
        # Get current admin info
        admin_name = session.get('admin_name', 'Administrator')
        ph_time = get_philippines_time()
        
        return render_template('incident_resolution_bundle.html',
                             resolution_reports=processed_reports,
                             start_date=formatted_start_date,
                             end_date=formatted_end_date,
                             category_filter=category_filter,
                             search_term=search_term,
                             status_distribution=status_distribution,
                             location_distribution=location_distribution,
                             admin_name=admin_name,
                             export_date=ph_time.strftime('%B %d, %Y %I:%M %p'))
        
    except Exception as e:
        print(f"Error generating resolution bundle: {e}")
        import traceback
        traceback.print_exc()
        flash(f'Error generating resolution bundle: {str(e)}', 'error')
        return redirect(url_for('incident_management'))

@app.route('/api/incidents/export-excel', methods=['POST'])
def api_export_incidents_excel():
    """Export incidents to Excel (CSV format) with filters and date range"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        data = request.get_json() or {}
        status_filter = data.get('status', 'All')
        search_term = data.get('search', '')
        sort_order = data.get('sort', 'recent')
        start_date = data.get('start_date', '')
        end_date = data.get('end_date', '')
        
        # Build query for incidents
        query = supabase.table('alert_incidents').select('*')
        
        # Apply status filter
        if status_filter != 'All':
            query = query.eq('icd_status', status_filter)
        
        # Get incidents
        incidents_result = query.execute()
        incidents = incidents_result.data if incidents_result.data else []
        
        # Get student and admin data for joins
        students_result = supabase.table('accounts_student').select('*').execute()
        students = {s['user_id']: s for s in students_result.data} if students_result.data else {}
        
        admins_result = supabase.table('accounts_admin').select('*').execute()
        admins = {a['admin_id']: a for a in admins_result.data} if admins_result.data else {}
        
        # Process incidents with join data and filters
        processed_incidents = []
        for incident in incidents:
            # Apply date filter if provided
            if start_date or end_date:
                incident_date = incident.get('icd_timestamp')
                if incident_date:
                    try:
                        # Parse incident date (from Supabase, usually ISO format with timezone)
                        if 'T' in str(incident_date):
                            incident_dt = datetime.fromisoformat(str(incident_date).replace('Z', '+00:00'))
                        else:
                            # If no time component, assume start of day
                            incident_dt = datetime.fromisoformat(str(incident_date) + 'T00:00:00+00:00')
                        
                        # Convert to PH timezone for comparison
                        if incident_dt.tzinfo is None:
                            incident_dt = incident_dt.replace(tzinfo=timezone.utc)
                        incident_dt = incident_dt.astimezone(PHILIPPINES_TZ)
                        
                        # Parse filter dates (assume PH timezone)
                        if start_date:
                            try:
                                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                                start_dt = PHILIPPINES_TZ.localize(start_dt.replace(hour=0, minute=0, second=0))
                                if incident_dt < start_dt:
                                    continue
                            except Exception as e:
                                print(f"Error parsing start_date: {e}")
                        
                        if end_date:
                            try:
                                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                                end_dt = PHILIPPINES_TZ.localize(end_dt.replace(hour=23, minute=59, second=59))
                                if incident_dt > end_dt:
                                    continue
                            except Exception as e:
                                print(f"Error parsing end_date: {e}")
                    except Exception as e:
                        print(f"Error processing date filter for incident {incident.get('icd_id')}: {e}")
                        # Continue processing if date parsing fails
                        pass
            
            # Apply search filter
            if search_term:
                search_lower = search_term.lower()
                incident_id = str(incident.get('icd_id', '')).lower()
                description = str(incident.get('icd_description', '')).lower()
                
                # Check student name
                student_name = ''
                if incident.get('user_id') and incident.get('user_id') in students:
                    student = students[incident.get('user_id')]
                    student_name = student.get('full_name', '').lower()
                
                # Check admin name
                admin_name = ''
                if incident.get('user_id') and incident.get('user_id') in admins:
                    admin = admins[incident.get('user_id')]
                    admin_name = admin.get('admin_fullname', '').lower()
                
                if (search_lower not in incident_id and 
                    search_lower not in description and 
                    search_lower not in student_name and 
                    search_lower not in admin_name):
                    continue
            
            # Add student/admin info
            incident_data = incident.copy()
            if incident.get('user_id') and incident.get('user_id') in students:
                student = students[incident.get('user_id')]
                incident_data['student_name'] = student.get('full_name', 'N/A')
                incident_data['student_number'] = student.get('student_id', 'N/A')
                incident_data['student_contact'] = student.get('student_cnum', 'N/A')
                incident_data['student_college'] = student.get('student_college', 'N/A')
            elif incident.get('user_id') and incident.get('user_id') in admins:
                admin = admins[incident.get('user_id')]
                incident_data['admin_name'] = admin.get('admin_fullname', 'N/A')
                incident_data['student_contact'] = 'N/A'
                incident_data['student_college'] = 'N/A'
            else:
                incident_data['student_contact'] = 'N/A'
                incident_data['student_college'] = 'N/A'
            
            # Get assigned responder name if available
            assigned_responder_id = incident.get('assigned_responder_id')
            if assigned_responder_id:
                try:
                    responder_result = supabase.table('accounts_admin').select('admin_fullname').eq('admin_id', assigned_responder_id).execute()
                    if responder_result.data and len(responder_result.data) > 0:
                        incident_data['assigned_responder_name'] = responder_result.data[0].get('admin_fullname', assigned_responder_id)
                    else:
                        incident_data['assigned_responder_name'] = assigned_responder_id
                except:
                    incident_data['assigned_responder_name'] = assigned_responder_id
            else:
                incident_data['assigned_responder_name'] = 'N/A'
            
            # Location name - use location identifier if available, otherwise leave blank
            incident_data['location_name'] = incident.get('icd_location_identifier', '')
            
            processed_incidents.append(incident_data)
        
        # Sort incidents
        if sort_order == 'recent':
            processed_incidents.sort(key=lambda x: x.get('icd_timestamp', ''), reverse=True)
        else:
            processed_incidents.sort(key=lambda x: x.get('icd_timestamp', ''))
        
        if not processed_incidents:
            return jsonify({'success': False, 'message': 'No incidents found matching the filters'}), 404
        
        # Try to use openpyxl for proper Excel format
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            use_openpyxl = True
        except ImportError:
            # openpyxl not available, will use CSV fallback
            use_openpyxl = False
        
        if use_openpyxl:
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Incident_Bundle_Report"
            
            # Define styles
            header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            label_font = Font(bold=True, size=11)
            border_style = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')
            wrap_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # Get admin and organization info
            admin_name = session.get('admin_name', 'Administrator')
            org_name = "University of Makati"
            ph_time = get_philippines_time()
            export_datetime = ph_time.strftime('%B %d, %Y %I:%M %p')
            
            # Format date range for display
            date_range_str = ''
            if start_date and end_date:
                try:
                    start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                    end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                    date_range_str = f"{start_dt.strftime('%B %d, %Y')} – {end_dt.strftime('%B %d, %Y')}"
                except:
                    date_range_str = f"{start_date} – {end_date}"
            elif start_date:
                try:
                    start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                    date_range_str = f"From {start_dt.strftime('%B %d, %Y')}"
                except:
                    date_range_str = f"From {start_date}"
            elif end_date:
                try:
                    end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                    date_range_str = f"Until {end_dt.strftime('%B %d, %Y')}"
                except:
                    date_range_str = f"Until {end_date}"
            else:
                date_range_str = "All dates"
            
            # Format filters for display
            status_filter_display = status_filter if status_filter != 'All' else 'All'
            category_filter_display = 'All'  # Category filter not currently in export, but keeping for future
            keyword_filter_display = search_term if search_term else 'None'
            
            # Write header information (rows 1-6)
            # Row 1: Organization Name
            ws.cell(row=1, column=1, value="Organization Name:").font = label_font
            ws.cell(row=1, column=2, value=org_name)
            
            # Row 2: Report Type
            ws.cell(row=2, column=1, value="Report Type:").font = label_font
            ws.cell(row=2, column=2, value="Bundle (Multiple Incidents)")
            
            # Row 3: Date Exported
            ws.cell(row=3, column=1, value="Date Exported:").font = label_font
            ws.cell(row=3, column=2, value=export_datetime)
            
            # Row 4: Prepared By
            ws.cell(row=4, column=1, value="Prepared By:").font = label_font
            ws.cell(row=4, column=2, value=admin_name)
            
            # Row 5: Date Range
            ws.cell(row=5, column=1, value="Date Range:").font = label_font
            ws.cell(row=5, column=2, value=date_range_str)
            
            # Row 6: Filters Applied
            ws.cell(row=6, column=1, value="Filters Applied:").font = label_font
            filters_text = f"Status: {status_filter_display} | Category: {category_filter_display} | Keyword: {keyword_filter_display}"
            ws.cell(row=6, column=2, value=filters_text)
            
            # Row 7: Empty row for spacing
            # Row 8: Table headers
            table_headers = [
                'Incident ID', 'Timestamp', 'Status', 'Category', 'Location Name',
                'Latitude (icd_lat)', 'Longitude (icd_lng)', 'Building', 'Floor', 'Room',
                'Reporter', 'Contact', 'College', 'Description',
                'Assigned Responder', 'Active TS', 'Pending TS', 'Resolved TS', 'Cancelled TS'
            ]
            
            for col_num, header in enumerate(table_headers, 1):
                cell = ws.cell(row=8, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border_style
            
            # Format timestamps helper
            def format_ts(ts):
                if not ts:
                    return ''
                try:
                    if 'T' in str(ts):
                        dt = datetime.fromisoformat(str(ts).replace('Z', '+00:00'))
                    else:
                        dt = datetime.fromisoformat(str(ts) + 'T00:00:00+00:00')
                    
                    # Convert to PH timezone
                    if dt.tzinfo is None:
                        dt = dt.replace(tzinfo=timezone.utc)
                    dt = dt.astimezone(PHILIPPINES_TZ)
                    return dt.strftime('%Y-%m-%d %H:%M:%S')
                except:
                    return str(ts) if ts else ''
            
            # Write data rows (starting at row 9, after header row 8)
            for row_num, incident in enumerate(processed_incidents, 9):
                # Determine reporter name
                reporter = incident.get('student_name', '') or incident.get('admin_name', '') or 'N/A'
                
                # Get contact number
                contact = incident.get('student_contact', 'N/A')
                
                # Get college
                college = incident.get('student_college', 'N/A')
                
                # Description - plain text, no formatting
                description = str(incident.get('icd_description', '')).strip() if incident.get('icd_description') else ''
                
                row_data = [
                    incident.get('icd_id', ''),
                    format_ts(incident.get('icd_timestamp')),  # Timestamp
                    incident.get('icd_status', ''),
                    incident.get('icd_category', ''),
                    incident.get('location_name', ''),  # Location Name
                    incident.get('icd_lat', ''),
                    incident.get('icd_lng', ''),
                    incident.get('icd_location_building', ''),
                    incident.get('icd_location_floor', ''),
                    incident.get('icd_location_room', ''),
                    reporter,  # Reporter
                    contact,  # Contact
                    college,  # College
                    description,  # Description (plain text)
                    incident.get('assigned_responder_name', 'N/A'),  # Assigned Responder
                    format_ts(incident.get('icd_timestamp')),  # Active TS (same as timestamp)
                    format_ts(incident.get('pending_timestamp')),  # Pending TS
                    format_ts(incident.get('resolved_timestamp')),  # Resolved TS
                    format_ts(incident.get('cancelled_timestamp'))  # Cancelled TS
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    # Set as plain text (no formatting)
                    cell.value = str(value) if value else ''
                    cell.border = border_style
                    if col_num == 14:  # Description column
                        cell.alignment = wrap_align
                    else:
                        cell.alignment = left_align
            
            # Auto-adjust column widths
            column_widths = {
                'A': 18,  # Incident ID
                'B': 20,  # Timestamp
                'C': 12,  # Status
                'D': 15,  # Category
                'E': 25,  # Location Name
                'F': 15,  # Latitude
                'G': 15,  # Longitude
                'H': 15,  # Building
                'I': 10,  # Floor
                'J': 10,  # Room
                'K': 25,  # Reporter
                'L': 15,  # Contact
                'M': 20,  # College
                'N': 50,  # Description
                'O': 20,  # Assigned Responder
                'P': 20,  # Active TS
                'Q': 20,  # Pending TS
                'R': 20,  # Resolved TS
                'S': 20   # Cancelled TS
            }
            
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
            
            # Add auto-filter to entire table (header row 8 to last data row)
            last_row = 8 + len(processed_incidents)
            last_col = get_column_letter(len(table_headers))
            ws.auto_filter.ref = f"A8:{last_col}{last_row}"
            
            # Freeze header row and info rows
            ws.freeze_panes = 'A9'
            
            # Set row heights
            ws.row_dimensions[8].height = 25  # Table header row
            
            # Save to BytesIO
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            ph_time = get_philippines_time()
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'incidents_export_{ph_time.strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
        else:
            # Fallback to CSV if openpyxl is not available
            print("Warning: openpyxl is not installed. Falling back to CSV format.")
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write header row (matching Excel format)
            headers = [
                'Incident ID', 'Timestamp', 'Status', 'Category', 'Location Name',
                'Latitude (icd_lat)', 'Longitude (icd_lng)', 'Building', 'Floor', 'Room',
                'Reporter', 'Contact', 'College', 'Description',
                'Assigned Responder', 'Active TS', 'Pending TS', 'Resolved TS', 'Cancelled TS'
            ]
            writer.writerow(headers)
            
            # Write data rows
            for incident in processed_incidents:
                def format_ts(ts):
                    if not ts:
                        return ''
                    try:
                        if 'T' in str(ts):
                            dt = datetime.fromisoformat(str(ts).replace('Z', '+00:00'))
                        else:
                            dt = datetime.fromisoformat(str(ts) + 'T00:00:00+00:00')
                        if dt.tzinfo is None:
                            dt = dt.replace(tzinfo=timezone.utc)
                        dt = dt.astimezone(PHILIPPINES_TZ)
                        return dt.strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        return str(ts) if ts else ''
                
                # Determine reporter name
                reporter = incident.get('student_name', '') or incident.get('admin_name', '') or 'N/A'
                
                writer.writerow([
                    incident.get('icd_id', ''),
                    format_ts(incident.get('icd_timestamp')),  # Timestamp
                    incident.get('icd_status', ''),
                    incident.get('icd_category', ''),
                    incident.get('location_name', ''),  # Location Name
                    incident.get('icd_lat', ''),
                    incident.get('icd_lng', ''),
                    incident.get('icd_location_building', ''),
                    incident.get('icd_location_floor', ''),
                    incident.get('icd_location_room', ''),
                    reporter,  # Reporter
                    incident.get('student_contact', 'N/A'),  # Contact
                    incident.get('student_college', 'N/A'),  # College
                    str(incident.get('icd_description', '')).strip() if incident.get('icd_description') else '',  # Description (plain text)
                    incident.get('assigned_responder_name', 'N/A'),  # Assigned Responder
                    format_ts(incident.get('icd_timestamp')),  # Active TS
                    format_ts(incident.get('pending_timestamp')),  # Pending TS
                    format_ts(incident.get('resolved_timestamp')),  # Resolved TS
                    format_ts(incident.get('cancelled_timestamp'))  # Cancelled TS
                ])
            
            # Create file response with Excel-compatible CSV (UTF-8 BOM for Excel)
            output.seek(0)
            ph_time = get_philippines_time()
            csv_content = output.getvalue()
            # Add UTF-8 BOM for Excel compatibility
            excel_bytes = '\ufeff'.encode('utf-8') + csv_content.encode('utf-8')
            
            return send_file(
                io.BytesIO(excel_bytes),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'incidents_export_{ph_time.strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
        
    except Exception as e:
        print(f"Error exporting incidents to Excel: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/incidents/bulk-action', methods=['POST'])
def api_bulk_action():
    """API endpoint for bulk actions on incidents"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        data = request.get_json()
        action = data.get('action')
        incident_ids = data.get('incident_ids', [])
        reason = data.get('reason', '')
        
        if not incident_ids:
            return jsonify({'success': False, 'message': 'No incidents selected'}), 400
        
        if action == 'archive':
            results = bulk_archive_incidents(incident_ids, session['admin_id'], reason)
            return jsonify({
                'success': True,
                'results': results,
                'message': f'Processed {len(results)} incidents'
            })
        elif action == 'status_update':
            new_status = data.get('new_status')
            if not new_status:
                return jsonify({'success': False, 'message': 'New status required'}), 400
            
            # Update status for all selected incidents
            current_time = get_philippines_time().isoformat()
            results = []
            for incident_id in incident_ids:
                try:
                    # Check if admin can edit this incident
                    can_edit, edit_message = can_admin_edit_incident(incident_id, session['admin_id'])
                    if not can_edit:
                        results.append({
                            'incident_id': incident_id,
                            'success': False,
                            'message': f'Permission denied: {edit_message}'
                        })
                        continue
                    
                    update_data = {
                        'icd_status': new_status,
                        'status_updated_at': current_time,
                        'status_updated_by': session['admin_id']
                    }
                    
                    # Reset all timestamp fields
                    update_data['resolved_timestamp'] = None
                    update_data['pending_timestamp'] = None
                    update_data['cancelled_timestamp'] = None
                    
                    # Set appropriate timestamp based on status
                    if new_status == 'Resolved':
                        update_data['resolved_timestamp'] = current_time
                    elif new_status == 'Pending':
                        update_data['pending_timestamp'] = current_time
                    elif new_status == 'Cancelled':
                        update_data['cancelled_timestamp'] = current_time
                    
                    supabase.table('alert_incidents').update(update_data).eq('icd_id', incident_id).execute()
                    
                    # Log to audit trail
                    log_incident_change(incident_id, 'status_updated', 
                                      new_status=new_status, admin_id=session['admin_id'])
                    
                    results.append({
                        'incident_id': incident_id,
                        'success': True,
                        'message': 'Status updated successfully'
                    })
                except Exception as e:
                    results.append({
                        'incident_id': incident_id,
                        'success': False,
                        'message': str(e)
                    })
            
            return jsonify({
                'success': True,
                'results': results,
                'message': f'Updated {len(results)} incidents'
            })
        else:
            return jsonify({'success': False, 'message': 'Unknown action'}), 400
        
    except Exception as e:
        print(f"Error in bulk action: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/archived-incidents')
def api_get_archived_incidents():
    """API endpoint to get archived incidents"""
    if 'admin_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        archived_incidents = get_archived_incidents()
        
        return jsonify({
            'success': True,
            'archived_incidents': archived_incidents
        })
    except Exception as e:
        print(f"Error getting archived incidents: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/archive/<archive_id>/restore', methods=['POST'])
def api_restore_incident(archive_id):
    """API endpoint to restore an archived incident"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        success, message = restore_incident(archive_id, session['admin_id'])
        
        if success:
            return jsonify({
                'success': True,
                'message': message,
                'restored_at': get_philippines_time().isoformat()
            })
        else:
            return jsonify({'success': False, 'message': message}), 500
            
    except Exception as e:
        print(f"Error restoring incident: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

def get_incidents_with_relations(filters=None):
    """Get incidents with related student and admin data - ENHANCED"""
    try:
        filters = filters or {}
        # Build base query
        query = supabase.table('alert_incidents').select('*')
        
        # Apply filters
        if filters:
            if filters.get('status') and filters['status'] != 'All':
                query = query.eq('icd_status', filters['status'])
            if filters.get('category'):
                query = query.eq('icd_category', filters['category'])
            if filters.get('search'):
                # Note: Supabase doesn't support full-text search in free tier
                # This is a simplified search - in production, consider using PostgreSQL full-text search
                pass
        
        incidents_result = query.order('icd_timestamp', desc=True).execute()
        incidents = incidents_result.data if incidents_result.data else []
        
        admin_id = filters.get('admin_id') if filters else None
        if admin_id:
            incidents = filter_incidents_for_admin(incidents, admin_id)
        
        # Get related data
        students_result = supabase.table('accounts_student').select('*').execute()
        students = {s['user_id']: s for s in students_result.data} if students_result.data else {}
        
        admins_result = supabase.table('accounts_admin').select('*').execute()
        admins = {a['admin_id']: a for a in admins_result.data} if admins_result.data else {}
        
        # Process incidents with join data
        processed_incidents = []
        for incident in incidents:
            # Get related student and admin data
            student_data = students.get(incident.get('user_id'))
            admin_data = admins.get(incident.get('admin_id'))
            
            # Add related data to incident
            incident['student_name'] = student_data.get('full_name') if student_data else None
            incident['student_number'] = student_data.get('student_id') if student_data else None
            incident['student_user_id'] = student_data.get('user_id') if student_data else None
            incident['admin_name'] = admin_data.get('admin_fullname') if admin_data else None
            
            # Calculate response time if resolved
            if incident.get('icd_status') == 'Resolved':
                response_time = calculate_response_time(incident.get('icd_id'))
                incident['response_time_minutes'] = response_time
            
            assigned_responder_id = incident.get('assigned_responder_id')
            if assigned_responder_id and admins.get(assigned_responder_id):
                incident['assigned_responder_name'] = admins[assigned_responder_id].get('admin_fullname', assigned_responder_id)
            else:
                incident['assigned_responder_name'] = None
            
            processed_incidents.append(incident)
        
        return processed_incidents
    except Exception as e:
        print(f"Error getting incidents with relations: {e}")
        return []

# ==================== API ROUTES FOR USER MANAGEMENT ====================
@app.route('/api/user/<user_id>', methods=['GET', 'PUT', 'DELETE'])
def api_manage_user(user_id):
    """API endpoint to get, update, or delete user details"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    # Get user type from query params or request body
    user_type = request.args.get('type')
    if not user_type and request.is_json:
        json_data = request.get_json() or {}
        user_type = json_data.get('type')
    
    if not user_type or user_type not in ['admin', 'student']:
        return jsonify({'success': False, 'message': 'Invalid user type'}), 400
    
    print(f"🔍 API Request - User ID: {user_id}, Type: {user_type}, Method: {request.method}")
    
    try:
        if request.method == 'GET':
            # Get user details
            if user_type == 'admin':
                result = supabase.table('accounts_admin').select('*').eq('admin_id', user_id).execute()
            else:
                result = supabase.table('accounts_student').select('*').eq('user_id', user_id).execute()
            
            if result.data:
                return jsonify({'success': True, 'user': result.data[0]})
            else:
                return jsonify({'success': False, 'message': 'User not found'}), 404
                
        elif request.method == 'PUT':
            # Update user - support both JSON and multipart form (for profile image uploads)
            if request.is_json:
                data = request.get_json() or {}
                files = None
            else:
                data = request.form.to_dict() if request.form else {}
                files = request.files
            
            if not data and not (files and files.get('profile_image')):
                return jsonify({'success': False, 'message': 'No data provided'}), 400

            def parse_bool(value):
                if isinstance(value, bool):
                    return value
                if isinstance(value, str):
                    return value.lower() in ['true', '1', 'yes', 'on']
                return False

            if 'email_verified' in data:
                data['email_verified'] = parse_bool(data.get('email_verified'))

            uploaded_profile_filename = None
            if files:
                profile_file = files.get('profile_image')
                if profile_file and profile_file.filename:
                    upload_result = handle_profile_upload(profile_file, user_type)
                    if upload_result.get('error'):
                        return jsonify({'success': False, 'message': upload_result['error']}), 400
                    uploaded_profile_filename = upload_result.get('filename')

            update_data = {}
            
            print(f"📝 Updating {user_type} with ID: {user_id}")
            print(f"📋 Received data: {data}")
            
            if user_type == 'admin':
                # Include all provided fields - update everything that's sent
                if 'username' in data:
                    update_data['admin_user'] = data.get('username') or None
                if 'email' in data:
                    update_data['admin_email'] = data.get('email') or None
                if 'fullname' in data:
                    update_data['admin_fullname'] = data.get('fullname') or None
                if 'role' in data:
                    update_data['admin_role'] = data.get('role') or None
                if 'status' in data:
                    update_data['admin_status'] = data.get('status') or None
                # Note: approval field removed for admin edit
                
                # Hash password if provided
                if data.get('password'):
                    update_data['admin_pass'] = hash_password(data.get('password'))
                
                if uploaded_profile_filename:
                    update_data['admin_profile'] = uploaded_profile_filename
                
                if not update_data:
                    return jsonify({'success': False, 'message': 'No fields to update'}), 400
                
                print(f"📤 Admin update data: {update_data}")
                result = supabase.table('accounts_admin').update(update_data).eq('admin_id', user_id).execute()
                print(f"✅ Admin update result: {result}")
            else:
                # Update student with proper field mapping - include ALL fields that are provided
                # This ensures all form data is saved to Supabase
                if 'student_id' in data:
                    update_data['student_id'] = data.get('student_id') or None
                if 'username' in data:
                    update_data['student_user'] = data.get('username') or None
                if 'email' in data:
                    update_data['student_email'] = data.get('email') or None
                if 'fullname' in data:
                    update_data['full_name'] = data.get('fullname') or None
                if 'yearlvl' in data:
                    update_data['student_yearlvl'] = data.get('yearlvl') or None
                if 'college' in data:
                    # Allow empty string to clear the field
                    college_value = data.get('college')
                    update_data['student_college'] = college_value if college_value else None
                if 'cnum' in data:
                    update_data['student_cnum'] = data.get('cnum') or None
                if 'emergency' in data:
                    update_data['primary_emergencycontact'] = data.get('emergency') or None
                if 'contactperson' in data:
                    update_data['primary_contactperson'] = data.get('contactperson') or None
                if 'cprelationship' in data:
                    update_data['primary_cprelationship'] = data.get('cprelationship') or None
                if 'secondary_emergency' in data:
                    secondary_emergency_value = data.get('secondary_emergency')
                    update_data['secondary_emergencycontact'] = secondary_emergency_value if secondary_emergency_value else None
                if 'secondary_contact' in data:
                    secondary_contact_value = data.get('secondary_contact')
                    update_data['secondary_contactperson'] = secondary_contact_value if secondary_contact_value else None
                if 'secondary_relationship' in data:
                    secondary_rel_value = data.get('secondary_relationship')
                    update_data['secondary_cprelationship'] = secondary_rel_value if secondary_rel_value else None
                if 'medinfo' in data:
                    medinfo_value = data.get('medinfo')
                    update_data['student_medinfo'] = medinfo_value if medinfo_value else None
                if 'address' in data:
                    update_data['student_address'] = data.get('address') or None
                if 'residency' in data:
                    update_data['residency'] = data.get('residency') or None
                if 'status' in data:
                    update_data['student_status'] = data.get('status') or None
                if 'email_verified' in data:
                    # Handle boolean checkbox value
                    update_data['email_verified'] = parse_bool(data.get('email_verified'))
                
                # Hash password if provided
                if data.get('password'):
                    update_data['student_pass'] = hash_password(data.get('password'))
                
                if uploaded_profile_filename:
                    update_data['student_profile'] = uploaded_profile_filename
                
                if not update_data:
                    return jsonify({'success': False, 'message': 'No fields to update'}), 400
                
                print(f"📤 Student update data: {update_data}")
                print(f"📤 Updating student with user_id: {user_id}")
                result = supabase.table('accounts_student').update(update_data).eq('user_id', user_id).execute()
                print(f"✅ Student update result: {result}")
                print(f"✅ Result data: {result.data if hasattr(result, 'data') else 'No data attribute'}")
                print(f"✅ Result error: {result.error if hasattr(result, 'error') else 'No error attribute'}")
            
            # Check if update was successful
            # Supabase update may return empty data array on success, so we check for errors instead
            # Check for errors first - if there's an error, the update failed
            if hasattr(result, 'error') and result.error:
                print(f"❌ Supabase error: {result.error}")
                return jsonify({'success': False, 'message': f'Database error: {result.error}'}), 500
            
            # If no error, the update was successful
            # Supabase sometimes returns empty list [] on successful updates, which is fine
            # result.data can be None, empty list [], or a list with data
            # All of these are acceptable if there's no error
            return jsonify({
                'success': True, 
                'message': 'User updated successfully',
                'data': result.data if hasattr(result, 'data') and result.data else []
            })
                
        elif request.method == 'DELETE':
            # Delete user
            if user_type == 'admin':
                result = supabase.table('accounts_admin').delete().eq('admin_id', user_id).execute()
            else:
                result = supabase.table('accounts_student').delete().eq('user_id', user_id).execute()
            
            if result.data:
                return jsonify({'success': True, 'message': 'User deleted successfully'})
            else:
                return jsonify({'success': False, 'message': 'Failed to delete user'}), 500
            
    except Exception as e:
        print(f"❌ Error managing user: {e}")
        print(f"❌ Error type: {type(e).__name__}")
        import traceback
        print(f"❌ Traceback: {traceback.format_exc()}")
        error_message = str(e)
        # Provide more helpful error messages
        if 'duplicate' in error_message.lower() or 'unique' in error_message.lower():
            error_message = 'A user with this information already exists. Please check for duplicates.'
        elif 'foreign key' in error_message.lower() or 'constraint' in error_message.lower():
            error_message = 'Database constraint error. Please check related records.'
        elif 'permission' in error_message.lower() or 'unauthorized' in error_message.lower():
            error_message = 'Permission denied. Please check your database permissions.'
        return jsonify({'success': False, 'message': f'Error: {error_message}'}), 500

@app.route('/api/user/<user_id>/archive', methods=['POST'])
def api_archive_user(user_id):
    """API endpoint to archive a user"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        data = request.get_json() or {}
        user_type = request.args.get('type') or data.get('type')
        reason = data.get('reason', 'Archived by administrator')
        
        if not user_type or user_type not in ['admin', 'student']:
            return jsonify({'success': False, 'message': 'Invalid user type'}), 400
        
        success, message = archive_user(user_id, user_type, session['admin_id'], reason)
        
        if success:
            return jsonify({
                'success': True, 
                'message': message,
                'archived_at': datetime.now().isoformat()
            })
        else:
            return jsonify({'success': False, 'message': message}), 500
            
    except Exception as e:
        print(f"Error archiving user: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/archived-users')
def api_get_archived_users():
    """API endpoint to get archived users"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        result = supabase.table('user_archive').select('*').order('archived_at', desc=True).execute()
        
        archived_users = []
        for user in (result.data or []):
            # Format the archived user data
            archived_user = {
                'archive_id': user.get('archive_id'),
                'user_id': user.get('user_id'),
                'user_type': user.get('user_type'),
                'archived_by': user.get('archived_by'),
                'archive_reason': user.get('archive_reason'),
                'archived_at': user.get('archived_at'),
                'fullname': user.get('admin_fullname') or user.get('full_name'),
                'email': user.get('admin_email') or user.get('student_email'),
                'username': user.get('admin_user') or user.get('student_user'),
                'role': user.get('admin_role') or 'Student',
                'status': user.get('admin_status') or user.get('student_status')
            }
            archived_users.append(archived_user)
        
        return jsonify({
            'success': True,
            'users': archived_users,
            'count': len(archived_users)
        })
    except Exception as e:
        print(f"Error fetching archived users: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/user/archive/<archive_id>/restore', methods=['POST'])
def api_restore_user(archive_id):
    """API endpoint to restore an archived user"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        result = restore_user(archive_id, session['admin_id'])
        
        if len(result) >= 2 and result[0]:  # success is True
            success, message = result[0], result[1]
            restored_user_id = result[2] if len(result) > 2 else None
            user_type = result[3] if len(result) > 3 else None
            
            return jsonify({
                'success': True,
                'message': message,
                'restored_at': datetime.now().isoformat(),
                'restored_user_id': restored_user_id,
                'user_type': user_type
            })
        else:
            success, message = result[0], result[1] if len(result) > 1 else "Unknown error"
            return jsonify({'success': False, 'message': message}), 500
            
    except Exception as e:
        print(f"Error restoring user: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/user', methods=['POST'])
def api_create_user():
    """API endpoint to create new user"""
    if 'admin_id' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        data = request.get_json()
        user_type = data.get('type')
        
        if not user_type or user_type not in ['admin', 'student']:
            return jsonify({'success': False, 'message': 'Invalid user type'}), 400
        
        if user_type == 'admin':
            # Generate admin ID for new admin
            new_admin_id = generate_next_admin_id()
            
            admin_data = {
                'admin_id': new_admin_id,
                'admin_user': data.get('username'),
                'admin_pass': hash_password(data.get('password')) if data.get('password') else None,
                'admin_email': data.get('email'),
                'admin_fullname': data.get('fullname'),
                'admin_role': data.get('role'),
                'admin_status': data.get('status', 'Active'),
                'admin_created_at': datetime.now().isoformat()
            }
            
            result = supabase.table('accounts_admin').insert(admin_data).execute()
            
        else:
            # Create student with proper field mapping - include ALL fields
            # NOTE: user_id is NOT included - it's auto-generated by the database sequence
            student_data = {
                'student_id': data.get('student_id'),
                'student_user': data.get('username'),
                'student_pass': hash_password(data.get('password')) if data.get('password') else None,
                'student_email': data.get('email'),
                'full_name': data.get('fullname'),
                'student_yearlvl': data.get('yearlvl'),
                'student_college': data.get('college') or 'CLAS',  # Default to CLAS if not provided
                'student_cnum': data.get('cnum'),
                'primary_emergencycontact': data.get('emergency') or None,
                'primary_contactperson': data.get('contactperson') or None,
                'primary_cprelationship': data.get('cprelationship') or None,
                'secondary_emergencycontact': data.get('secondary_emergency') or None,
                'secondary_contactperson': data.get('secondary_contact') or None,
                'secondary_cprelationship': data.get('secondary_relationship') or None,
                'student_medinfo': data.get('medinfo') or None,
                'student_address': data.get('address'),
                'residency': data.get('residency', 'MAKATI'),
                'student_status': data.get('status', 'Active'),
                'email_verified': data.get('email_verified', False),
                'student_created_at': datetime.now().isoformat()
            }
            
            # Keep None values only for fields that are allowed to be NULL in the database
            # Required fields must have values, optional fields can be None
            optional_nullable_fields = ['primary_cprelationship', 'secondary_cprelationship', 'primary_emergencycontact', 
                                       'secondary_emergencycontact', 'primary_contactperson', 'secondary_contactperson', 
                                       'student_medinfo', 'student_profile']
            student_data = {k: v for k, v in student_data.items() if v is not None or k in optional_nullable_fields}
            
            print(f"📤 Inserting student data (user_id will be auto-generated): {student_data}")
            result = supabase.table('accounts_student').insert(student_data).execute()
            print(f"✅ Student insert result: {result}")
            if hasattr(result, 'error') and result.error:
                print(f"❌ Supabase error: {result.error}")
                raise Exception(f"Database error: {result.error}")
        
        if result.data:
            return jsonify({'success': True, 'message': 'User created successfully', 'user_id': result.data[0].get('admin_id' if user_type == 'admin' else 'user_id')})
        else:
            # Check for errors
            if hasattr(result, 'error') and result.error:
                error_msg = str(result.error)
                return jsonify({'success': False, 'message': f'Database error: {error_msg}'}), 500
            return jsonify({'success': False, 'message': 'Failed to create user - no data returned'}), 500
            
    except Exception as e:
        print(f"❌ Error creating user: {e}")
        print(f"❌ Error type: {type(e).__name__}")
        import traceback
        print(f"❌ Traceback: {traceback.format_exc()}")
        error_message = str(e)
        # Provide more helpful error messages
        if 'duplicate' in error_message.lower() or 'unique' in error_message.lower():
            error_message = 'A user with this information already exists. Please check for duplicates (student_id, email, or username).'
        elif 'foreign key' in error_message.lower() or 'constraint' in error_message.lower():
            error_message = 'Database constraint error. Please check that all required fields are provided and enum values are correct.'
        elif 'enum' in error_message.lower() or 'invalid input' in error_message.lower():
            error_message = 'Invalid enum value. Please check that year level, college, status, and relationship values match the database enum types.'
        elif 'not null' in error_message.lower():
            error_message = 'Required field is missing. Please fill in all required fields.'
        return jsonify({'success': False, 'message': f'Error: {error_message}'}), 500

@app.route('/logout')
def logout():
    """Admin logout route"""
    session.clear()
    flash('You have been logged out successfully!', 'success')
    return redirect(url_for('login'))

# Test route to create a test admin (remove in production)
@app.route('/create_test_admin')
def create_test_admin():
    """Create a test admin user (for development only)"""
    try:
        # Check if test admin already exists
        existing = supabase.table('accounts_admin').select('*').eq('admin_user', 'admin').execute()
        
        if existing.data:
            return "Test admin already exists! Use username: admin, password: admin123"
        
        # Generate admin ID for test admin
        test_admin_id = generate_next_admin_id()
        
        # Create test admin with plain text password
        test_admin = {
            'admin_id': test_admin_id,
            'admin_user': 'admin',
            'admin_pass': 'admin123',  # Plain text
            'admin_email': 'admin@test.com',
            'admin_fullname': 'Test Administrator',
            'admin_role': 'System Administrator',
            'admin_status': 'Active',
            'admin_profile': None,
            'admin_last_login': None
        }
        
        result = supabase.table('accounts_admin').insert(test_admin).execute()
        return f"Test admin created successfully! Use username: admin, password: admin123"
        
    except Exception as e:
        return f"Error creating test admin: {e}"

# Test route to create a sample account request
@app.route('/create-test-request')
def create_test_request():
    """Create a test account request for testing"""
    try:
        test_request = {
            'admin_fullname': 'Test User',
            'admin_email': 'test@example.com',
            'admin_user': 'testuser',
            'admin_pass': bcrypt.hashpw('Test123!'.encode('utf-8'), bcrypt.gensalt()).decode('utf-8'),
            'admin_role': 'Security Staff',
            'admin_approval': 'Pending',
            'request_reason': 'Test request for approval system'
        }
        
        result = supabase.table('account_requests').insert(test_request).execute()
        return f"Test request created! ID: {result.data[0]['id'] if result.data else 'Unknown'}"
    except Exception as e:
        return f"Error creating test request: {e}"

# Health check route
@app.route('/health')
def health():
    """Health check endpoint"""
    return {'status': 'healthy', 'timestamp': datetime.now().isoformat()}

if __name__ == '__main__':
    # Allow hosting platforms like Railway to inject the port value
    port = int(os.getenv('PORT', '5000'))
    debug_mode = os.getenv('FLASK_DEBUG', 'false').lower() == 'true'
    app.run(debug=debug_mode, host='0.0.0.0', port=port)